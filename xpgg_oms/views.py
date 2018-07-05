#!/usr/bin/env python3
# -.- coding:utf-8 -.-

from django.shortcuts import render, redirect, HttpResponse
import json
from django.http import JsonResponse, HttpResponseRedirect, FileResponse  # 1.7以后版本json数据返回方法
import re
import os
import time
from .forms import *
from .models import *
from . import cron
from .scripts import netscantool
from django.conf import settings
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


from django.views.generic import TemplateView, ListView, DetailView
from django.core.paginator import Paginator
from django.contrib.auth import logout, login, authenticate  # 登陆注销注册django自带模块引用
from django.contrib.auth.hashers import make_password  # django自带密码加密模块
import ast  # 去掉字符串的一层""
import openpyxl  # 操作excel读写
from io import BytesIO
from django.utils.encoding import escape_uri_path  # 下载文件中文名时使用
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone  # 调用django的时间参数timezone.now()


import logging
# Create your views here.
logger = logging.getLogger('xpgg_oms.views')


def global_setting(request):
    return {
            'SITE_NAME': settings.SITE_NAME,
            'SITE_DESC': settings.SITE_DESC,
            'SITE_URL': settings.SITE_URL}


# 封装salt-api的调用
class SaltAPI(object):
    # 给获取token的几个参数设置默认值，这样如果真突然出现调用另一个salt-api的情况，直接把对应参数传入一样适用哈哈
    # 第一个参数session是为了给with as传入使用的，因为用with as会在程序执行完成后回收资源，不然Session是长连接占着连接不知道会不会造成影响后期
    def __init__(self, session, apiurl=settings.SITE_SALT_API_URL, username=settings.SITE_SALT_API_NAME, password=settings.SITE_SALT_API_PWD, eauth='pam'):
        self.url = apiurl
        self.session = session
        self.username = username
        self.password = password
        self.eauth = eauth

    # 获取token
    def get_token(self):
        count = 2
        connect_test = 1
        while count:
            try:
                # 初始化获取api的token
                token = self.session.post(self.url+'/login', json={'username': self.username, 'password': self.password, 'eauth': self.eauth, }, timeout=10)
                token.raise_for_status()
            except Exception as e:
                response_data = '第%s次尝试链接saltapi获取token失败:' % connect_test
                logger.error(response_data + str(e))
                count -= 1
                connect_test += 1
                time.sleep(4)
                continue
            else:
                return True
        #     比如当api服务忘了开就会false
        else:
            return False

    # 先做一个最通用的方法，就是不定义data的各个东西，在使用的时候定义好带入，好处是任何一个saltapi的操作都能支持，而且可以单独使用
    def public(self, data, message='public'):
        count = 2
        connect_test = 1
        while count:
            try:
                response_data = self.session.post(self.url, data=data)
                response_data.raise_for_status()
            except Exception as e:
                response_data = '第%s次尝试SaltAPI调用%s请求出错' % (connect_test, message)
                logger.error(response_data + str(e))
                count -= 1
                connect_test += 1
                time.sleep(4)
                continue
            else:
                # 正确执行后返回值一般有这几种情况：
                # 1、返回需要的值，字典key是return值是[{xxx}]如{'return': [{'192.168.100.171': True}]}
                # 2、返回{'return': [{}]}，第一种没有这个minion_id，第二种是在使用salt-run jobs.lookup_jid任务结果还未返回时候也会如此
                # 出现这种情况说实话不好做判断，所以最好的办法是minion表要多实时刷新保持最新，从源头避免minion不存在的可能
                # 3、{'return': [{'192.168.100.170': False}]} 说明有这个minion_id但是连不上有可能停止了反正就是通不了哈
                # 同样的这种情况最好的办法不是在代码做判断，也是保持minion表中minion状态的实时在线离线，源头避免
                # 所以如果调用的时候有这种情况需要做下判断
                return response_data.json()
        else:
            return False

    # 封装test.ping,默认执行salt '*' test.ping
    def test_api(self,  client='local', tgt='*', tgt_type='glob', fun='test.ping', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'test_api'
        return self.public(data, message)

    # 封装cmd.run,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def cmd_run_api(self, client='local', tgt='*', tgt_type='glob', fun='cmd.run', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'cmd_run_api'
        return self.public(data, message)

    # 封装异步cmd.run,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def async_cmd_run_api(self, client='local_async', tgt='*', tgt_type='glob', fun='cmd.run', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'async_cmd_run_api'
        return self.public(data, message)

    # 封装state.sls,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def state_api(self, client='local', tgt='*', tgt_type='glob', fun='state.sls', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'state_api'
        return self.public(data, message)

    # 封装异步state.sls,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入，得到结果为jid号
    def async_state_api(self, client='local_async', tgt='*', tgt_type='glob', fun='state.sls', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'async_state_api'
        return self.public(data, message)

    # 封装通过jid查询任务执行状态，以便后续操作，返回[{}]表示执行完毕，返回数据表示还在执行
    def job_active_api(self, tgt, arg, tgt_type='glob'):
        data = {'client': 'local', 'tgt': tgt, 'tgt_type': tgt_type, 'fun': 'saltutil.find_job', 'arg': arg}
        message = 'job_active_api'
        return self.public(data, message)

    # 封装查询jid执行状态,使用的时候只要代入jid既可以，返回true表示执行结束并且成功退出，false表示没有成功或者还没执行完毕
    def job_exit_success_api(self, client='runner', fun='jobs.exit_success', jid=None):
        data = {'client': client,
                'fun': fun,
                'jid': jid,
                }
        message = 'job_exit_success_api'
        return self.public(data, message)

    # 封装查询jid结果方法,使用的时候只要代入jid既可以
    def jid_api(self, client='runner', fun='jobs.lookup_jid', jid=None):
        data = {'client': client,
                'fun': fun,
                'jid': jid,
                }
        message = 'jid_api'
        return self.public(data, message)

    # 封装archive.zip,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def archive_zip_api(self, client='local', tgt='*', tgt_type='glob', fun='archive.zip', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg
                }
        message = 'archive_zip_api'
        return self.public(data, message)

    # 封装archive.tar,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def archive_tar_api(self, client='local', tgt='*', tgt_type='glob', fun='archive.tar', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg
                }
        message = 'archive_tar_api'
        return self.public(data, message)

    # 封装cp.get_file,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def cp_get_file_api(self, client='local', tgt='*', tgt_type='glob', fun='cp.get_file', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'cp_get_file_api'
        return self.public(data, message)

    # 封装cp.get_dir,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def cp_get_dir_api(self, client='local', tgt='*', tgt_type='glob', fun='cp.get_dir', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'cp_get_dir_api'
        return self.public(data, message)

    # 封装salt-key -L
    def saltkey_listall_api(self, client='wheel',  fun='key.list_all'):
        data = {'client': client,
                'fun': fun,
                }
        message = 'saltkey_listall_api'
        return self.public(data, message)

    # 封装salt-key -d删除指令,有点坑就是发现这个返回结果只要api请求正常返回结果都是这样的：
    # {'return': [{'tag': 'salt/wheel/20170914183529100717', 'data': {'tag': 'salt/wheel/20170914183529100717',
    #  'success': True, 'fun': 'wheel.key.delete', 'user': 'saltapi', 'jid': '20170914183529100717',
    #  '_stamp': '2017-09-14T10:35:29.106380', 'return': {}}}]} 看到没return值为{}空无语，所以简单点就是判断success是true
    # 说明api请求返回成功了
    def saltkey_delete_api(self, client='wheel', fun='key.delete', match=None):
        data = {'client': client,
                'fun': fun,
                'match': match
                }
        message = 'saltkey_delete_api'
        return self.public(data, message)

    # 接受salt-key的方法奶奶的include_rejected和include_denied就算设置为True也无效测试发现！！
    def saltkey_accept_api(self, client='wheel', fun='key.accept', match=None, include_rejected=False, include_denied=False):
        data = {'client': client,
                'fun': fun,
                'match': match,
                'include_rejected': include_rejected,
                'include_denied': include_denied
                }
        message = 'saltkey_accept_api'
        return self.public(data, message)

    # 拒绝salt-key的方法奶奶的include_accepted和include_denied就算设置为True也无效测试发现！！
    def saltkey_reject_api(self, client='wheel', fun='key.reject', match=None, include_accepted=False, include_denied=False):
        data = {'client': client,
                'fun': fun,
                'match': match,
                'include_accepted': include_accepted,
                'include_denied': include_denied
                }
        message = 'saltkey_reject_api'
        return self.public(data, message)

    # salt-run manage.status 查看minion在线离线状态，速度比较慢但是没BUG不像salt-run manage.alived
    def saltrun_manage_status_api(self, client='runner', fun='manage.status', arg=None):
        data = {'client': client,
                'fun': fun,
                'arg': arg,
                }
        message = 'saltrun_manage_status_api'
        return self.public(data, message)

    # salt-run manage.alived 查看在线的minion，非常快速方便可惜有bug后来启用(而且可以带参数show_ipv4=True获取到和master通信的ip是什么，默认False)
    def saltrun_manage_alive_api(self, client='runner', fun='manage.alived', arg=None):
        data = {'client': client,
                'fun': fun,
                'arg': arg,
                }
        message = 'saltrun_manage_alive_api'
        return self.public(data, message)

    # salt-run manage.not_alived 查看不在线的minion，非常快速方便可惜有bug后来启用
    def saltrun_manage_notalive_api(self, client='runner', fun='manage.not_alived', arg=None):
        data = {'client': client,
                'fun': fun,
                'arg': arg,
                }
        message = 'saltrun_manage_notalive_api'
        return self.public(data, message)

    # 封装grains.itmes,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def grains_itmes_api(self, client='local', tgt='*', tgt_type='glob', fun='grains.items', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'grains_itmes_api'
        return self.public(data, message)

    # 封装service.available查看服务是否存在Ture or False,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def service_available_api(self, client='local', tgt='*', tgt_type='glob', fun='service.available', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'service_available_api'
        return self.public(data, message)

    # 封装service.status查看启动服务状态,使用的时候只要代入tgt和arg即可，最多把tgt_type也代入
    def service_status_api(self, client='local', tgt='*', tgt_type='glob', fun='service.status', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'service_status_api'
        return self.public(data, message)

    # 封装service.start启动系统服务windows和linux通用，
    def service_start_api(self, client='local', tgt='*', tgt_type='glob', fun='service.start', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'service_start_api'
        return self.public(data, message)

    # 封装service.stop停止系统服务windows和linux通用，salt '*' service.stop <service name>，由于有发现停止成功但是返回结果是
    # 一堆错误提示，所以最好使用的时候最后做一步service.status，返回服务状态True说明启动，False说明停止了
    def service_stop_api(self, client='local', tgt='*', tgt_type='glob', fun='service.stop', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'service_stop_api'
        return self.public(data, message)

    # 封装ps.pgrep查看name的进程号windows和linux通用带模糊查询效果，
    def ps_pgrep_api(self, client='local', tgt='*', tgt_type='glob', fun='ps.pgrep', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'ps_pgrep_api'
        return self.public(data, message)

    # 封装ps.proc_info通过进程号查看详细信息，
    # {'client':'local', 'tgt':'id','fun':'ps.proc_info', 'arg':['pid=123','attrs=["cmdline","pid","name","status"]']}
    def ps_proc_info_api(self, client='local', tgt='*', tgt_type='glob', fun='ps.proc_info', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'ps_proc_info_api'
        return self.public(data, message)

    # 封装ps.kill_pid结束某个进程，{'client':'local','fun':'ps.kill_pid','tgt':'192.168.68.1', 'arg':['pid=11932']}
    def ps_kill_pid_api(self, client='local', tgt='*', tgt_type='glob', fun='ps.kill_pid', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'ps_kill_pid_api'
        return self.public(data, message)

    # 封装task.create_task创建windows计划任务，salt '192.168.68.1' task.create_task ooxx  action_type=Execute
    # cmd='"C:\ooxx\Shadowsocks.exe"' force=true execution_time_limit=False  user_name=administrator
    def task_create_api(self, client='local', tgt='*', tgt_type='glob', fun='task.create_task', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'task_create_api'
        return self.public(data, message)

    # 封装task.run启动windows计划任务，salt '192.168.100.171' task.run test1
    # ！坑！官方文档里命令是salt '192.168.100.171' task.list_run test1 根本就不行！
    def task_run_api(self, client='local', tgt='*', tgt_type='glob', fun='task.run', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'task_run_api'
        return self.public(data, message)

    # 封装task.stop启动windows计划任务，salt '192.168.100.171' task.run test1
    def task_stop_api(self, client='local', tgt='*', tgt_type='glob', fun='task.stop', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'task_stop_api'
        return self.public(data, message)

    # 封装file.mkdir,创建目录最后可以不需要/号，另一个file.makedirs则需要最后/，不然只创建到有/那一层这点也是可以利用的呵呵
    def file_mkdir_api(self, client='local', tgt='*', tgt_type='glob', fun='file.mkdir', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'file_mkdir_api'
        return self.public(data, message)

    # 封装file.makedirs,创建目录最后可以需要/号，不然只创建到有/那一层这点也是可以利用的呵呵
    def file_makedirs_api(self, client='local', tgt='*', tgt_type='glob', fun='file.makedirs', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'file_makedirs_api'
        return self.public(data, message)

    # 封装file_exists,检查文件是否存在
    def file_exists_api(self, client='local', tgt='*', tgt_type='glob', fun='file.file_exists', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'file_exists_api'
        return self.public(data, message)

    # 封装file_exists,检查文件是否存在
    def file_write_api(self, client='local', tgt='*', tgt_type='glob', fun='file.write', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'file_write_api'
        return self.public(data, message)

    # 封装file.remove,移除文件，如果是目录则递归删除
    def file_remove_api(self, client='local', tgt='*', tgt_type='glob', fun='file.remove', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'file_remove_api'
        return self.public(data, message)

    # 封装file.directory_exists,检测目录是否存在返回True/False
    def file_directory_exists_api(self, client='local', tgt='*', tgt_type='glob', fun='file.directory_exists', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'file_directory_exists_api'
        return self.public(data, message)

    # 封装file.symlink,创建软连接
    def file_symlink_api(self, client='local', tgt='*', tgt_type='glob', fun='file.symlink', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'file_symlink_api'
        return self.public(data, message)

    # 封装supervisord.status,检测supervisor守护名称的状态，返回结果有这几种情况：
    # 1、没这个命令，salt需要安装supervisor才能用{'return': [{'192.168.100.171': "'supervisord.status' is not available."}]}
    # 2、没这个守护名称{'return': [{'192.168.68.50-master': {'1:': {'reason': '(no such process)', 'state': 'ERROR'}}}]}
    # 3、安装了supervisor但是没启动{'return': [{'192.168.100.170': {'unix:///var/run/supervisor/supervisor.sock': {'reason'
    # : 'such file', 'state': 'no'}}}]}
    # 4、正常获取结果的情况：
    # 启动{'return': [{'192.168.68.50-master': {'djangoproject.runserver': {'state': 'RUNNING', 'reason': 'pid 1233,
    #  uptime 1 day, 6:56:14'}}}]}
    # 停止{'return': [{'192.168.100.170': {'test': {'state': 'STOPPED', 'reason': 'Dec 13 05:23 PM'}}}]}
    def supervisord_status_api(self, client='local', tgt='*', tgt_type='glob', fun='supervisord.status', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'supervisord_status_api'
        return self.public(data, message)

    # 封装supervisord.stop,停止supervisor守护名称，返回结果除了上面status的前面3中情况还有以下几种情况：
    # 1、程序已经停止情况：{'return': [{'192.168.100.170': 'test: ERROR (not running)'}]}
    # 2、正常停止：{'return': [{'192.168.100.170': 'test: stopped'}]}
    # 3、没这个程序名称{'return': [{'192.168.100.170': 'test1: ERROR (no such process)'}]}
    def supervisord_stop_api(self, client='local', tgt='*', tgt_type='glob', fun='supervisord.stop', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'supervisord_stop_api'
        return self.public(data, message)

    # 封装supervisord.start,启动supervisor守护名称，返回结果有这几种情况：
    # 1、正常启动：{'return': [{'192.168.100.170': 'test: started'}]}
    # 2、已经启动过了{'return': [{'192.168.100.170': 'test: ERROR (already started)'}]}
    # 3、没这个程序名称{'return': [{'192.168.100.170': 'test1: ERROR (no such process)'}]}
    def supervisord_start_api(self, client='local', tgt='*', tgt_type='glob', fun='supervisord.start', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'supervisord_start_api'
        return self.public(data, message)

    # supervisord配置重载会启动新添加的程序
    def supervisord_update_api(self, client='local', tgt='*', tgt_type='glob', fun='supervisord.update', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'supervisord_update_api'
        return self.public(data, message)

    # 封装rsync.rsync同步命令
    def rsync_rsync_api(self, client='local', tgt='*', tgt_type='glob', fun='rsync.rsync', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'rsync_rsync_api'
        return self.public(data, message)

    # 封装异步rsync.rsync同步命令
    def async_rsync_rsync_api(self, client='local_async', tgt='*', tgt_type='glob', fun='rsync.rsync', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'async_rsync_rsync_api'
        return self.public(data, message)

    # 封装sys.doc查询模块帮助命令
    def sys_doc_api(self, client='local', tgt='*', tgt_type='glob', fun='sys.doc', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'sys_doc_api'
        return self.public(data, message)

    # 封装sys.doc查询模块帮助命令
    def sys_runner_doc_api(self, client='local', tgt='*', tgt_type='glob', fun='sys.runner_doc', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'sys_runner_doc_api'
        return self.public(data, message)

    # 封装sys.doc查询模块帮助命令
    def sys_state_doc_api(self, client='local', tgt='*', tgt_type='glob', fun='sys.state_doc', arg=None):
        data = {'client': client,
                'tgt': tgt,
                'tgt_type': tgt_type,
                'fun': fun,
                'arg': arg,
                }
        message = 'sys_state_doc_api'
        return self.public(data, message)


# 分页代码，第一个参数要request是因为里头代码要request.GET东西需要有request支持
def getPage(request, data_list, page_num=10):
    # 传2参数，一个是要分页的列表或者queryset，一个是每页显示数量默认10
    paginator = Paginator(data_list, page_num)  # import引入的django自带分页模块Paginator，data_list是数据库查询后的queryset，每页10条记录
    try:
        page = int(request.GET.get('page', 1))  # 从页面上的?page获取值，看html里分页设置了这个值,如果没有就赋值1，第一页
        data_list = paginator.page(page)
    except Exception:
        data_list = paginator.page(1)
    return data_list


# 登录页
def do_login(request):
    try:
        if request.method == 'GET':
            return render(request, 'login.html')
        elif request.is_ajax():
            username = request.POST.get('username')
            password = request.POST.get('password')
            next_url = request.POST.get('next')
            user = authenticate(username=username, password=password)  # 调用django自带验证模块key是Myuser的
            if user is not None and user.is_active:  # user如果验证成功返回user对象失败返回None，is_active是判断用户是否激活状态的，自带的用户注册默认状态是true激活
                login(request, user)
                if next_url:
                    return JsonResponse({'result': next_url, 'status': True})

                else:
                    return JsonResponse({'result': '/', 'status': True})
            else:
                return JsonResponse({'status': False})
            # else:
            #     #设置一个变量login_err来把错误内容输出到源界面，而不是用下面注释掉的跳转错误页
            #     return render(request, 'login.html', {'login_err': login_form.errors})
            #     #return render(request,'failure.html',{'reason':login_form.errors})

    except Exception as e:
        logger.error(e)
        return render(request, 'login.html')


# 退出
def do_logout(request):
    try:
        logout(request)
    except Exception as e:
        logger.error(e)
    return redirect(settings.LOGIN_URL)#返回到登录页


# 首页仪表盘
def index(request):
    # physical_pc_count = ServerList.objects.filter(server_type='0').count()
    # virtual_machine = ServerList.objects.filter(server_type='1').count()
    # sys_type_windows_count = ServerList.objects.filter(sys_type='windows').count()
    # sys_type_linux_count = ServerList.objects.filter(sys_type='linux').count()
    # saltkey_accepted_count = SaltKeyList.objects.filter(certification_status='accepted').count()
    # saltkey_denied_count = SaltKeyList.objects.filter(certification_status='denied').count()
    # saltkey_rejected_count = SaltKeyList.objects.filter(certification_status='rejected').count()
    # saltkey_unaccepted_count = SaltKeyList.objects.filter(certification_status='unaccepted').count()
    # minion_up_count = MinionList.objects.filter(minion_status='在线').count()
    # minion_down_count = MinionList.objects.filter(minion_status='离线').count()
    # minion_error_count = MinionList.objects.filter(minion_status='异常').count()
    # data = {'physical_pc_count': physical_pc_count, 'virtual_machine': virtual_machine}
    return render(request, 'index.html', locals())


# 首页仪表盘ajax
def index_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.GET.get('index_tag_key') == 'index_get_data':
                result['result'] = {}
                result['result']['physical_pc_count'] = ServerList.objects.filter(server_type='0').count()
                result['result']['virtual_machine'] = ServerList.objects.filter(server_type='1').count()
                result['result']['sys_type_windows_count'] = ServerList.objects.filter(sys_type='windows').count()
                result['result']['sys_type_linux_count'] = ServerList.objects.filter(sys_type='linux').count()
                result['result']['saltkey_accepted_count'] = SaltKeyList.objects.filter(certification_status='accepted').count()
                result['result']['saltkey_denied_count'] = SaltKeyList.objects.filter(certification_status='denied').count()
                result['result']['saltkey_rejected_count'] = SaltKeyList.objects.filter(certification_status='rejected').count()
                result['result']['saltkey_unaccepted_count'] = SaltKeyList.objects.filter(certification_status='unaccepted').count()
                result['result']['minion_up_count'] = MinionList.objects.filter(minion_status='在线').count()
                result['result']['minion_down_count'] = MinionList.objects.filter(minion_status='离线').count()
                result['result']['minion_error_count'] = MinionList.objects.filter(minion_status='异常').count()
                result['result']['minion_windows_count'] = MinionList.objects.filter(sys='Windows').count()
                result['result']['minion_linux_count'] = MinionList.objects.filter(sys='Linux').count()
                result['status'] = True
                # 返回字典之外的需要把参数safe改成false如：JsonResponse([1, 2, 3], safe=False)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'add_server_list':
                obj = ServerListAddForm(request.POST)
                if obj.is_valid():
                    ServerList.objects.create(server_name=obj.cleaned_data["server_name"],
                                              server_type=obj.cleaned_data["server_type"],
                                              localhost=obj.cleaned_data["localhost"],
                                              ip=obj.cleaned_data["ip"],
                                              system_issue=obj.cleaned_data["system_issue"],
                                              sn=obj.cleaned_data["sn"],
                                              cpu_num=obj.cleaned_data["cpu_num"],
                                              cpu_model=obj.cleaned_data["cpu_model"],
                                              sys_type=obj.cleaned_data["sys_type"],
                                              kernel=obj.cleaned_data["kernel"],
                                              product_name=obj.cleaned_data["product_name"],
                                              ipv4_address=obj.cleaned_data["ipv4_address"],
                                              mac_address=obj.cleaned_data["mac_address"],
                                              mem_total=obj.cleaned_data["mem_total"],
                                              mem_explain=obj.cleaned_data["mem_explain"],
                                              disk_total=obj.cleaned_data["disk_total"],
                                              disk_explain=obj.cleaned_data["disk_explain"],
                                              minion_id=obj.cleaned_data["minion_id"],
                                              idc_name=obj.cleaned_data["idc_name"],
                                              idc_num=obj.cleaned_data["idc_num"],
                                              login_ip=obj.cleaned_data["login_ip"],
                                              login_port=obj.cleaned_data["login_port"],
                                              login_user=obj.cleaned_data["login_user"],
                                              login_password=obj.cleaned_data["login_password"],
                                              description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'update_server_list':
                obj = ServerListUpdateForm(request.POST)
                if obj.is_valid():
                    ServerList.objects.filter(server_name=obj.cleaned_data["server_name"]).update(
                        server_type=obj.cleaned_data["server_type"],
                        localhost=obj.cleaned_data["localhost"],
                        ip=obj.cleaned_data["ip"],
                        system_issue=obj.cleaned_data["system_issue"],
                        sn=obj.cleaned_data["sn"],
                        cpu_num=obj.cleaned_data["cpu_num"],
                        cpu_model=obj.cleaned_data["cpu_model"],
                        sys_type=obj.cleaned_data["sys_type"],
                        kernel=obj.cleaned_data["kernel"],
                        product_name=obj.cleaned_data["product_name"],
                        ipv4_address=obj.cleaned_data["ipv4_address"],
                        mac_address=obj.cleaned_data["mac_address"],
                        mem_total=obj.cleaned_data["mem_total"],
                        mem_explain=obj.cleaned_data["mem_explain"],
                        disk_total=obj.cleaned_data["disk_total"],
                        disk_explain=obj.cleaned_data["disk_explain"],
                        minion_id=obj.cleaned_data["minion_id"],
                        idc_name=obj.cleaned_data["idc_name"],
                        idc_num=obj.cleaned_data["idc_num"],
                        login_ip=obj.cleaned_data["login_ip"],
                        login_port=obj.cleaned_data["login_port"],
                        login_user=obj.cleaned_data["login_user"],
                        login_password=obj.cleaned_data["login_password"],
                        update_time=timezone.now(),
                        description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'delete_server_list':
                server_name = request.POST.get('server_name')
                try:
                    ServerList.objects.get(server_name=server_name).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'import_server_list':
                # 获取的数据类型如果有文件，则要像下面这么写form才能正确获取到文件
                obj = ServerListImportForm(request.POST, request.FILES)
                if obj.is_valid():
                    # request.FILES['file'].read()是从内存中直接读取文件内容
                    # openpyxl的filename除了能打开某个路径的文件，还可以接受BytesIO读取的二进制数据所以如下即可获取xlsx内容
                    wb = openpyxl.load_workbook(filename=BytesIO(request.FILES['file'].read()))
                    table = wb[wb.sheetnames[0]]
                    server_list_cover_select = request.POST.get('server_list_cover_select')
                    # 结果需要存一些内容所以修改成列表
                    result['result'] = []
                    for row in table.iter_rows(min_row=2, max_col=25):
                        data = {'server_name': row[0].value, 'server_type': row[1].value, 'localhost': row[2].value,
                                'ip': row[3].value, 'system_issue': row[4].value, 'sn': row[5].value,
                                'cpu_num': row[6].value, 'cpu_model': row[7].value, 'sys_type': row[8].value,
                                'kernel': row[9].value, 'product_name': row[10].value, 'ipv4_address': row[11].value,
                                'mac_address': row[12].value, 'mem_total': row[13].value, 'mem_explain': row[14].value,
                                'disk_total': row[15].value, 'disk_explain': row[16].value, 'minion_id': row[17].value,
                                'idc_name': row[18].value, 'idc_num': row[17].value, 'login_ip': row[20].value,
                                'login_port': row[21].value, 'login_user': row[22].value,
                                'login_password': row[23].value, 'description': row[24].value}
                        obj = ServerListAddForm(data)
                        if obj.is_valid():
                            ServerList.objects.create(server_name=obj.cleaned_data["server_name"],
                                                      server_type=obj.cleaned_data["server_type"],
                                                      localhost=obj.cleaned_data["localhost"],
                                                      ip=obj.cleaned_data["ip"],
                                                      system_issue=obj.cleaned_data["system_issue"],
                                                      sn=obj.cleaned_data["sn"],
                                                      cpu_num=obj.cleaned_data["cpu_num"],
                                                      cpu_model=obj.cleaned_data["cpu_model"],
                                                      sys_type=obj.cleaned_data["sys_type"],
                                                      kernel=obj.cleaned_data["kernel"],
                                                      product_name=obj.cleaned_data["product_name"],
                                                      ipv4_address=obj.cleaned_data["ipv4_address"],
                                                      mac_address=obj.cleaned_data["mac_address"],
                                                      mem_total=obj.cleaned_data["mem_total"],
                                                      mem_explain=obj.cleaned_data["mem_explain"],
                                                      disk_total=obj.cleaned_data["disk_total"],
                                                      disk_explain=obj.cleaned_data["disk_explain"],
                                                      minion_id=obj.cleaned_data["minion_id"],
                                                      idc_name=obj.cleaned_data["idc_name"],
                                                      idc_num=obj.cleaned_data["idc_num"],
                                                      login_ip=obj.cleaned_data["login_ip"],
                                                      login_port=obj.cleaned_data["login_port"],
                                                      login_user=obj.cleaned_data["login_user"],
                                                      login_password=obj.cleaned_data["login_password"],
                                                      description=obj.cleaned_data["description"])
                        else:
                            error_str = json.loads(obj.errors.as_json())
                            if error_str.get('server_name') and server_list_cover_select == 'cover':
                                if error_str['server_name'][0]['message'] == '服务器名称已存在，请检查':
                                    obj = ServerListUpdateForm(data)
                                    if obj.is_valid():
                                        ServerList.objects.filter(server_name=obj.cleaned_data["server_name"]).update(
                                            server_type=obj.cleaned_data["server_type"],
                                            localhost=obj.cleaned_data["localhost"],
                                            ip=obj.cleaned_data["ip"],
                                            system_issue=obj.cleaned_data["system_issue"],
                                            sn=obj.cleaned_data["sn"],
                                            cpu_num=obj.cleaned_data["cpu_num"],
                                            cpu_model=obj.cleaned_data["cpu_model"],
                                            sys_type=obj.cleaned_data["sys_type"],
                                            kernel=obj.cleaned_data["kernel"],
                                            product_name=obj.cleaned_data["product_name"],
                                            ipv4_address=obj.cleaned_data["ipv4_address"],
                                            mac_address=obj.cleaned_data["mac_address"],
                                            mem_total=obj.cleaned_data["mem_total"],
                                            mem_explain=obj.cleaned_data["mem_explain"],
                                            disk_total=obj.cleaned_data["disk_total"],
                                            disk_explain=obj.cleaned_data["disk_explain"],
                                            minion_id=obj.cleaned_data["minion_id"],
                                            idc_name=obj.cleaned_data["idc_name"],
                                            idc_num=obj.cleaned_data["idc_num"],
                                            login_ip=obj.cleaned_data["login_ip"],
                                            login_port=obj.cleaned_data["login_port"],
                                            login_user=obj.cleaned_data["login_user"],
                                            login_password=obj.cleaned_data["login_password"],
                                            update_time=timezone.now(),
                                            description=obj.cleaned_data["description"])
                                    else:
                                        error_str = json.loads(obj.errors.as_json())
                                        result['result'].append(error_str)
                            else:
                                if error_str['server_name'][0]['message'] == '服务器名称已存在，请检查':
                                    result['result'].append('服务器名称:{} 已存在'.format(data['server_name']))
                                else:
                                    result['result'].append(error_str)
                    if result['result']:
                        pass
                    else:
                        result['result'] = '成功'
                        result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            else:
                result['result'] = '主机管理页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('主机管理页ajax提交处理有问题', e)
        result['result'] = '主机管理页ajax提交处理有问题'
        return JsonResponse(result)


# salt执行state.sls的返回结果格式化，因为通过api返回的结果不怎么好看呵呵
def format_state(result):
    a = result
    # b是返回minion列表
    b = (a['return'])
    # 用来存放所有minion格式化后的结果的
    result_data = []
    try:
        # i是return后面的列表其实就是a['return'][0]
        for i in b:
            # key是minion的ID,value是这个ID执行的所有结果又是一个字典
            for key, value in i.items():
                succeeded = 0
                failed = 0
                changed = 0
                Total_states_run = 0
                Total_run_time = 0
                minion_id = key
                run_num = len(value)  # 得到执行的state个数
                result_list = [k for k in range(run_num)] #把列表先用数字撑大，因为接收的数据随机的顺序如（3,5,6），先撑开列表到时候假设是3过来就插3的位子这样顺序就有序了
                for key1, value1 in value.items():
                    # print(value1)
                    # key1是一个个state的ID，value1是每个state的结果
                    key1 = key1.split('_|-')
                    Function = key1[0] + '_' + key1[-1]
                    ID = key1[1]
                    Name = key1[2]
                    aaa = '----------\n' + 'ID: '.rjust(14) + ID + '\n' + 'Function: '.rjust(
                        14) + Function + '\n' + 'Name: '.rjust(14) + Name + '\n' + 'Result: '.rjust(14) + str(
                        value1['result']) + '\n' + 'Comment: '.rjust(14) + value1['comment'] + '\n'
                    # start_time有的没有有的有
                    if value1.get('start_time'):
                        aaa += 'Started: '.rjust(14) + str(value1['start_time']) + '\n'
                    # duration有的没有有的有
                    if value1.get('duration'):
                        aaa += 'Duration: '.rjust(14) + str(value1['duration']) + ' ms' + '\n'
                        Total_run_time += value1['duration']
                    # changes都有，就算没值也是一个空的{}
                    if value1['changes'] == {}:
                        aaa += 'Changes: '.rjust(14)+'\n'
                    elif type(value1['changes']) == str:
                        aaa += 'ChangesIs: '.rjust(14) + '\n' + ''.rjust(14) + '----------\n'
                        aaa += ''.rjust(14) + value1['changes'] + ':\n' + ''.rjust(18) + '----------\n'
                    else:
                        aaa += 'ChangesIs: '.rjust(14) + '\n' + ''.rjust(14) + '----------\n'
                        for key in value1['changes'].keys():

                            if type(value1['changes'][key]) == dict:
                                aaa += ''.rjust(14) + key + ':\n' + ''.rjust(18) + '----------\n'
                                for ckey, cvalue in value1['changes'][key].items():
                                    aaa += ''.rjust(18) + ckey + ':\n' + ''.rjust(22) + str(cvalue).replace('\n','\n'+' '*18) + '\n'
                            else:
                                aaa += ''.rjust(14) + key + ':\n' + ''.rjust(18) + str(value1['changes'][key]).replace('\n','\n'+' '*18) + '\n'
                        changed += 1
                    if value1.get('__run_num__') is None:
                        result_list.append(aaa)
                    else:
                        result_list[value1.get('__run_num__')] = aaa
                    if value1['result']:
                        succeeded += 1
                    else:
                        failed += 1
                    Total_states_run += 1
                Total_run_time = Total_run_time / 1000
                bbb =74*'-'+ '\nSummary for %s\n-------------\nSucceeded: %d (changed=%d)\nFailed:    %2d\n-------------\nTotal states run:     %d\nTotal run time:    %.3f s\n\n' % (
                minion_id, succeeded, changed, failed, Total_states_run, Total_run_time)
                result_list.insert(0, bbb)
                result_data.extend(result_list)
        return result_data
    #如果格式化有问题，就把原来的以str来返回，然后在调用这个格式化的方法里写判断如果为str说明格式化失败，然后该怎么处理就怎么处理呵呵
    except Exception as e:
        logger.error('格式化不成功'+str(e))
        return str(a)


# 新的模块部署方法，采用了异步，并且加强了逻辑判断，加强了容错率，后期准备再加入入库用来做salt任务管理使用
def module_deploy(request):
    if request.method == 'GET':
        return render(request, 'module_deploy.html')
    elif request.is_ajax() and request.POST.get('module_deploy_tag_key') == 'state_exe':
        # 这里没有做对传入tgt判断是否存在并且是否离线，自己给自己埋坑了因为可以支持多种tgt_type判断太麻烦了蛋疼，不过下面有通过async_state_api返回值判断！！！！！！！！！！！！！！！
        data_list = []  # 用来存结果的
        try:
            tgt = request.POST.get('tgt', None)
            tgt_type = request.POST.get('tgt_type')
            arg = request.POST.getlist('arg')
            # 下面这个是执行state用队列模式，还有一种是可以用并行，用队列安全性高一些不会出现同时执行同一个state对同一台minion但是相应速度慢
            # 并且我发现如果在下面api接口async_state_api里直接写arg=arg.append('queue=True')是不行的奶奶的，弄半天才发现！
            arg.append('queue=True')
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                # 当调用api失败的时候比如salt-api服务stop了会返回false
                if saltapi.get_token() is False:
                    response_data = '模块部署失败，SaltAPI调用get_token请求出错'
                    return JsonResponse({'result': response_data, 'status': False})
                else:
                    # 调用async_state_api方法来执行部署
                    jid = saltapi.async_state_api(tgt=tgt, tgt_type=tgt_type, arg=arg)
                    # 当调用api失败的时候比如salt-api服务stop了会返回false
                    if jid is False:
                        response_data = '模块部署失败，SaltAPI调用async_state_api请求出错'
                        return JsonResponse({'result': response_data, 'status': False})
                    # 判断jid如果返回值如果为[{}]表明没有这个minion，我会在后期保证不出现这种现象也就不需要这个判断了，从源头解决！
                    elif jid['return'] == [{}]:
                        response_data = '模块部署失败，请确认minion是否存在。。'
                        return JsonResponse({'result': response_data, 'status': False})
                    else:
                        jid = jid['return'][0]['jid']
                        check_count = 60
                        re_count = 0
                        time.sleep(15)
                        while check_count:
                            false_count = 0
                            job_status = saltapi.job_active_api(tgt=tgt, tgt_type=tgt_type, arg=jid)
                            if job_status is False:
                                response_data = '模块部署失败，SaltAPI调用job_active_api请求出错'
                                return JsonResponse({'result': response_data, 'status': False})
                            else:
                                # 获取minion的数量用来给下面做对比使用
                                minion_count = len(job_status['return'][0].items())

                                '''这步是因为结果是一个字典套字典，如：{'return': [{'192.168.68.51': {'fun': 'cmd.script', 'arg': ['salt://script/test.py'],
                                'tgt_type': 'glob', 'user': 'saltapi', 'pid': 3015, 'jid': '20170807144951631265', 'ret': '', 'tgt': '192.168.68.51'},
                                '192.168.68.50-master': {}}]}里面有2个minion'''
                                for key, value in job_status['return'][0].items():
                                    # !!使用job_active_api方法检测任务状态，如果有值说明还在运行，如果值为{}即空说明运行结束或为False表示minion离线，这步经典想了好久！！
                                    # 因为如果直接判断为空或者为False没想过如何确认全部都是为空，最后终于想到反着判断哈哈
                                    if value:
                                        # 为真说明job还在执行，刚好用来恢复断线false的计数器
                                        if re_count > 0:
                                            re_count = 0
                                        # 如果为真说明还没运行结束，break跳出不会运行下面的else，则jid_data为空，然后跳到计数-1再睡眠30秒继续下一轮
                                        break
                                    # 这个留在这里做个说明，我发现在调用job_active_api接口的时候经常失败返回false了，感觉是接口有问题
                                    # 而如果出现都是false用jid_api接口取到的结果就会是[{}]所以下面对这个要做一层判断，以免因为接口不稳导致没取到结果
                                    # 另外注意这里value is False看上去好像和上面是if value是相反的可以直接用else代替，但是不行！因为当执行完毕返回是{}而{}和False是不同的！
                                    if value is False:
                                        # 一次falst就计数器+1直到等于你执行任务的minion的数量，说明特么全断线了
                                        false_count += 1
                                        # 如果全部是minion（假设你是批量部署）都false并且连续监测2次都是那就不用跑了直接返回离线结束呵呵
                                        if false_count == minion_count and re_count == 2:
                                            response_data = {'result': '对不起您要部署的所有minion都离线了',
                                                             'status': False}
                                            return JsonResponse(response_data)
                                        # re计数器不到3次则+1，继续下一轮循环
                                        elif false_count == minion_count:
                                            # print('全fasle出现了', re_count)
                                            re_count += 1
                                            break

                                # 当上面的for正常执行结束没有因为break退出时候说明job执行完毕了(或者是有部分执行完毕有部分返回了false)，则执行下面
                                else:
                                    jid_data = saltapi.jid_api(jid=jid)
                                    # 注意[{}] ！= False所以不能用if jid_data['return']判断是否有数据，这个坑埋了好久奶奶的！！！
                                    if jid_data is False:
                                        response_data = '模块部署失败，SaltAPI调用jid_api请求出错'
                                        return JsonResponse({'result': response_data, 'status': False})
                                    elif jid_data['return'] == [{}]:
                                        # 这个判断没必要，只是留这里做个说明，我之前上面没有做if value is False判断的时候，如果job_active_api
                                        # 的结果全部false了也会正常跳出for循环，然后在这里会出现jid_data['return'] == [{}]的情况，因为false
                                        # 说明minion断线了，结果肯定取到空了；还有另一种情况就是还没有返回值的时候也会等于[{}],
                                        # 不过后面我在上面加了对false做判断这里就没必要了呵呵
                                        pass
                                    else:
                                        format_result = format_state(jid_data)
                                        # 这个是对格式化输出的一个判断，类型str说明格式化出错了呵呵，一般在minion一个sls未执行完成又执行会出现
                                        if type(format_result) == str:
                                            data_list.append(format_state(jid_data))
                                            response_data = {'result': jid_data['return'][0], 'status': False}
                                            return JsonResponse(response_data)
                                        else:
                                            data_list.extend(format_state(jid_data))
                                            break
                                check_count -= 1
                                time.sleep(10)
                        else:
                            response_data = {
                                'result': '执行时间已经超过10分钟了，建议您可以去salt命令执行里有任务查看快捷通道，jid：%s 的结果以免耽误泡妞时间。。' % jid,
                                'status': False}
                            return JsonResponse(response_data)
        except Exception as e:
            response_data = {'result': '模块部署失败，内部代码可能有问题：%s。。' % str(e), 'status': False}
            return JsonResponse(response_data)
        response_data = {'result': data_list, 'status': True}
        return JsonResponse(response_data)  # 新版1.7以上写法


# 网络扫描
def net_tool(request):
    try:
        if request.method == 'GET':
            return render(request, 'net_tool.html')
        elif request.method == 'POST':
            # 网段扫描
            if request.POST.get('scan_type') == 'ipscan':
                ipscan_ip = request.POST.get('ipscan_ip', None)
                try:
                    result_data = netscantool.ipscan(ipscan_ip)
                    response_data = {'result': result_data, 'status': True}
                    return JsonResponse(response_data)
                except Exception as e:
                    logger.error('无法获取IP网段扫描结果'+str(e))
                    return JsonResponse({'result': '无法获取IP网段扫描结果', 'status': False})
            # TCP端口扫描
            elif request.POST.get('scan_type') == 'portscan':
                portscan_ip = request.POST.get('portscan_ip', None)
                portscan_startport = ast.literal_eval(request.POST.get('portscan_startport', None))
                portscan_endport = ast.literal_eval(request.POST.get('portscan_endport', None))
                try:
                    result_data = netscantool.portscan(portscan_ip,portscan_startport,portscan_endport)

                    response_data = {'result': result_data, 'status': True}
                    return JsonResponse(response_data)
                except Exception as e:
                    logger.error('无法获取端口扫描结果'+str(e))
                    return JsonResponse({'result': '无法获取端口扫描结果', 'status': False})
            # 路由跟踪
            elif request.POST.get('scan_type') == 'traceroutescan':
                tracert_data = request.POST.get('tracert_data', None)
                try:
                    result_data = netscantool.traceroutescan(tracert_data)
                    response_data = {'result': result_data, 'status': True}
                    return JsonResponse(response_data)
                except Exception as e:
                    logger.error('无法获取路由跟踪结果'+str(e))
                    return JsonResponse({'result': '无法获取路由跟踪结果', 'status': False})

    except Exception as e:
        logger.error('网络扫描报错：', e)


# salt-ssh运行state的输出格式化，和salt运行state差不多，有时间可以试着整合
def ssh_format_state(result):
    a = result
    succeeded = 0
    failed = 0
    changed = 0
    Total_states_run = 0
    Total_run_time = 0
    # b是返回minion列表
    b = a['return']
    try:
        for i in b:
            for key, value in i.items():
                minion_id = key
                run_num = len(value['return'])  # 得到执行的state个数,就是这里和salt格式化不同
                result_list = [k for k in range(run_num)]  # 把列表先用数字撑大，因为接收的数据随机的顺序如（3,5,6），先撑开列表到时候假设是3过来就插3的位子这样顺序就有序了
                for key1, value1 in value['return'].items():
                    # print(value1)
                    # key1是一个个state的ID，value1是每个state的结果
                    key1 = key1.split('_|-')
                    Function = key1[0] + '_' + key1[-1]
                    ID = key1[1]
                    Name = key1[2]
                    aaa = '----------\n' + 'ID: '.rjust(14) + ID + '\n' + 'Function: '.rjust(
                        14) + Function + '\n' + 'Name: '.rjust(14) + Name + '\n' + 'Result: '.rjust(14) + str(
                        value1['result']) + '\n' + 'Comment: '.rjust(14) + value1['comment'] + '\n'
                    # start_time有的没有有的有
                    if value1.get('start_time'):
                        aaa += 'Started: '.rjust(14) + str(value1['start_time']) + '\n'
                    # duration有的没有有的有
                    if value1.get('duration'):
                        aaa += 'Duration: '.rjust(14) + str(value1['duration']) + ' ms' + '\n'
                        Total_run_time += value1['duration']
                    # changes都有，就算没值也是一个空的{}
                    # print(value1['changes'])
                    if value1['changes'] == {}:
                        aaa += 'Changes: '.rjust(14) + '\n'
                    else:
                        #这里用ChangesIs而不是Changes是为了前端区分方便改颜色哈哈
                        aaa += 'ChangesIs: '.rjust(14) + '\n' + ''.rjust(14) + '----------\n'
                        for key in value1['changes'].keys():

                            if type(value1['changes'][key]) == dict:
                                aaa += ''.rjust(14) + key + ':\n' + ''.rjust(18) + '----------\n'
                                # print(value1['changes'][key])
                                # aaa+=''.rjust(14) + key + ':\n'+ ''.rjust(18) + str(value1['changes'][key]) + '\n'+ '\n' + ''.rjust(14) + '----------\n'
                                # aaa +=''.rjust(18) + '----------\n' + ''.rjust(18) + '----------\n'
                                for ckey, cvalue in value1['changes'][key].items():
                                    aaa += ''.rjust(18) + ckey + ':\n' + ''.rjust(22) + str(cvalue).replace('\n',
                                                                                                            '\n' + ' ' * 18) + '\n'
                                    # aaa+=''.rjust(14) + ckey + ':\n'+ ''.rjust(18) + str(value1['changes'][key][ckey]) + '\n'
                            else:
                                aaa += ''.rjust(14) + key + ':\n' + ''.rjust(18) + str(value1['changes'][key]).replace('\n',
                                                                                                                       '\n' + ' ' * 18) + '\n'
                        changed += 1
                    result_list[value1['__run_num__']] = aaa
                    if value1['result']:
                        succeeded += 1
                    else:
                        failed += 1
                    Total_states_run += 1
                Total_run_time = Total_run_time / 1000
                bbb = 74 * '-' + '\nSummary for %s\n-------------\nSucceeded: %d (changed=%d)\nFailed:    %2d\n-------------\nTotal states run:     %d\nTotal run time:    %.3f s\n\n' % (
                    minion_id, succeeded, changed, failed, Total_states_run, Total_run_time)
                result_list.insert(0, bbb)
                return result_list
                # 如果格式化有问题，就把原来的以str来返回，然后在调用这个格式化的方法里写判断如果为str说明格式化失败，然后该怎么处理就怎么处理呵呵
    except Exception as e:
        logger.error('格式化不成功' + str(e))
        return str(a)


# minion客户端salt-ssh部署
def minion_client_install(request):
    try:
        if request.method == 'GET':
            return render(request, 'minion_client_install.html')
        else:
            server_ip = request.POST.get('server_ip')
            server_username = request.POST.get('server_username')
            server_password = request.POST.get('server_password')
            # 我在前端设置了如果没输入minion_id则会返回''空，不过这个还是会被django的get收到只是变成minion_id=''还是存在minion_id的
            # 所以我在后端的sls文件里加了判断如果为''的判断
            minion_id = request.POST.get('minion_id')
            if request.POST.get('master'):
                master = request.POST.get('master')
            else:
                master = settings.SITE_SALT_MASTER_IP
            server_port = request.POST.get('server_port')
            try:
                # 直接打开耦合性就比较差了，我这为了方便，其实最好是用salt的cmd.run来打开写入比较正规，而且低耦合
                with open(r'/etc/salt/roster','w') as roster:
                    # 本来我是没打算用覆盖写入的方式，后来想想如果不用覆盖写入判断起来更耗性能，所以干脆简单点改成覆盖了
                    roster.write('\n%s:\n  host: %s\n  user: %s\n  passwd: %s\n  port: %s\n' % (server_ip, server_ip, server_username, server_password, server_port))
                print('主机信息写入成功')
            except Exception as e:
                print('写入主机信息失败')
                return JsonResponse({'result': '写入主机信息失败', 'status': False})
            else:
                data = {'client': 'ssh',
                        'tgt': server_ip,
                        'fun': 'test.ping'
                        }
                try:
                    rpost = requests.post('%s/run' % settings.SITE_SALT_API_URL, json=data)
                    rpost.raise_for_status()
                    #这里要注意！！！！！，这步测试客户机连通性基本100%不会报错，因为报错也会返回错误结果，所以这里的try是
                    #为了验证接口是否能正常通信
                except Exception as e:
                    reponse_data = 'ssh接口有问题，请联系管理员处理'
                    print(reponse_data + str(e))
                    logger.error(reponse_data + str(e))
                    return JsonResponse({'result': reponse_data, 'status': False})
                else:
                    if 'Permission denied, please try again' in json.dumps(rpost.json()):
                        response_data = {'result': '用户名或密码错误', 'status': False}
                        print('minion部署失败了用户名或密码错误',rpost.json())
                        return JsonResponse(response_data)
                    elif 'Connection refused' in json.dumps(rpost.json()):
                        response_data = {'result': '无法连接%s,请确认IP是否正确' % server_ip, 'status': False}
                        print('minion部署失败了无法连接该地址', rpost.json())
                        return JsonResponse(response_data)
                    elif '"return": true' in json.dumps(rpost.json()):
                        print('连通性以及用户名密码验证成功')
                        # 如果目标主机验证没有问题则开始执行部署工作，我有个想法是ajax写2层嵌套，就是验证做一层成功先在页面
                        # 返回一个成功的提示并且显示开始部署然后再到后台执行部署，不过没仔细想如何实现，这里就先不做两层直接后台一次处理
                        minion_data = {'client': 'ssh',
                                'tgt': server_ip,
                                'fun': 'state.sls',
                                # 注意下面pillar的写法，主要是双引号内带单引号或者单引号内带双引号，不要全用单引号或双引号不然会报错
                                'arg': ["minion", "pillar={'minion_id':'%s','master':'%s'}" % (minion_id, master)]
                                }
                        try:
                            minion_post = requests.post('%s/run' % settings.SITE_SALT_API_URL, json=minion_data)
                            minion_post.raise_for_status()
                        except Exception as e:
                            minion_response = '主机%s部署minion客户端失败请联系管理员：' % server_ip
                            print(minion_response + str(e))
                            logger.error(minion_response + str(e))
                            # return logger.info(e + ' 无法获取数据')
                            return JsonResponse({'result': minion_response, 'status': False})
                        else:
                            # 格式化输出
                            data_list = ssh_format_state(minion_post.json())
                            minion_response = {'result': data_list, 'status': True}
                            return JsonResponse(minion_response)
                    else:
                        response_data = {'result': '无法连接%s,请确认master是否可以和客户机通信' % server_ip, 'status': False}
                        logger.error('minion部署失败了请确认master是否可以和客户机通信' + str(rpost.json()))
                        return JsonResponse(response_data)
    except Exception as e:
        logger.error('minion安装部署报错：', e)


# 主机资源列表
def server_list(request):
    try:
        # 这里主要是担心通过ajax提交了get请求，其实多此一举因为这个页面没有ajax的get请求哈哈
        if request.is_ajax() is not True:
            if request.method == 'GET':
                search_field = request.GET.get('search_field', '')
                search_content = request.GET.get('search_content', '')
                if search_content is '':
                    server_data = ServerList.objects.all().order_by('create_date')
                    data_list = getPage(request, server_data, 9)
                else:
                    if search_field == 'search_server_name':
                        server_data = ServerList.objects.filter(server_name__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                    elif search_field == 'search_ip':
                        server_data = ServerList.objects.filter(ip__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                    elif search_field == 'search_sys':
                        server_data = ServerList.objects.filter(sys__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                    else:
                        server_data = ServerList.objects.filter(server_type__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                return render(request, 'server_list.html',
                              {'data_list': data_list, 'search_field': search_field, 'search_content': search_content})

    except Exception as e:
        logger.error('主机列表管理页面有问题', e)
        return render(request, 'server_list.html')


# 主机资源列表ajax
def server_list_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.GET.get('server_list_tag_key') == 'modal_search_minion_id':
                minion_id = request.GET.get('minion_id')
                sys = request.GET.get('sys_type')
                minion_id_list = MinionList.objects.filter(minion_id__icontains=minion_id, sys=sys).order_by(
                    'create_date').values_list('minion_id', flat=True)
                result['result'] = list(minion_id_list)
                result['status'] = True
                # 返回字典之外的需要把参数safe改成false如：JsonResponse([1, 2, 3], safe=False)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'add_server_list':
                obj = ServerListAddForm(request.POST)
                if obj.is_valid():
                    ServerList.objects.create(server_name=obj.cleaned_data["server_name"],
                                              server_type=obj.cleaned_data["server_type"],
                                              localhost=obj.cleaned_data["localhost"],
                                              ip=obj.cleaned_data["ip"],
                                              system_issue=obj.cleaned_data["system_issue"],
                                              sn=obj.cleaned_data["sn"],
                                              cpu_num=obj.cleaned_data["cpu_num"],
                                              cpu_model=obj.cleaned_data["cpu_model"],
                                              sys_type=obj.cleaned_data["sys_type"],
                                              kernel=obj.cleaned_data["kernel"],
                                              product_name=obj.cleaned_data["product_name"],
                                              ipv4_address=obj.cleaned_data["ipv4_address"],
                                              mac_address=obj.cleaned_data["mac_address"],
                                              mem_total=obj.cleaned_data["mem_total"],
                                              mem_explain=obj.cleaned_data["mem_explain"],
                                              disk_total=obj.cleaned_data["disk_total"],
                                              disk_explain=obj.cleaned_data["disk_explain"],
                                              minion_id=obj.cleaned_data["minion_id"],
                                              idc_name=obj.cleaned_data["idc_name"],
                                              idc_num=obj.cleaned_data["idc_num"],
                                              login_ip=obj.cleaned_data["login_ip"],
                                              login_port=obj.cleaned_data["login_port"],
                                              login_user=obj.cleaned_data["login_user"],
                                              login_password=obj.cleaned_data["login_password"],
                                              description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'update_server_list':
                obj = ServerListUpdateForm(request.POST)
                if obj.is_valid():
                    ServerList.objects.filter(server_name=obj.cleaned_data["server_name"]).update(
                        server_type=obj.cleaned_data["server_type"],
                        localhost=obj.cleaned_data["localhost"],
                        ip=obj.cleaned_data["ip"],
                        system_issue=obj.cleaned_data["system_issue"],
                        sn=obj.cleaned_data["sn"],
                        cpu_num=obj.cleaned_data["cpu_num"],
                        cpu_model=obj.cleaned_data["cpu_model"],
                        sys_type=obj.cleaned_data["sys_type"],
                        kernel=obj.cleaned_data["kernel"],
                        product_name=obj.cleaned_data["product_name"],
                        ipv4_address=obj.cleaned_data["ipv4_address"],
                        mac_address=obj.cleaned_data["mac_address"],
                        mem_total=obj.cleaned_data["mem_total"],
                        mem_explain=obj.cleaned_data["mem_explain"],
                        disk_total=obj.cleaned_data["disk_total"],
                        disk_explain=obj.cleaned_data["disk_explain"],
                        minion_id=obj.cleaned_data["minion_id"],
                        idc_name=obj.cleaned_data["idc_name"],
                        idc_num=obj.cleaned_data["idc_num"],
                        login_ip=obj.cleaned_data["login_ip"],
                        login_port=obj.cleaned_data["login_port"],
                        login_user=obj.cleaned_data["login_user"],
                        login_password=obj.cleaned_data["login_password"],
                        update_time=timezone.now(),
                        description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'delete_server_list':
                server_name = request.POST.get('server_name')
                try:
                    ServerList.objects.get(server_name=server_name).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'import_server_list':
                # 获取的数据类型如果有文件，则要像下面这么写form才能正确获取到文件
                obj = ServerListImportForm(request.POST, request.FILES)
                if obj.is_valid():
                    # request.FILES['file'].read()是从内存中直接读取文件内容
                    # openpyxl的filename除了能打开某个路径的文件，还可以接受BytesIO读取的二进制数据所以如下即可获取xlsx内容
                    wb = openpyxl.load_workbook(filename=BytesIO(request.FILES['file'].read()))
                    table = wb[wb.sheetnames[0]]
                    server_list_cover_select = request.POST.get('server_list_cover_select')
                    # 结果需要存一些内容所以修改成列表
                    result['result'] = []
                    for row in table.iter_rows(min_row=2, max_col=25):
                        data = {'server_name': row[0].value, 'server_type': row[1].value, 'localhost': row[2].value,
                                'ip': row[3].value, 'system_issue': row[4].value, 'sn': row[5].value,
                                'cpu_num': row[6].value, 'cpu_model': row[7].value, 'sys_type': row[8].value,
                                'kernel': row[9].value, 'product_name': row[10].value, 'ipv4_address': row[11].value,
                                'mac_address': row[12].value, 'mem_total': row[13].value, 'mem_explain': row[14].value,
                                'disk_total': row[15].value, 'disk_explain': row[16].value, 'minion_id': row[17].value,
                                'idc_name': row[18].value, 'idc_num': row[17].value, 'login_ip': row[20].value,
                                'login_port': row[21].value, 'login_user': row[22].value,
                                'login_password': row[23].value, 'description': row[24].value}
                        obj = ServerListAddForm(data)
                        if obj.is_valid():
                            ServerList.objects.create(server_name=obj.cleaned_data["server_name"],
                                                      server_type=obj.cleaned_data["server_type"],
                                                      localhost=obj.cleaned_data["localhost"],
                                                      ip=obj.cleaned_data["ip"],
                                                      system_issue=obj.cleaned_data["system_issue"],
                                                      sn=obj.cleaned_data["sn"],
                                                      cpu_num=obj.cleaned_data["cpu_num"],
                                                      cpu_model=obj.cleaned_data["cpu_model"],
                                                      sys_type=obj.cleaned_data["sys_type"],
                                                      kernel=obj.cleaned_data["kernel"],
                                                      product_name=obj.cleaned_data["product_name"],
                                                      ipv4_address=obj.cleaned_data["ipv4_address"],
                                                      mac_address=obj.cleaned_data["mac_address"],
                                                      mem_total=obj.cleaned_data["mem_total"],
                                                      mem_explain=obj.cleaned_data["mem_explain"],
                                                      disk_total=obj.cleaned_data["disk_total"],
                                                      disk_explain=obj.cleaned_data["disk_explain"],
                                                      minion_id=obj.cleaned_data["minion_id"],
                                                      idc_name=obj.cleaned_data["idc_name"],
                                                      idc_num=obj.cleaned_data["idc_num"],
                                                      login_ip=obj.cleaned_data["login_ip"],
                                                      login_port=obj.cleaned_data["login_port"],
                                                      login_user=obj.cleaned_data["login_user"],
                                                      login_password=obj.cleaned_data["login_password"],
                                                      description=obj.cleaned_data["description"])
                        else:
                            error_str = json.loads(obj.errors.as_json())
                            if error_str.get('server_name') and server_list_cover_select == 'cover':
                                if error_str['server_name'][0]['message'] == '服务器名称已存在，请检查':
                                    obj = ServerListUpdateForm(data)
                                    if obj.is_valid():
                                        ServerList.objects.filter(server_name=obj.cleaned_data["server_name"]).update(
                                            server_type=obj.cleaned_data["server_type"],
                                            localhost=obj.cleaned_data["localhost"],
                                            ip=obj.cleaned_data["ip"],
                                            system_issue=obj.cleaned_data["system_issue"],
                                            sn=obj.cleaned_data["sn"],
                                            cpu_num=obj.cleaned_data["cpu_num"],
                                            cpu_model=obj.cleaned_data["cpu_model"],
                                            sys_type=obj.cleaned_data["sys_type"],
                                            kernel=obj.cleaned_data["kernel"],
                                            product_name=obj.cleaned_data["product_name"],
                                            ipv4_address=obj.cleaned_data["ipv4_address"],
                                            mac_address=obj.cleaned_data["mac_address"],
                                            mem_total=obj.cleaned_data["mem_total"],
                                            mem_explain=obj.cleaned_data["mem_explain"],
                                            disk_total=obj.cleaned_data["disk_total"],
                                            disk_explain=obj.cleaned_data["disk_explain"],
                                            minion_id=obj.cleaned_data["minion_id"],
                                            idc_name=obj.cleaned_data["idc_name"],
                                            idc_num=obj.cleaned_data["idc_num"],
                                            login_ip=obj.cleaned_data["login_ip"],
                                            login_port=obj.cleaned_data["login_port"],
                                            login_user=obj.cleaned_data["login_user"],
                                            login_password=obj.cleaned_data["login_password"],
                                            update_time=timezone.now(),
                                            description=obj.cleaned_data["description"])
                                    else:
                                        error_str = json.loads(obj.errors.as_json())
                                        result['result'].append(error_str)
                            else:
                                if error_str['server_name'][0]['message'] == '服务器名称已存在，请检查':
                                    result['result'].append('服务器名称:{} 已存在'.format(data['server_name']))
                                else:
                                    result['result'].append(error_str)
                    if result['result']:
                        pass
                    else:
                        result['result'] = '成功'
                        result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            else:
                result['result'] = '主机管理页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('主机管理页ajax提交处理有问题', e)
        result['result'] = '主机管理页ajax提交处理有问题'
        return JsonResponse(result)


# 主机资源管理模板下载
def server_list_template_down(request):
    try:
        # 基本上django下载文件就按这个模板来就可以，更复杂的参考下面的主机管理列表导出方法
        # FileResponse方法继承了StreamingHttpResponse并且封装了迭代方法，是django最好的大文件流传送方式了，用法直接按下面，很简单
        file = settings.STATICFILES_DIRS[0] + "/download_files/主机模板.xlsx"
        response = FileResponse(open(file, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        # 带中文的文件名需要如下用escape_uri_path和utf8才能识别
        response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path(os.path.basename(file)))
        return response
    except Exception as e:
        logger.error(str(e))
        return str(e)


# 主机资源列表理导出下载
def server_list_down(request):
    from openpyxl.utils import get_column_letter
    from openpyxl.writer.excel import save_virtual_workbook
    # 只写模式会加快速度，不过写法和普通读写有区别，按下面一行一行插入即可
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = "Sheet1"
    columns = ("服务器名称", "服务器类型(0是物理机，1是虚拟机)", "主机名", "IP地址", "系统版本", "SN", "CPU核数", "CPU型号",
               "系统类型", "内核", "品牌名称", "ipv4列表", "mac地址列表", "内存大小(M)", "内存说明", "磁盘大小(G)", "磁盘说明",
               "minion_id", "机房名称", "机柜号", "远程访问IP", "远程访问端口", "远程访问用户", "远程访问密码", "描述备注")

    ws.append(columns)
    queryset = ServerList.objects.all()
    for obj in queryset:
        row = (obj.server_name, int(obj.server_type), obj.localhost, obj.ip, obj.system_issue, obj.sn, obj.cpu_num,
               obj.cpu_model, obj.sys_type, obj.kernel, obj.product_name, obj.ipv4_address, obj.mac_address,
               obj.mem_total, obj.mem_explain, obj.disk_total, obj.disk_explain, obj.minion_id, obj.idc_name,
               obj.idc_num, obj.login_ip, obj.login_port, obj.login_user, obj.login_password, obj.update_time, obj.description)
        ws.append(row)
    # 用(save_virtual_workbook(wb)来保存到内存中供django调用,无法使用StreamingHttpResponse或者FileResponse在openpyxl官方例子就是这样的
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path('主机列表.xlsx'))
    # wb.save(response)
    return response


# 网络设备列表
def network_list(request):
    try:
        # 这里主要是担心通过ajax提交了get请求，其实多此一举因为这个页面没有ajax的get请求
        if request.is_ajax() is not True:
            if request.method == 'GET':
                search_field = request.GET.get('search_field', '')
                search_content = request.GET.get('search_content', '')
                if search_content is '':
                    device_data = NetworkList.objects.all().order_by('create_date')
                    data_list = getPage(request, device_data, 9)
                else:
                    if search_field == 'search_device_name':
                        device_data = NetworkList.objects.filter(device_name__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 9)
                    elif search_field == 'search_manage_ip':
                        device_data = NetworkList.objects.filter(manage_ip__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 9)
                    elif search_field == 'search_device_type':
                        device_data = NetworkList.objects.filter(device_type__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 9)
                    else:
                        device_data = NetworkList.objects.filter(server_type__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 9)
                return render(request, 'network_list.html',
                              {'data_list': data_list, 'search_field': search_field, 'search_content': search_content})

    except Exception as e:
        logger.error('网络设备列表管理页面有问题', e)
        return render(request, 'network_list.html')


# 网络设备列表ajax
def network_list_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.POST.get('network_list_tag_key') == 'add_network_list':
                obj = NetworkListAddForm(request.POST)
                if obj.is_valid():
                    NetworkList.objects.create(device_name=obj.cleaned_data["device_name"],
                                               device_type=obj.cleaned_data["device_type"],
                                               manage_ip=obj.cleaned_data["manage_ip"],
                                               product_name=obj.cleaned_data["product_name"],
                                               product_type=obj.cleaned_data["product_type"],
                                               sn=obj.cleaned_data["sn"],
                                               idc_name=obj.cleaned_data["idc_name"],
                                               idc_num=obj.cleaned_data["idc_num"],
                                               login_ip=obj.cleaned_data["login_ip"],
                                               login_port=obj.cleaned_data["login_port"],
                                               login_user=obj.cleaned_data["login_user"],
                                               login_password=obj.cleaned_data["login_password"],
                                               description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('network_list_tag_key') == 'update_network_list':
                obj = NetworkListUpdateForm(request.POST)
                if obj.is_valid():
                    NetworkList.objects.filter(device_name=obj.cleaned_data["device_name"]).update(
                        device_type=obj.cleaned_data["device_type"],
                        manage_ip=obj.cleaned_data["manage_ip"],
                        product_name=obj.cleaned_data["product_name"],
                        product_type=obj.cleaned_data["product_type"],
                        sn=obj.cleaned_data["sn"],
                        idc_name=obj.cleaned_data["idc_name"],
                        idc_num=obj.cleaned_data["idc_num"],
                        login_ip=obj.cleaned_data["login_ip"],
                        login_port=obj.cleaned_data["login_port"],
                        login_user=obj.cleaned_data["login_user"],
                        login_password=obj.cleaned_data["login_password"],
                        update_time=timezone.now(),
                        description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('network_list_tag_key') == 'delete_server_list':
                device_name = request.POST.get('device_name')
                try:
                    NetworkList.objects.get(device_name=device_name).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('network_list_tag_key') == 'import_server_list':
                # 获取的数据类型如果有文件，则要像下面这么写form才能正确获取到文件
                obj = NetworkListImportForm(request.POST, request.FILES)
                if obj.is_valid():
                    # request.FILES['file'].read()是从内存中直接读取文件内容
                    # openpyxl的filename除了能打开某个路径的文件，还可以接受BytesIO读取的二进制数据所以如下即可获取xlsx内容
                    wb = openpyxl.load_workbook(filename=BytesIO(request.FILES['file'].read()))
                    table = wb[wb.sheetnames[0]]
                    network_list_cover_select = request.POST.get('network_list_cover_select')
                    # 结果需要存一些内容所以修改成列表
                    result['result'] = []
                    for row in table.iter_rows(min_row=2, max_col=25):
                        data = {'device_name': row[0].value, 'device_type': row[1].value, 'manage_ip': row[2].value,
                                'product_name': row[3].value, 'product_type': row[4].value, 'sn': row[5].value,
                                'idc_name': row[18].value, 'idc_num': row[17].value, 'login_ip': row[20].value,
                                'login_port': row[21].value, 'login_user': row[22].value,
                                'login_password': row[23].value, 'description': row[24].value}
                        obj = NetworkListAddForm(data)
                        if obj.is_valid():
                            NetworkList.objects.create(device_name=obj.cleaned_data["device_name"],
                                                       device_type=obj.cleaned_data["device_type"],
                                                       manage_ip=obj.cleaned_data["manage_ip"],
                                                       product_name=obj.cleaned_data["product_name"],
                                                       product_type=obj.cleaned_data["product_type"],
                                                       sn=obj.cleaned_data["sn"],
                                                       idc_name=obj.cleaned_data["idc_name"],
                                                       idc_num=obj.cleaned_data["idc_num"],
                                                       login_ip=obj.cleaned_data["login_ip"],
                                                       login_port=obj.cleaned_data["login_port"],
                                                       login_user=obj.cleaned_data["login_user"],
                                                       login_password=obj.cleaned_data["login_password"],
                                                       description=obj.cleaned_data["description"])
                        else:
                            error_str = json.loads(obj.errors.as_json())
                            if error_str.get('device_name') and network_list_cover_select == 'cover':
                                if error_str['device_name'][0]['message'] == '设备名称已存在，请检查':
                                    obj = NetworkListUpdateForm(data)
                                    if obj.is_valid():
                                        NetworkList.objects.filter(device_name=obj.cleaned_data["device_name"]).update(
                                            device_type=obj.cleaned_data["device_type"],
                                            manage_ip=obj.cleaned_data["manage_ip"],
                                            product_name=obj.cleaned_data["product_name"],
                                            product_type=obj.cleaned_data["product_type"],
                                            sn=obj.cleaned_data["sn"],
                                            idc_name=obj.cleaned_data["idc_name"],
                                            idc_num=obj.cleaned_data["idc_num"],
                                            login_ip=obj.cleaned_data["login_ip"],
                                            login_port=obj.cleaned_data["login_port"],
                                            login_user=obj.cleaned_data["login_user"],
                                            login_password=obj.cleaned_data["login_password"],
                                            update_time=timezone.now(),
                                            description=obj.cleaned_data["description"])
                                    else:
                                        error_str = json.loads(obj.errors.as_json())
                                        result['result'].append(error_str)
                            else:
                                if error_str['device_name'][0]['message'] == '设备名称已存在，请检查':
                                    result['result'].append('设备名称:{} 已存在'.format(data['device_name']))
                                else:
                                    result['result'].append(error_str)
                    if result['result']:
                        pass
                    else:
                        result['result'] = '成功'
                        result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            else:
                result['result'] = '网络设备管理页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('网络设备管理页ajax提交处理有问题', e)
        result['result'] = '网络设备管理页ajax提交处理有问题'
        return JsonResponse(result)


# 网络设备列表模板下载
def network_list_template_down(request):
    try:
        # 基本上django下载文件就按这个模板来就可以，更复杂的参考下面的主机管理列表导出方法
        # FileResponse方法继承了StreamingHttpResponse并且封装了迭代方法，是django最好的大文件流传送方式了，用法直接按下面，很简单
        file = settings.STATICFILES_DIRS[0] + "/download_files/网络设备模板.xlsx"
        response = FileResponse(open(file, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        # 带中文的文件名需要如下用escape_uri_path和utf8才能识别
        response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path(os.path.basename(file)))
        return response
    except Exception as e:
        logger.error(str(e))
        return str(e)


# 网络设备列表理导出下载
def network_list_down(request):
    from openpyxl.utils import get_column_letter
    from openpyxl.writer.excel import save_virtual_workbook
    # 只写模式会加快速度，不过写法和普通读写有区别，按下面一行一行插入即可
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = "Sheet1"
    columns = ("设备名称", "设备类型", "主机名", "管理IP", "设备厂家", "产品型号", "序列号",
               "机房名称", "机柜号", "远程管理IP", "远程管理端口", "远程管理用户", "远程管理密码", "描述备注")
    ws.append(columns)
    queryset = NetworkList.objects.all()
    for obj in queryset:
        row = (obj.device_name, obj.server_type, obj.manage_ip, obj.product_name, obj.product_type, obj.sn, obj.idc_name,
               obj.idc_num, obj.login_ip, obj.login_port, obj.login_user, obj.login_password, obj.update_time, obj.description)
        ws.append(row)
    # 用(save_virtual_workbook(wb)来保存到内存中供django调用,无法使用StreamingHttpResponse或者FileResponse在openpyxl官方例子就是这样的
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path('网络设备列表.xlsx'))
    # wb.save(response)
    return response


# minion管理
def minion_manage(request):
    try:
        # 这里主要是担心通过ajax提交了get请求，其实多此一举因为这个页面没有ajax的get请求哈哈
        if request.is_ajax() is not True:
            if request.method == 'GET':
                search_field = request.GET.get('search_field', '')
                search_content = request.GET.get('search_content', '')
                if search_content is '':
                    minion_data = MinionList.objects.all().order_by('create_date')
                    data_list = getPage(request, minion_data, 12)
                else:
                    if search_field == 'search_minion_id':
                        minion_data = MinionList.objects.filter(minion_id__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, minion_data, 12)
                    elif search_field == 'search_minion_sys':
                        minion_data = MinionList.objects.filter(sys__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, minion_data, 12)
                    elif search_field == 'search_minion_status':
                        minion_data = MinionList.objects.filter(minion_status__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, minion_data, 12)
                    else:
                        minion_data = MinionList.objects.filter(ip__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, minion_data, 12)
                return render(request, 'minion_manage.html',
                              {'data_list': data_list, 'search_field': search_field, 'search_content': search_content})

    except Exception as e:
        logger.error('minion管理页面有问题', e)
        return render(request, 'minion_manage.html')


# minion管理ajax
def minion_manage_ajax(request):
    result = {'result': None, 'status': False}
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.POST.get('minion_manage_key') == 'update_minion_description':
                minion_id = request.POST.get('minion_id')
                description = request.POST.get('description')
                if len(description) > 200:
                    result['result'] = '备注内容不得超过200字'
                else:
                    MinionList.objects.filter(minion_id=minion_id).update(
                        update_time=time.strftime('%Y年%m月%d日 %X'),
                        description=description)
                    result['result'] = '修改成功'
                    result['status'] = True
                return JsonResponse(result)
            elif request.POST.get('minion_manage_key') == 'update_minion_list':
                update_result = cron.minion_status()
                if update_result:
                    result['result'] = '更新成功'
                    result['status'] = True
                else:
                    result['result'] = '更新失败'
                return JsonResponse(result)
            elif request.POST.get('minion_manage_key') == 'update_minion_status':
                minion_list = MinionList.objects.values_list('minion_id', flat=True)
                id_list = []
                with requests.Session() as s:
                    saltapi = SaltAPI(session=s)
                    if saltapi.get_token() is False:
                        logger.error('minion管理更新状态操作获取SaltAPI调用get_token请求出错')
                        result['result'] = 'minion管理更新状态操作获取SaltAPI调用get_token请求出错'
                        return JsonResponse(result)
                    else:
                        # salt检测minion最准的方法salt-run manage.status
                        response_data = saltapi.saltrun_manage_status_api()
                        if response_data is False:
                            logger.error('minion管理更新状态，操作saltrun_manage_status_api调用API失败了')
                            result['result'] = 'minion管理更新状态，操作saltrun_manage_status_api调用API失败了'
                            return JsonResponse(result)
                        else:
                            status_up = response_data['return'][0]['up']
                            for minion_id in status_up:
                                updated_values = {'minion_id': minion_id, 'minion_status': '在线',
                                                  'update_time': time.strftime('%Y年%m月%d日 %X')}
                                MinionList.objects.update_or_create(minion_id=minion_id, defaults=updated_values)
                            status_down = response_data['return'][0]['down']
                            for minion_id in status_down:
                                updated_values = {'minion_id': minion_id, 'minion_status': '离线',
                                                  'update_time': time.strftime('%Y年%m月%d日 %X')}
                                MinionList.objects.update_or_create(minion_id=minion_id, defaults=updated_values)
                            id_list.extend(status_up)
                            id_list.extend(status_down)
                            for minion_id in minion_list:
                                if minion_id not in id_list:
                                    MinionList.objects.filter(minion_id=minion_id).delete()
                            result['result'] = '更新成功'
                            result['status'] = True
                return JsonResponse(result)
            else:
                result['result'] = 'minion管理页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('minion管理ajax提交处理有问题', e)
        result['result'] = 'minion管理ajax提交处理有问题'
        return JsonResponse(result)


# salt命令集
def salt_cmd_manage(request):
    try:
        if request.is_ajax() is not True:
            if request.method == 'GET':
                search_field = request.GET.get('search_field', '')
                search_content = request.GET.get('search_content', '')
                if search_content is '':
                    data = SaltCmdInfo.objects.all().order_by('salt_cmd_type','salt_cmd')
                    data_list = getPage(request, data, 12)
                else:
                    if search_field == 'any':
                        data = SaltCmdInfo.objects.filter(salt_cmd__icontains=search_content).order_by('salt_cmd_type',
                                                                                                       'salt_cmd')
                        data_list = getPage(request, data, 12)
                    elif search_field in ['module', 'state', 'runner']:
                        data = SaltCmdInfo.objects.filter(salt_cmd_type__icontains=search_field,
                                                          salt_cmd__icontains=search_content).order_by('salt_cmd_type',
                                                                                                       'salt_cmd')
                        data_list = getPage(request, data, 12)
                    else:
                        return render(request, 'salt_cmd_manage.html')
                return render(request, 'salt_cmd_manage.html',
                              {'data_list': data_list, 'search_field': search_field, 'search_content': search_content})
    except Exception as e:
        logger.error('salt命令集页面有问题', e)
        return render(request, 'salt_cmd_manage.html')


# salt命令集ajax操作
def salt_cmd_manage_ajax(request):
    result = {'result': None, 'status': False}
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.GET.get('salt_cmd_tag_key') == 'modal_search_minion_id':
                minion_id = request.GET.get('minion_id')
                minion_id_list = MinionList.objects.filter(minion_id__icontains=minion_id).order_by(
                    'create_date').values_list('minion_id', flat=True)
                result['result'] = list(minion_id_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.POST.get('salt_cmd_tag_key') == 'collection_info':
                minion_id = request.POST.get('minion_id')
                collection_style = request.POST.get('collection_style')
                try:
                    with requests.Session() as s:
                        saltapi = SaltAPI(session=s)
                        if saltapi.get_token() is False:
                            error_data = 'salt命令集采集信息获取SaltAPI调用get_token请求出错'
                            result['result'] = error_data
                            return JsonResponse(result)
                        else:
                            if collection_style == 'state':
                                response_data = saltapi.sys_state_doc_api(tgt=minion_id, tgt_type='list')
                            elif collection_style == 'runner':
                                response_data = saltapi.sys_runner_doc_api(tgt=minion_id, tgt_type='list')
                            else:
                                response_data = saltapi.sys_doc_api(tgt=minion_id, tgt_type='list')
                            # 当调用api失败的时候会返回false
                            if response_data is False:
                                error_data = 'salt命令集采集信息失败，SaltAPI调用采集api请求出错'
                                result['result'] = error_data
                                return JsonResponse(result)
                            else:
                                response_data = response_data['return'][0]
                                try:
                                    # 用来存放掉线或者访问不到的minion_id信息
                                    info = ''
                                    # state的使用帮助特殊，比如cmd.run会有一个头cmd的说明，所以要对cmd这样做一个处理把他加入到cmd.run的使用帮助中
                                    if collection_style == 'state':
                                        a = {}
                                        b = {}
                                        for min_id, cmd_dict in response_data.items():
                                            if isinstance(cmd_dict, dict):
                                                for salt_cmd, salt_cmd_doc in cmd_dict.items():
                                                    if len(salt_cmd.split('.')) == 1:
                                                        a[salt_cmd] = salt_cmd_doc
                                                    else:
                                                        b[salt_cmd] = salt_cmd_doc
                                                for salt_cmd in b.keys():
                                                    try:
                                                        b[salt_cmd] = salt_cmd.split('.')[0] + ':\n' + str(
                                                            a[salt_cmd.split('.')[0]]).replace('\n',
                                                                                           '\n    ') + '\n\n' + salt_cmd + ':\n' + str(
                                                            b[salt_cmd])
                                                    except Exception as e:
                                                        logger.error('state采集后台错误：' + str(e))
                                                        result['result'] = 'state采集后台错误：' + str(e)
                                                        return JsonResponse(result)
                                                    updated_values = {'salt_cmd': salt_cmd, 'salt_cmd_type': collection_style,
                                                                      'salt_cmd_module': salt_cmd.split('.')[0],
                                                                      'salt_cmd_source': 'minion', 'salt_cmd_doc': b[salt_cmd],
                                                                      'update_time': time.strftime('%Y年%m月%d日 %X')}
                                                    SaltCmdInfo.objects.update_or_create(salt_cmd=salt_cmd, salt_cmd_type=collection_style, defaults=updated_values)
                                            elif isinstance(cmd_dict, bool):
                                                info += ' 不过minion_id:' + min_id + '掉线了没有从它采集到数据'
                                        result['result'] = '采集完成' + info
                                        result['status'] = True
                                        return JsonResponse(result)
                                    else:
                                        for min_id, cmd_dict in response_data.items():
                                            if isinstance(cmd_dict, dict):
                                                for salt_cmd, salt_cmd_doc in cmd_dict.items():
                                                    salt_cmd_doc = str(salt_cmd) + ':\n' + str(salt_cmd_doc)
                                                    updated_values = {'salt_cmd': salt_cmd, 'salt_cmd_type': collection_style,
                                                                      'salt_cmd_module': salt_cmd.split('.')[0],
                                                                      'salt_cmd_source': 'minion', 'salt_cmd_doc': salt_cmd_doc,
                                                                      'update_time': time.strftime('%Y年%m月%d日 %X')}
                                                    SaltCmdInfo.objects.update_or_create(salt_cmd=salt_cmd, salt_cmd_type=collection_style, defaults=updated_values)
                                            elif isinstance(cmd_dict, bool):
                                                info += ' 不过minion_id:' + min_id + '掉线了没有从它采集到数据'
                                        result['result'] = '采集完成' + info
                                        result['status'] = True
                                        return JsonResponse(result)
                                except Exception as e:
                                    logger.error('采集后台错误：' + str(e))
                                    result['result'] = '采集后台错误：' + str(e)
                                    return JsonResponse(result)
                except Exception as e:
                    logger.error('采集信息出错：'+str(e))
                    result['result'] = '采集信息出错'
                    return JsonResponse(result)
            elif request.POST.get('salt_cmd_tag_key') == 'update_salt_cmd_description':
                salt_cmd = request.POST.get('salt_cmd')
                description = request.POST.get('description')
                try:
                    SaltCmdInfo.objects.filter(salt_cmd=salt_cmd).update(update_time=time.strftime('%Y年%m月%d日 %X'),
                                                                         description=description)
                    result['result'] = '修改成功'
                    result['status'] = True
                except Exception as e:
                    message = '修改失败', str(e)
                    logger.error(message)
                    result['result'] = message
                return JsonResponse(result)
            elif request.POST.get('salt_cmd_tag_key') == 'salt_cmd_delete':
                salt_cmd = request.POST.get('salt_cmd')
                try:
                    SaltCmdInfo.objects.filter(salt_cmd=salt_cmd).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    message = '修改失败', str(e)
                    logger.error(message)
                    result['result'] = message
                return JsonResponse(result)
    except Exception as e:
        logger.error('salt命令集ajax提交处理有问题', e)
        result['result'] = 'salt命令集ajax提交处理有问题'
        return JsonResponse(result)


# SaltKey管理
def saltkey_manage(request):
    try:
        if request.method == 'GET':
            accepted_count = SaltKeyList.objects.filter(certification_status='accepted').count()
            unaccepted_count = SaltKeyList.objects.filter(certification_status='unaccepted').count()
            denied_count = SaltKeyList.objects.filter(certification_status='denied').count()
            rejected_count = SaltKeyList.objects.filter(certification_status='rejected').count()
            if request.GET.get('status') is None:
                return HttpResponseRedirect('/saltkey_manage/?status=accepted&search=')
            if request.GET.get('status') == 'accepted':
                if request.GET.get('search').strip() is "":
                    accepted_data = SaltKeyList.objects.filter(certification_status='accepted')
                    data_list = getPage(request, accepted_data, 8)
                    return render(request, 'saltkey_manage.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': ""})
                else:
                    search_data = request.GET.get('search').strip()
                    accepted_data = SaltKeyList.objects.filter(minion_id__icontains=search_data, certification_status='accepted')
                    data_list = getPage(request, accepted_data, 8)
                    return render(request, 'saltkey_manage.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': search_data})
            elif request.GET.get('status') == 'unaccepted':
                if request.GET.get('search').strip() is "":
                    unaccepted_data = SaltKeyList.objects.filter(certification_status='unaccepted')
                    data_list = getPage(request, unaccepted_data, 8)
                    return render(request, 'saltkey_manage_unaccepted.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': ""})
                else:
                    search_data = request.GET.get('search').strip()
                    unaccepted_data = SaltKeyList.objects.filter(minion_id__icontains=search_data, certification_status='unaccepted')
                    data_list = getPage(request, unaccepted_data, 8)
                    return render(request, 'saltkey_manage_unaccepted.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': search_data})
            elif request.GET.get('status') == 'denied':
                if request.GET.get('search').strip() is "":
                    denied_data = SaltKeyList.objects.filter(certification_status='denied')
                    data_list = getPage(request, denied_data, 8)
                    return render(request, 'saltkey_manage_denied.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': ""})
                else:
                    search_data = request.GET.get('search').strip()
                    denied_data = SaltKeyList.objects.filter(minion_id__icontains=search_data, certification_status='denied')
                    data_list = getPage(request, denied_data, 8)
                    return render(request, 'saltkey_manage_denied.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': search_data})
            elif request.GET.get('status') == 'rejected':
                if request.GET.get('search').strip() is "":
                    rejected_data = SaltKeyList.objects.filter(certification_status='rejected')
                    data_list = getPage(request, rejected_data, 8)
                    return render(request, 'saltkey_manage_rejected.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': ""})
                else:
                    search_data = request.GET.get('search').strip()
                    rejected_data = SaltKeyList.objects.filter(minion_id__icontains=search_data, certification_status='rejected')
                    data_list = getPage(request, rejected_data, 8)
                    return render(request, 'saltkey_manage_rejected.html',
                                  {'data_list': data_list, 'accepted_count': accepted_count,
                                   'unaccepted_count': unaccepted_count, 'denied_count': denied_count,
                                   'rejected_count': rejected_count, 'search': search_data})

    except Exception as e:
        logger.error('SaltKey管理页面有问题', e)
        return render(request, 'saltkey_manage.html')


# SaltKey全局操作
def salt_key_global(request):
    try:
        if request.is_ajax():
            if 'global_flush_salt_key' in request.POST:
                if cron.saltkey_list():
                    return JsonResponse({'result': '操作成功', 'status': True})
                else:
                    return JsonResponse({'result': '操作失败', 'status': False})
    except Exception as e:
        logger.error('全局操作出错了' + str(e))
        response_data = {'result': '操作失败', 'status': False}
        return JsonResponse(response_data)


# salt的test.ping方法
def salt_test_ping(request):
    try:
        if request.is_ajax():
            minion_id = request.POST.get('minion_id')
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                if saltapi.get_token() is False:
                    logger.error('test.ping操作获取SaltAPI调用get_token请求出错')
                    response_data = {'result': '检测失败', 'status': False}
                    return JsonResponse(response_data)
                else:
                    response_data = saltapi.test_api(tgt=minion_id)
                    # 当调用api失败的时候比如salt-api服务stop了会返回false
                    if response_data is False:
                        logger.error('test.ping失败可能代入的参数有问题，SaltAPI调用test_api请求出错')
                        response_data = {'result': '检测失败', 'status': False}
                        return JsonResponse(response_data)
                    # 判断返回值如果为[{}]表明没有这个minion_id
                    elif response_data['return'] != [{}]:
                        # 正常结果类似这样：{'return': [{'192.168.68.51': False, '192.168.68.1': True, '192.168.68.50-master': True}]}
                        data_source = response_data['return'][0]
                        response_data = {'result': data_source, 'status': True}
                        return JsonResponse(response_data)
                    else:
                        logger.error('test.ping检测失败，请确认minion是否存在。。')
                        response_data = {'result': '检测失败', 'status': False}
                        return JsonResponse(response_data)
    except Exception as e:
        logger.error('test.ping检测出错了' + str(e))
        response_data = {'result': '检测失败', 'status': False}
        return JsonResponse(response_data)


# salt的删除key方法
def salt_key_delete(request):
    try:
        if request.is_ajax():
            minion_id = request.POST.getlist('minion_id')
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                if saltapi.get_token() is False:
                    logger.error('删除key操作获取SaltAPI调用get_token请求出错')
                    response_data = {'result': '删除失败', 'status': False}
                    return JsonResponse(response_data)
                else:
                    response_data = saltapi.saltkey_delete_api(match=minion_id)
                    # 当调用api失败的时候会返回false
                    if response_data is False:
                        logger.error('删除key失败可能代入的参数有问题，SaltAPI调用saltkey_delete_api请求出错')
                        response_data = {'result': '删除失败', 'status': False}
                        return JsonResponse(response_data)
                    # 返回值不会有为[{}]的情况所以没出现api失败就表示成功，这是这个delete的api接口的坑奶奶的
                    else:
                        if cron.saltkey_list():
                            response_data = {'result': '删除成功', 'status': True}
                            return JsonResponse(response_data)
                        else:
                            logger.error('删除key在执行刷新saltkey操作即cron.py里的方法时候出错了')
                            response_data = {'result': '删除失败', 'status': False}
                            return JsonResponse(response_data)
    except Exception as e:
        logger.error('删除key出错了' + str(e))
        response_data = {'result': '删除失败', 'status': False}
        return JsonResponse(response_data)


# salt的删除denied下的key方法
def salt_key_denied_delete(request):
    try:
        if request.is_ajax():
            # 默认获取列表主要是为了后面做批量删除可以复用这个方法
            minion_id = request.POST.getlist('minion_id')
            # 因为在linux中批量删除文件要用{1.txt,2.txt,a.txt}或者1.txt 2.txt a.txt空格隔开即可所以要格式化下把上面列表转成这样子
            # 本来我是要用{1.txt,2.txt,a.txt}但是在遇到单个文件的时候{1.txt}不行{1.txt,}可以但是删除的含义不同了会导致master挂掉草
            # 所以最终使用进入到minions_denied文件夹然后1.txt 2.txt a.txt空格隔开每个文件删除的方式删除
            minion_id = ' '.join(minion_id)
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                if saltapi.get_token() is False:
                    logger.error('删除key操作获取SaltAPI调用get_token请求出错')
                    response_data = {'result': '删除失败', 'status': False}
                    return JsonResponse(response_data)
                else:
                    response_data = saltapi.cmd_run_api(tgt=settings.SITE_SALT_MASTER, arg='cd /etc/salt/pki/master/minions_denied/ && rm -rf %s' % minion_id)
                    # 当调用api失败的时候会返回false
                    if response_data is False:
                        logger.error('删除denied的key失败可能代入的参数有问题，SaltAPI调用cmd.run请求出错')
                        response_data = {'result': '删除失败', 'status': False}
                        return JsonResponse(response_data)
                    # 命令rm删除返回值为空，所以return值是[{}]这个值不是空哟所以没出现api失败就表示成功
                    else:
                        if cron.saltkey_list():
                            response_data = {'result': '删除成功', 'status': True}
                            return JsonResponse(response_data)
                        else:
                            logger.error('删除denied的key在执行刷新saltkey操作即cron.py里的方法时候出错了')
                            response_data = {'result': '删除失败', 'status': False}
                            return JsonResponse(response_data)
    except Exception as e:
        logger.error('删除denied的key出错了' + str(e))
        response_data = {'result': '删除失败', 'status': False}
        return JsonResponse(response_data)


# 接受salt-key方法saltkey_accept_api
def salt_key_accept(request):
    try:
        if request.is_ajax():
            minion_id = request.POST.getlist('minion_id')
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                if saltapi.get_token() is False:
                    logger.error('接受key操作获取SaltAPI调用get_token请求出错')
                    response_data = {'result': '删除失败', 'status': False}
                    return JsonResponse(response_data)
                else:
                    response_data = saltapi.saltkey_accept_api(match=minion_id)
                    # 当调用api失败的时候会返回false
                    if response_data is False:
                        logger.error('接受key失败可能代入的参数有问题，SaltAPI调用saltkey_accept_api请求出错')
                        response_data = {'result': '接受失败', 'status': False}
                        return JsonResponse(response_data)
                    # 返回值不会有为[{}]的情况所以没出现api失败就表示成功，这是这个delete的api接口的坑奶奶的
                    else:
                        if cron.saltkey_list():
                            response_data = {'result': '接受成功', 'status': True}
                            return JsonResponse(response_data)
                        else:
                            logger.error('接受key在执行刷新saltkey操作即cron.py里的方法时候出错了')
                            response_data = {'result': '接受失败', 'status': False}
                            return JsonResponse(response_data)
    except Exception as e:
        logger.error('接受key出错了' + str(e))
        response_data = {'result': '接受失败', 'status': False}
        return JsonResponse(response_data)


# 拒绝salt-key方法saltkey_accept_api
def salt_key_reject(request):
    try:
        if request.is_ajax():
            minion_id = request.POST.getlist('minion_id')
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                if saltapi.get_token() is False:
                    logger.error('拒绝key操作获取SaltAPI调用get_token请求出错')
                    response_data = {'result': '拒绝失败', 'status': False}
                    return JsonResponse(response_data)
                else:
                    response_data = saltapi.saltkey_reject_api(match=minion_id)
                    # 当调用api失败的时候会返回false
                    if response_data is False:
                        logger.error('拒绝key失败可能代入的参数有问题，SaltAPI调用saltkey_reject_api请求出错')
                        response_data = {'result': '拒绝失败', 'status': False}
                        return JsonResponse(response_data)
                    # 返回值不会有为[{}]的情况所以没出现api失败就表示成功，这是这个delete的api接口的坑奶奶的
                    else:
                        if cron.saltkey_list():
                            response_data = {'result': '拒绝成功', 'status': True}
                            return JsonResponse(response_data)
                        else:
                            logger.error('拒绝key在执行刷新saltkey操作即cron.py里的方法时候出错了')
                            response_data = {'result': '拒绝失败', 'status': False}
                            return JsonResponse(response_data)
    except Exception as e:
        logger.error('拒绝key出错了' + str(e))
        response_data = {'result': '拒绝失败', 'status': False}
        return JsonResponse(response_data)


# salt_exe执行salt命令页
def salt_exe(request):
    try:
        if request.method == 'GET' and not request.is_ajax():
            data_list = SaltCmdInfo.objects.filter(salt_cmd_type='module').values('salt_cmd_module').distinct().order_by('salt_cmd_module')
            return render(request, 'salt_exe.html', {'data_list': data_list})
    except Exception as e:
        logger.error('salt命令执行页面有问题：', e)
        return render(request, 'salt_exe.html')


# salt_exe_ajax执行salt命令ajax操作
def salt_exe_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.GET.get('salt_exe_tag_key') == 'modal_search_minion_id':
                minion_id = request.GET.get('minion_id')
                minion_id_list = MinionList.objects.filter(minion_id__icontains=minion_id).order_by(
                    'create_date').values_list('minion_id', flat=True)
                result['result'] = list(minion_id_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.GET.get('salt_exe_tag_key') == 'search_salt_module':
                salt_cmd_type = request.GET.get('salt_cmd_type')
                salt_cmd_list = SaltCmdInfo.objects.filter(salt_cmd_type=salt_cmd_type).values_list(
                    'salt_cmd_module', flat=True).distinct().order_by('salt_cmd_module')
                result['result'] = list(salt_cmd_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.GET.get('salt_exe_tag_key') == 'search_salt_cmd':
                salt_cmd_type = request.GET.get('salt_cmd_type')
                salt_cmd_module = request.GET.get('salt_cmd_module')
                salt_cmd_list = SaltCmdInfo.objects.filter(salt_cmd_type=salt_cmd_type,
                                                           salt_cmd_module=salt_cmd_module).values_list('salt_cmd',
                                                                                                        flat=True)
                result['result'] = list(salt_cmd_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.GET.get('salt_exe_tag_key') == 'search_salt_cmd_doc':
                salt_cmd = request.GET.get('salt_cmd')
                salt_cmd_type = request.GET.get('salt_cmd_type')
                salt_cmd_data = SaltCmdInfo.objects.filter(salt_cmd=salt_cmd, salt_cmd_type=salt_cmd_type).first()
                result['result'] = salt_cmd_data.salt_cmd_doc if salt_cmd_data else '查询结果为空，请确认模块和命令是否填写正确'
                result['status'] = True
                return JsonResponse(result)
            elif request.POST.get('salt_exe_tag_key') == 'salt_exe':
                client = request.POST.get('client')
                tgt = request.POST.get('tgt')
                tgt_type = request.POST.get('tgt_type')
                fun = request.POST.get('fun')
                arg = request.POST.getlist('arg')
                # 这是判断arg是否传输值过来，如果没有前端会传个['']过来，这是由于我前端设置了的
                if arg == ['']:
                    arg = None
                if tgt_type == 'list':
                    tgt = [tgt]
                if client != 'runner':
                    data = {'client': client, 'tgt': tgt, 'tgt_type': tgt_type, 'fun': fun, 'arg': arg}
                else:
                    data = {'client': client, 'fun': fun, 'arg': arg}
                with requests.Session() as s:
                    saltapi = SaltAPI(session=s)
                    if saltapi.get_token() is False:
                        app_log.append('\nsalt命令执行后台出错_error(0)，请联系管理员')
                        result['result'] = app_log
                        return JsonResponse(result)
                    else:
                        response_data = saltapi.public(data=data)
                        # 当调用api失败的时候会返回false
                        if response_data is False:
                            app_log.append('\nsalt命令执行后台出错_error(1)，请联系管理员')
                            result['result'] = app_log
                            return JsonResponse(result)
                        else:
                            try:
                                response_data = response_data['return'][0]
                                result['status'] = True
                                result['result'] = response_data
                                return JsonResponse(result)
                            except Exception as e:
                                app_log.append('\n' + 'salt命令执行失败_error(2):' + str(response_data))
                                result['result'] = app_log
                                return JsonResponse(result)
            elif request.GET.get('salt_exe_tag_key') == 'search_jid_status':
                jid = request.GET.get('jid')
                with requests.Session() as s:
                    saltapi = SaltAPI(session=s)
                    if saltapi.get_token() is False:
                        app_log.append('\nsalt命令执行查询jid后台出错_error(0)，请联系管理员')
                        result['result'] = app_log
                        return JsonResponse(result)
                    else:
                        response_data = saltapi.job_exit_success_api(jid=jid)
                        # 当调用api失败的时候会返回false
                        if response_data is False:
                            app_log.append('\nsalt命令执行查询jid后台出错_error(1)，请联系管理员')
                            result['result'] = app_log
                            return JsonResponse(result)
                        else:
                            try:
                                response_data = response_data['return'][0]
                                result['status'] = True
                                result['result'] = response_data
                                return JsonResponse(result)
                            except Exception as e:
                                app_log.append('\n' + 'salt命令执行执行查询jid失败_error(2):' + str(response_data))
                                result['result'] = app_log
                                return JsonResponse(result)
            elif request.GET.get('salt_exe_tag_key') == 'search_jid_result':
                jid = request.GET.get('jid')
                with requests.Session() as s:
                    saltapi = SaltAPI(session=s)
                    if saltapi.get_token() is False:
                        app_log.append('\nsalt命令执行查询job结果后台出错_error(0)，请联系管理员')
                        result['result'] = app_log
                        return JsonResponse(result)
                    else:
                        response_data = saltapi.jid_api(jid=jid)
                        # 当调用api失败的时候会返回false
                        if response_data is False:
                            app_log.append('\nsalt命令执行查询job结果后台出错_error(1)，请联系管理员')
                            result['result'] = app_log
                            return JsonResponse(result)
                        else:
                            try:
                                response_data = response_data['return'][0]
                                result['status'] = True
                                result['result'] = response_data
                                return JsonResponse(result)
                            except Exception as e:
                                app_log.append('\n' + 'salt命令执行执行查询job结果失败_error(2):' + str(response_data))
                                result['result'] = app_log
                                return JsonResponse(result)
    except Exception as e:
        logger.error('salt命令执行ajax提交处理有问题', e)
        result['result'] = 'salt命令执行ajax提交处理有问题'
        return JsonResponse(result)


# salt_tool执行salt各种封装好的命令
def salt_tool(request):
    try:
        if request.method == 'GET' and not request.is_ajax():
            return render(request, 'salt_tool.html')
    except Exception as e:
        logger.error('salt命令集页面有问题', e)
        return render(request, 'salt_tool.html')


# salt_tool执行salt各种封装好的命令ajax操作
def salt_tool_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.GET.get('salt_tool_tag_key') == 'modal_search_minion_id':
                minion_id = request.GET.get('minion_id')
                minion_id_list = MinionList.objects.filter(minion_id__icontains=minion_id).order_by(
                    'create_date').values_list('minion_id', flat=True)
                result['result'] = list(minion_id_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.POST.get('salt_tool_tag_key') == 'create_task':
                tgt = request.POST.get('tgt')
                tgt_type = request.POST.get('tgt_type')
                task_name = request.POST.get('task_name')
                username = request.POST.get('username')
                cmd = request.POST.get('cmd')
                force = request.POST.get('force')
                start_in = request.POST.get('start_in')
                task_arg = request.POST.get('task_arg')
                arg = []
                arg.extend([task_name, 'action_type=Execute', 'user_name=%s' % username, 'cmd=%s' % cmd, 'force=%s' % force, 'execution_time_limit=False', 'trigger_type=OnBoot'])
                # 这是判断arg是否传输值过来，如果没有前端会传个['']过来，这是由于我前端设置了的
                if start_in != '':
                    arg.append('start_in=%s' % start_in)
                if task_arg != '':
                    arg.append('arguments=%s' % task_arg)
                with requests.Session() as s:
                    saltapi = SaltAPI(session=s)
                    if saltapi.get_token() is False:
                        app_log.append('\nwindows创建计划任务后台出错_error(0)，请联系管理员')
                        result['result'] = app_log
                        return JsonResponse(result)
                    else:
                        response_data = saltapi.task_create_api(tgt=tgt, tgt_type=tgt_type, arg=arg)
                        # 当调用api失败的时候会返回false
                        if response_data is False:
                            app_log.append('\nwindows创建计划任务后台出错_error(1)，请联系管理员')
                            result['result'] = app_log
                            return JsonResponse(result)
                        else:
                            try:
                                response_data = response_data['return'][0]
                                result['status'] = True
                                result['result'] = response_data
                                return JsonResponse(result)
                            except Exception as e:
                                app_log.append('\n' + 'windows创建计划任务失败_error(2):' + str(response_data))
                                result['result'] = app_log
                                return JsonResponse(result)
            elif request.GET.get('salt_tool_tag_key') == 'get_supervisor_content':
                program = request.GET.get('program')
                command = request.GET.get('command')
                directory = request.GET.get('directory')
                env = request.GET.get('env')
                logfile = request.GET.get('logfile')
                errorlogfile = request.GET.get('errorlogfile')
                arg = []
                arg.extend(['[program:%s]' % program, 'command=%s' % command, 'stdout_logfile_maxbytes=10MB',
                            'stderr_logfile_maxbytes=10MB', 'stdout_logfile_backups=5', 'stderr_logfile_backups=5',
                            'startsecs=5', 'stopsignal=QUIT', 'stopasgroup=true', 'killasgroup=true'])
                # 这是判断arg是否传输值过来，如果没有前端会传个['']过来，这是由于我前端设置了的
                if logfile != '':
                    arg.append('stdout_logfile=%s' % logfile)
                else:
                    logfile = 'stdout_logfile=/var/log/supervisor/%s.log' % program
                    arg.append(logfile)
                if errorlogfile != '':
                    arg.append('stderr_logfile=%s' % errorlogfile)
                else:
                    errorlogfile = 'stderr_logfile=/var/log/supervisor/%s.log' % program
                    arg.append(errorlogfile)
                if env != '':
                    arg.append('environment=%s' % env)
                if directory != '':
                    arg.append('directory=%s' % directory)
                result['status'] = True
                result['result'] = arg
                return JsonResponse(result)
            elif request.POST.get('salt_tool_tag_key') == 'create_supervisor':
                tgt = request.POST.get('tgt')
                tgt_type = request.POST.get('tgt_type')
                program = request.POST.get('program')
                command = request.POST.get('command')
                directory = request.POST.get('directory')
                env = request.POST.get('env')
                logfile = request.POST.get('logfile')
                errorlogfile = request.POST.get('errorlogfile')
                force = request.POST.get('force')
                # path文件存放的目录必须是supervisord.conf的include定义好的，所以无法自定义，只能内置
                path = '/etc/supervisord.d/'
                # file_path是文件路径加文件名后缀是.ini，这个也是在supervisord.conf的include定义好的格式
                file_path = path + program + '.ini'
                arg = []
                arg.extend(['[program:%s]' % program, 'command=%s' % command,'stdout_logfile_maxbytes=10MB',
                            'stderr_logfile_maxbytes=10MB', 'stdout_logfile_backups=5', 'stderr_logfile_backups=5',
                            'startsecs=5', 'stopsignal=QUIT', 'stopasgroup=true', 'killasgroup=true'])
                # 这是判断arg是否传输值过来，如果没有前端会传个['']过来，这是由于我前端设置了的
                if logfile != '':
                    logfile_path = logfile
                    arg.append('stdout_logfile=%s' % logfile_path)
                else:
                    logfile_path = '/var/log/supervisor/%s.log' % program
                    arg.append('stdout_logfile=%s' % logfile_path)
                if errorlogfile != '':
                    errorlogfile_path = errorlogfile
                    arg.append('stderr_logfile=%s' % errorlogfile_path)
                else:
                    errorlogfile_path = '/var/log/supervisor/%s.log' % program
                    arg.append('stderr_logfile=%s' % errorlogfile_path)
                if env != '':
                    arg.append('environment=%s' % env)
                if directory != '':
                    arg.append('directory=%s' % directory)
                with requests.Session() as s:
                    saltapi = SaltAPI(session=s)
                    if saltapi.get_token() is False:
                        app_log.append('\n创建supervisor进程后台出错_error(0)，请联系管理员')
                        result['result'] = app_log
                        return JsonResponse(result)
                    else:
                        response_data = saltapi.file_exists_api(tgt=tgt, tgt_type=tgt_type, arg=file_path)
                        # 当调用api失败的时候会返回false
                        if response_data is False:
                            app_log.append('\n创建supervisor进程后台出错_error(1)，请联系管理员')
                            result['result'] = app_log
                            return JsonResponse(result)
                        else:
                            try:
                                response_data = response_data['return'][0]
                                success_tgt = []
                                for k, v in response_data.items():
                                    # 判断如果执行的minion中是否已经存在了该文件夹，如果存在并且强制没有选中True就不继续执行这个minion的下一步操作
                                    if v is True and force == 'false':
                                        app_log.append({k: 'minion_id:%s已存在同名文件，如需覆盖请选择强制' % k})
                                    else:
                                        success_tgt.append(k)
                                if success_tgt:
                                    success_tgt = ','.join(success_tgt)
                                    response_data = saltapi.file_write_api(tgt=success_tgt, tgt_type='list', arg=[file_path, 'args=%s' % arg])
                                    if response_data is False:
                                        app_log.append('\n创建supervisor进程后台出错_error(2)，请联系管理员')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    else:
                                        response_data = response_data['return']
                                        app_log.extend(response_data)
                                    # 日志文件目录创建
                                    response_data = saltapi.file_makedirs_api(tgt=success_tgt, tgt_type='list',
                                                                              arg=logfile_path)
                                    if response_data is False:
                                        app_log.append('\n创建supervisor进程日志创建出错_error(3)，请联系管理员')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    response_data = saltapi.file_makedirs_api(tgt=success_tgt, tgt_type='list',
                                                                              arg=errorlogfile_path)
                                    if response_data is False:
                                        app_log.append('\n创建supervisor进程日志创建出错_error(4)，请联系管理员')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    # 重载supervisord配置
                                    response_data = saltapi.supervisord_update_api(tgt=success_tgt, tgt_type='list')
                                    if response_data is False:
                                        app_log.append('\n创建supervisor进程重载配置出错_error(5)，请联系管理员')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                result['status'] = True
                                result['result'] = app_log
                                return JsonResponse(result)
                            except Exception as e:
                                app_log.append('\n' + '创建supervisor进程失败_error(3):' + str(response_data))
                                result['result'] = app_log
                                return JsonResponse(result)
            else:
                result['result'] = 'salt工具页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('salt命令执行ajax提交处理有问题', e)
        result['result'] = 'salt命令执行ajax提交处理有问题'
        return JsonResponse(result)

# #nginx添加功能
# def nginx_add(request):
#     pass


# nginx管理
def nginx_manage(request):
    try:
        if request.method == 'GET':
            nginx_list = NginxManage.objects.all()
            return render(request, 'nginx_manage.html', {'nginx_list': nginx_list})
        else:
            nginx_ip = request.POST.get('nginx_ip')
            nginx_vip = request.POST.get('nginx_vip')
            nginx_path = request.POST.get('nginx_path')
            nginxconf_path = request.POST.get('nginxconf_path')
            nginxvhosts_path = request.POST.get('nginxvhosts_path')
            nginxlogs_path = request.POST.get('nginxlogs_path')
            # 判断nginx管理表是否已经存在此IP了，存在就不添加
            if NginxManage.objects.filter(ip=nginx_ip):
                response_data = {'result': 'IP:%s已存在nginx管理中' % nginx_ip, 'status': False}
                return JsonResponse(response_data)
            # 判断ip是否在minion列表里如果不在返回拒绝添加，因为必须用salt来控制，这里还有一个点就是minion的状态离线问题不知道要不要判断
            if MinionList.objects.filter(ip=nginx_ip):
                minion_id = MinionList.objects.get(ip=nginx_ip).minion_id
                NginxManage.objects.create(ip=nginx_ip, vip=nginx_vip, path=nginx_path, conf_path=nginxconf_path,
                                           vhost_path=nginxvhosts_path, logs_path=nginxlogs_path, update_time=time.strftime('%Y-%m-%d %X'),
                                           minion_id=minion_id)
                response_data = {'result': '新增nginx成功', 'status': True}
                return JsonResponse(response_data)
            else:
                response_data = {'result': 'IP:%s未安装salt-minion，请安装salt-minion并加入SaltKey' % nginx_ip, 'status': False}
                return JsonResponse(response_data)
    except Exception as e:
        logger.error('nginx管理问题', e)


# nginx配置文件列表
def nginx_conflist(request):
    try:
        if request.method == 'GET':
            minionid = request.GET.get('minionid')
            nginxconfpaht = request.GET.get('confpath')
            nginxvhostpaht = request.GET.get('vhostspath')
            nginxip = request.GET.get('nginxip')
            with requests.Session() as s:
                try:
                    token = s.post('http://192.168.68.50:8080/login',
                                   json={'username': 'saltapi', 'password': '123456', 'eauth': 'pam',})
                    token.raise_for_status()
                except Exception as e:
                    print('nginx配置文件列表获取token报错：', e)
                    return render(request, 'nginx_conflist.html')
                else:
                    data = {'client': 'local',
                            'tgt': minionid,
                            'expr_form': 'glob',
                            'fun': 'cmd.run',
                            'arg': 'ls -A %s/*.conf %s/*.conf %s/.*.swp %s/.*.swp' % (nginxconfpaht, nginxvhostpaht, nginxconfpaht, nginxvhostpaht),
                            }
                    try:
                        response_data = s.post('http://192.168.68.50:8080', data=data)
                        response_data.raise_for_status()
                    except Exception as e:
                        print(e)
                        # logger.error(e)
                        # return logger.info(e + ' 无法获取数据')
                        return render(request, 'nginx_conflist.html')
                    else:
                        data = response_data.json()['return'][0][minionid]
                        datalist = data.split('\n')
                        # 获得配置文件列表（带绝对路径的），并且把没有查找到的排除掉No such file or directory
                        datalist = [x for x in datalist if 'No such file or directory' not in x]
                        # 筛选出conf文件排除掉swp文件
                        conf_list = [x for x in datalist if x[-4:] != '.swp']
                        # 因为conf文件有可能正在被人编辑，那么我们就无法操作不然会冲突，所以需要判断swp是否存在，swp就是在编辑时候的缓存文件
                        swp_list = [x for x in datalist if x[-4:] == '.swp']
                        conflist = []
                        # 循环conf文件列表判断是否有同名swp文件，如果有就在4元祖最后一个内容改成'正在编辑中'
                        for x in conf_list:
                            for y in swp_list:
                                if x.rsplit('/', 1)[1] == y.rsplit('/', 1)[1][1:-4]:
                                    conflist.append((x, x.rsplit('/', 1)[1], x.rsplit('/', 1)[0], '正在被编辑中'))
                                    break
                            else:
                                # 获得配置文件名，3元祖为了好给前端直接使用，第一个是类似/tmp/nginx.conf，第二个是nginx.conf，三个是/tmp,第四个是判断是否在编辑
                                conflist.append((x, x.rsplit('/', 1)[1], x.rsplit('/', 1)[0], '正常'))

                        # conflist = [(x, x.rsplit('/', 1)[1], x.rsplit('/', 1)[0]) for x in datalist]
                        return render(request, 'nginx_conflist.html', {'conflist': conflist, 'minionid':minionid,'nginxip':nginxip})
    except Exception as e:
        logger.error('miniion管理页面有问题', e)
        return render(request, 'nginx_conflist.html')


# 做一个nginx切换负载的小功能，作为nginx管理后台的起步
def nginx_upstream(request):
    reg1 = re.compile(r'\nupstream[^\}]+}')
    reg2 = re.compile(r'\nupstream[^\{]+')
    reg3 = re.compile(r'\#* *server \d[^;]+;')
    reg4 = re.compile(r'\#+ *')
    upstreamlist = []
    try:
        if request.method == 'GET':
            minionid = request.GET.get('minionid')
            nginxip = request.GET.get('nginxip')
            path = request.GET.get('path')
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                if saltapi.get_token() is False:
                    logger.error('nginx_upstream页get操作获取SaltAPI调用get_token请求出错')
                    return render(request, 'nginx_upstream.html')
                else:
                    response_data = saltapi.cmd_run_api(tgt=minionid, arg='cat %s' % path)
                    if response_data is False:
                        logger.error('获取upstream列表失败可能代入的参数有问题，SaltAPI调用cmd_run_api请求出错')
                        return render(request, 'nginx_upstream.html')
                    # 判断upstream_data如果返回值如果为[{}]表明没有这个minionid
                    elif response_data['return'] != [{}]:
                        data_source = response_data['return'][0][minionid]
                        data_list = re.findall(reg1, data_source)
                        for i in data_list:
                            # 获取upstream name
                            b2 = re.search(reg2, i)
                            # 获取upstream server列表
                            b3 = re.findall(reg3, i)
                            # 用空格切割字符串取第二个就是servername了
                            namekey = b2.group().split(' ')[1]
                            # 下面这个如果直接赋值b3会有一些问题，就是出现'##  server'这样的也会被前端输出，所以用了
                            # 正则把这种出现'###  '是全部替换成#这样不仅显示正常，在下面post中获取的前端upstream_server也会正确，很重要
                            upstreamlist.append([namekey, [re.sub(reg4, '#', x.strip()) for x in b3]])
                        return render(request, 'nginx_upstream.html',
                                      {'upstreamlist': upstreamlist, 'minionid': minionid, 'nginxip': nginxip, 'path': path})
                    else:
                        logger.error('获取upstream列表失败，请确认minion是否存在。。')
                        return render(request, 'nginx_upstream.html')
        # 切换里面server的up/down状态，其实就是注释加#和不注释而已
        else:
            upstream_name = request.POST.get('upstream_name')
            upstream_server = request.POST.get('upstream_server')
            upstream_status = request.POST.get('upstream_status')
            minionid = request.POST.get('minionid')
            path = request.POST.get('path')
            with requests.Session() as s:
                saltapi = SaltAPI(session=s)
                if saltapi.get_token() is False:
                    logger.error('nginx_upstream页状态变更操作获取SaltAPI调用get_token请求出错')
                    return JsonResponse({'result': '失败,后台问题', 'status': False})
                else:
                    if upstream_status == 'down':
                        arg = "sed -i '/^upstream *%s/,/^}$/{s/%s/#%s/g}' %s&&free -m" % (upstream_name, upstream_server, upstream_server, path)
                    else:
                        arg = "sed -i '/^upstream *%s/,/^}$/{s/#\+ *%s/%s/g}' %s&&free -m" % (upstream_name, upstream_server, upstream_server, path)
                    response_data = saltapi.cmd_run_api(tgt=minionid, arg=arg)
                    if response_data is False:
                        logger.error('nginx_upstream页状态变更SaltAPI调用cmd_run_api请求出错')
                        return JsonResponse({'result': '失败,后台问题', 'status': False})
                    # 返回值如果为[{}]表明没有这个minionid
                    elif response_data['return'] == [{}]:
                        logger.error('nginx_upstream页状态变更SaltAPI调用cmd_run_api获取到空值了')
                        return JsonResponse({'result': '失败,后台问题', 'status': False})
                    else:
                        return JsonResponse({'result': '成功', 'status': True})
    except Exception as e:
        logger.error('获取upstream列表失败'+str(e))
        return render(request, 'nginx_upstream.html')


# nginx的upstream负载切换页面新增一个server功能
def nginx_upstreamserver_add(request):
    minionid = request.POST.get('minionid')
    path = request.POST.get('path')
    upserverdata = request.POST.get('upserverdata')
    upstreamname = request.POST.get('upstreamname')
    srserver = request.POST.get('srserver')
    with requests.Session() as s:
        try:
            token = s.post('http://192.168.68.50:8080/login',
                           json={'username': 'saltapi', 'password': '123456', 'eauth': 'pam', })
            token.raise_for_status()
        except Exception as e:
            logger.error('upstream新增server操作post获取token报错：'+str(e))
            return JsonResponse({'result': 'upstream新增server操作post获取token报错', 'status': False})
        else:
            serverdata = 'server '+upserverdata+';'
            # 注意下面的命令和在shell下执行不一样地方就是\\n\\t在shell下直接\n\t，在这里之所以要这样因为不转义为普通的\n\t会直接先被
            # python给解析掉了。。。直接变成回车和tab了，无法传到客户端去，无语，坑啊，记住了！！！
            arg = "sed -i '/^upstream *%s/,/^}$/{s/%s/%s\\n\\t%s/g}' %s" % (upstreamname, srserver, srserver, serverdata, path)
            data = {'client': 'local',
                    'tgt': minionid,
                    'tgt_type': 'glob',
                    'fun': 'cmd.run',
                    'arg': arg,
                    }
            try:
                response_data = s.post('http://192.168.68.50:8080', data=data)
                response_data.raise_for_status()
            except Exception as e:
                reponse_data = '新增upstream的server配置操作提交出错：'
                logger.error(reponse_data + str(e))
                # return logger.info(e + ' 无法获取数据')
                return JsonResponse({'result': '失败', 'status': False})
            else:
                logger.error(response_data.json())
                return JsonResponse({'result': '成功', 'status': True})


# nginx reload操作,也是在nginx负载切换的页面。由于是ajax提交所以可以分开单独写呵呵
def nginx_reload(request):
    minionid = request.POST.get('minionid')
    nginxip = request.POST.get('nginxip')
    with requests.Session() as s:
        try:
            token = s.post('http://192.168.68.50:8080/login',
                           json={'username': 'saltapi', 'password': '123456', 'eauth': 'pam',})
            token.raise_for_status()
        except Exception as e:
            logger.error('nginx reload操作post获取token报错：', e)
            return JsonResponse({'result': 'nginx reload操作post获取token报错', 'status': False})
        else:
            # 通过nginxip获取nginx的path和conf路径来reload
            nginx = NginxManage.objects.get(ip=nginxip)
            path = nginx.path
            conf = nginx.confpath
            cmd = '%ssbin/nginx -t -c %snginx.conf && %ssbin/nginx -s reload' % (path, conf, path)
            data = {'client': 'local',
                    'tgt': minionid,
                    'expr_form': 'glob',
                    'fun': 'cmd.run',
                    'arg': cmd,
                    }
            try:
                response_data = s.post('http://192.168.68.50:8080', data=data)
                response_data.raise_for_status()
            except Exception as e:
                reponse_data = 'nginx重载配置操作提交出错：'
                logger.error(reponse_data + str(e))
                # return logger.info(e + ' 无法获取数据')
                return JsonResponse({'result': '失败', 'status': False})
            else:
                data = response_data.json()['return'][0][minionid]
                if 'test is successful' in data:
                    return JsonResponse({'result': 'reload成功', 'status': True})
                else:
                    return JsonResponse({'result': data, 'status': False})



# H5临时发布系统
def h5_issue(request):
    return render(request, 'h5_issue.html')


# H5——svn检出
def h5_svn_co(request):
    svn_addr = request.POST.get('svn_addr')
    svn_user = request.POST.get('svn_user')
    svn_pwd = request.POST.get('svn_pwd')
    svn_version = request.POST.get('svn_version')
    local_path = '/data/svn/'+svn_addr.rsplit('/', 1)[1]
    minionid='192.168.68.50-master'
    with requests.Session() as s:
        saltapi = SaltAPI(session=s)
        if saltapi.get_token() is False:
            logger.error('h5_svn检出获取SaltAPI调用get_token请求出错')
            return JsonResponse({'result': 'h5_svn检出获取SaltAPI调用get_token请求出错', 'status': False})
        else:
            response_data = saltapi.cmd_run_api(tgt=minionid, arg=['cd %s && svn up -r %s --non-interactive  --username=%s --password=%s ' % (local_path, svn_version, svn_user, svn_pwd), 'reset_system_locale=false'])
            if response_data is False:
                logger.error('h5_svn检出可能代入的参数有问题，SaltAPI调用cmd_run_api请求出错')
                return JsonResponse({'result': 'h5_svn检出可能代入的参数有问题，SaltAPI调用cmd_run_api请求出错', 'status': False})
            # 判断upstream_data如果返回值如果为[{}]表明没有这个minionid
            elif response_data['return'] != [{}]:
                data_source = response_data['return'][0][minionid]
                logger.error(data_source)
                if 'svn: E' in data_source:
                    return JsonResponse({'result': data_source, 'status': False})
                # 在首次更新的时候如果还没做过svn co检出是无法更新的，所以要判断是否已经存在了这个svn目录，不存在就要co一次
                elif 'No such file or directory' in data_source:
                    response_data = saltapi.cmd_run_api(tgt=minionid, arg=['svn co -r %s --non-interactive %s %s --username=%s --password=%s' % (svn_version, svn_addr, local_path, svn_user, svn_pwd), 'reset_system_locale=false'])
                    if response_data is False:
                        logger.error('h5_svn检出可能代入的参数有问题，SaltAPI调用cmd_run_api请求出错')
                        return JsonResponse({'result': 'h5_svn检出可能代入的参数有问题，SaltAPI调用cmd_run_api请求出错', 'status': False})
                    # 判断upstream_data如果返回值如果为[{}]表明没有这个minionid
                    elif response_data['return'] != [{}]:
                        data_source = response_data['return'][0][minionid]
                        if 'svn: E' in data_source:
                            return JsonResponse({'result': data_source, 'status': False})
                        else:
                            return JsonResponse({'result': data_source, 'status': True})
                    else:
                        logger.error('h5_svn检出失败，请确认minion是否存在。。')
                        return JsonResponse({'result': 'h5_svn检出失败，请确认minion是否存在。。', 'status': False})
                else:
                    return JsonResponse({'result': data_source, 'status': True})
            else:
                logger.error('h5_svn检出失败，请确认minion是否存在。。')
                return JsonResponse({'result': 'h5_svn检出失败，请确认minion是否存在。。', 'status': False})


# H5——svn压缩
def h5_svn_zip(request):
    svn_addr = request.POST.get('svn_addr')
    cwd = '/data/svn/'+svn_addr.rsplit('/', 1)[1]
    zip_path = request.POST.get('zip_path')
    minionid='192.168.68.50-master'
    with requests.Session() as s:
        saltapi = SaltAPI(session=s)
        if saltapi.get_token() is False:
            logger.error('h5_svn压缩获取SaltAPI调用get_token请求出错')
            return JsonResponse({'result': 'h5_svn压缩获取SaltAPI调用get_token请求出错', 'status': False})
        else:
            response_data = saltapi.cmd_run_api(tgt=minionid, arg='tar -C %s -cvf %s *' % (cwd, zip_path))
            if response_data is False:
                logger.error('h5_svn压缩获取可能代入的参数有问题，SaltAPI调用cmd_run_api请求出错')
                return JsonResponse({'result': 'h5_svn检出可能代入的参数有问题，SaltAPI调用cmd_run_api请求出错', 'status': False})
            # 判断upstream_data如果返回值如果为[{}]表明没有这个minionid
            elif response_data['return'] != [{}]:
                data_source = response_data['return'][0][minionid]
                return JsonResponse({'result': data_source, 'status': True})
            else:
                logger.error('h5_svn压缩获取失败，请确认minion是否存在。。')
                return JsonResponse({'result': 'h5_svn压缩获取，请确认minion是否存在。。', 'status': False})


# H5压缩文件查询，用来部署前选择部署文件
def h5_file(request):
    project_file = request.POST.get('project_file')
    minionid = '192.168.68.50-master'
    if project_file == 'H5_*.tar':
        with requests.Session() as s:
            saltapi = SaltAPI(session=s)
            if saltapi.get_token() is False:
                logger.error('h5_file获取SaltAPI调用get_token请求出错')
                return JsonResponse({'result': 'h5_file获取SaltAPI调用get_token请求出错', 'status': False})
            else:
                # 第一个要在master中定义好，第二个zip_path是minion的路径，解压可以解压到一个不存在的目录，会自动创建哟
                response_data = saltapi.cmd_run_api(tgt=minionid,
                                                    arg="find /data/svnzip/ -type f -name '%s' | xargs ls -t" % project_file)
                if response_data is False:
                    logger.error('h5_file可能代入的参数有问题，SaltAPI调用archive_unzip_api请求出错')
                    return JsonResponse(
                        {'result': 'h5_file可能代入的参数有问题，SaltAPI调用archive_unzip_api请求出错', 'status': False})
                # 判断upstream_data如果返回值如果为[{}]表明没有这个minionid
                elif response_data['return'] != [{}]:
                    data_source = response_data['return'][0][minionid]
                    data_source = [x.split('/')[-1] for x in data_source.split()]
                    return JsonResponse({'result': data_source, 'status': True})
                else:
                    logger.error('h5_file失败，请确认minion是否存在。。')
                return JsonResponse({'result': 'h5_file失败，请确认minion是否存在。。', 'status': False})


# H5-svn文件传输
def h5_svn_push(request):
    zip_file = request.POST.get('project_file')
    zip_path = '/data/svnzip/'+zip_file
    source_path = 'salt://%s?saltenv=svnzip' % zip_file
    # 下面这个到时候改成H5的minionid
    project_ip = request.POST.get('project_ip')
    minion_id = MinionList.objects.get(ip=project_ip).minion_id
    with requests.Session() as s:
        saltapi = SaltAPI(session=s)
        if saltapi.get_token() is False:
            logger.error('h5_svn推送获取SaltAPI调用get_token请求出错')
            return JsonResponse({'result': 'h5_svn推送获取SaltAPI调用get_token请求出错', 'status': False})
        else:
            # 第一个要在master中定义好，第二个zip_path是minion的路径
            response_data = saltapi.cp_get_file_api(tgt=minion_id, arg=[source_path, zip_path])
            if response_data is False:
                logger.error('h5_svn推送可能代入的参数有问题，SaltAPI调用cp_get_file_api请求出错')
                return JsonResponse({'result': 'h5_svn推送可能代入的参数有问题，SaltAPI调用cp_get_file_api请求出错', 'status': False})
            # 判断upstream_data如果返回值如果为[{}]表明没有这个minionid
            elif response_data['return'] != [{}]:
                data_source = response_data['return'][0][minion_id]
                return JsonResponse({'result': data_source, 'status': True})
            else:
                logger.error('h5_svn推送失败，请确认minion是否存在。。')
                return JsonResponse({'result': 'h5_svn推送失败，请确认minion是否存在。。', 'status': False})


# H5-svn文件解压
def h5_svn_unzip(request):
    zip_file = request.POST.get('project_file')
    zip_path = '/data/svnzip/' + zip_file
    project_addr = request.POST.get('project_addr')
    # 下面这个到时候改成H5的minionid
    project_ip = request.POST.get('project_ip')
    minion_id = MinionList.objects.get(ip=project_ip).minion_id
    with requests.Session() as s:
        saltapi = SaltAPI(session=s)
        if saltapi.get_token() is False:
            logger.error('h5_svn解压获取SaltAPI调用get_token请求出错')
            return JsonResponse({'result': 'h5_svn解压获取SaltAPI调用get_token请求出错', 'status': False})
        else:
            # 第一个要在master中定义好，第二个是minion的路径，解压可以解压到一个不存在的目录，会自动创建哟
            response_data = saltapi.cmd_run_api(tgt=minion_id, arg='tar -xvf %s -C %s' % (zip_path, project_addr))
            if response_data is False:
                logger.error('h5_svn解压可能代入的参数有问题，SaltAPI调用archive_unzip_api请求出错')
                return JsonResponse({'result': 'h5_svn解压可能代入的参数有问题，SaltAPI调用archive_unzip_api请求出错', 'status': False})
            # 判断upstream_data如果返回值如果为[{}]表明没有这个minionid
            elif response_data['return'] != [{}]:
                data_source = response_data['return'][0][minion_id]
                logger.error(data_source)
                return JsonResponse({'result': data_source, 'status': True})
            else:
                logger.error('h5_svn解压失败，请确认minion是否存在。。')
                return JsonResponse({'result': 'h5_svn解压失败，请确认minion是否存在。。', 'status': False})


# 发布系统 应用发布 主
def app_release(request):
    try:
        if request.method == 'GET':
            # 默认如果没有get到的话值为None，这里我需要为空''，所以下面修改默认值为''
            search_field = request.GET.get('search_field', '')
            search_content = request.GET.get('search_content', '')
            # 判断是否为超级管理员或者普通用户，按权限分配
            if request.user.is_superuser:
                if search_content is '':
                    app_data = AppRelease.objects.all().order_by('create_time')
                    data_list = getPage(request, app_data, 15)
                else:
                    if search_field == 'search_app_name':
                        app_data = AppRelease.objects.filter(
                            app_name__icontains=search_content).order_by(
                            'create_time')
                        data_list = getPage(request, app_data, 15)
                    elif search_field == 'search_minion_id':
                        app_data = AppRelease.objects.filter(
                            minion_id__icontains=search_content).order_by(
                            'create_time')
                        data_list = getPage(request, app_data, 15)
                    elif search_field == 'search_svn_url':
                        app_data = AppRelease.objects.filter(
                            app_svn_url__icontains=search_content).order_by(
                            'create_time')
                        data_list = getPage(request, app_data, 15)
                    else:
                        data_list = ""
                return render(request, 'app_release.html',
                          {'data_list': data_list, 'search_field': search_field,
                           'search_content': search_content})
            else:
                username = request.user.username
                try:
                    app_auth_app_data = AppAuth.objects.get(username=username).app_perms.split(',')
                except Exception as e:
                    app_auth_app_data = ''
                if search_content is '':
                    app_data = AppRelease.objects.filter(app_name__in=app_auth_app_data).order_by('create_time')
                    data_list = getPage(request, app_data, 15)
                else:
                    if search_field == 'search_app_name':
                        app_data = AppRelease.objects.filter(app_name__in=app_auth_app_data).filter(
                            app_name__icontains=search_content).order_by(
                            'create_time')
                        data_list = getPage(request, app_data, 15)
                    elif search_field == 'search_minion_id':
                        app_data = AppRelease.objects.filter(app_name__in=app_auth_app_data).filter(
                            minion_id__icontains=search_content).order_by(
                            'create_time')
                        data_list = getPage(request, app_data, 15)
                    else:
                        data_list = ""
                return render(request, 'app_release.html',
                          {'data_list': data_list, 'search_field': search_field,
                           'search_content': search_content})

    except Exception as e:
        logger.error('应用发布页面有问题:'+str(e))
        return render(request, 'app_release.html')


# 发布系统 应用发布页ajax提交处理
def app_release_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.GET.get('app_tag_key') == 'modal_search_minion_id':
                minion_id = request.GET.get('minion_id')
                sys = request.GET.get('sys_type')
                minion_id_list = MinionList.objects.filter(minion_id__icontains=minion_id, sys=sys).order_by(
                    'create_date').values_list('minion_id', flat=True)
                result['result'] = list(minion_id_list)
                result['status'] = True
                # 返回字典之外的需要把参数safe改成false如：JsonResponse([1, 2, 3], safe=False)
                return JsonResponse(result)
            elif request.POST.get('app_tag_key') == 'app_add' and request.user.is_superuser:
                obj = AppReleaseAddForm(request.POST)
                if obj.is_valid():
                    app_svn_co_path = settings.SITE_BASE_SVN_PATH + time.strftime('%Y%m%d_%H%M%S')
                    AppRelease.objects.create(app_name=obj.cleaned_data["app_name"],
                                              sys_type=obj.cleaned_data["sys_type"],
                                              minion_id=obj.cleaned_data["minion_id"],
                                              app_path=obj.cleaned_data["app_path"],
                                              app_path_owner=obj.cleaned_data["app_path_owner"],
                                              app_svn_url=obj.cleaned_data["app_svn_url"],
                                              app_svn_user=obj.cleaned_data["app_svn_user"],
                                              app_svn_password=obj.cleaned_data["app_svn_password"],
                                              app_svn_co_path=app_svn_co_path,
                                              execution_style=obj.cleaned_data["execution_style"],
                                              operation_content=obj.cleaned_data["operation_content"],
                                              operation_arguments=obj.cleaned_data["operation_arguments"],
                                              app_backup_path=obj.cleaned_data["app_backup_path"],
                                              description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('app_tag_key') == 'app_update' and request.user.is_superuser:
                obj = AppReleaseUpdateForm(request.POST)
                if obj.is_valid():
                    app_svn_co_path = AppRelease.objects.get(
                        app_name=obj.cleaned_data["app_name"]).app_svn_co_path
                    source_app_svn_url = AppRelease.objects.get(
                        app_name=obj.cleaned_data["app_name"]).app_svn_url
                    update_app_svn_url = obj.cleaned_data["app_svn_url"]
                    # 判断下svn地址有没改变，如果有改变删除master上的svn目录并重置svn两版本字段为none，下次有更新再重新检出，但注意目录路径是不改变的！！
                    # 主要是因为发现如果svn目录不重新生成会出现旧的svn文件不会在新svn地址检出后清除掉
                    if source_app_svn_url == update_app_svn_url:
                        AppRelease.objects.filter(app_name=obj.cleaned_data["app_name"]).update(
                            sys_type=obj.cleaned_data["sys_type"],
                            minion_id=obj.cleaned_data["minion_id"],
                            app_path=obj.cleaned_data["app_path"],
                            app_path_owner=obj.cleaned_data["app_path_owner"],
                            app_svn_url=obj.cleaned_data["app_svn_url"],
                            app_svn_user=obj.cleaned_data["app_svn_user"],
                            app_svn_password=obj.cleaned_data["app_svn_password"],
                            execution_style=obj.cleaned_data["execution_style"],
                            operation_content=obj.cleaned_data["operation_content"],
                            operation_arguments=obj.cleaned_data["operation_arguments"],
                            app_backup_path=obj.cleaned_data["app_backup_path"],
                            description=obj.cleaned_data["description"])
                        result['result'] = '成功'
                        result['status'] = True
                    else:
                        with requests.Session() as s:
                            saltapi = SaltAPI(session=s)
                            if saltapi.get_token() is False:
                                error_data = '更新应用删除应用旧检出目录获取SaltAPI调用get_token请求出错'
                                logger.error(error_data)
                                result['result'] = error_data
                                return JsonResponse(result)
                            else:
                                response_data = saltapi.file_directory_exists_api(tgt=settings.SITE_SALT_MASTER,
                                                                                  arg=[app_svn_co_path])
                                # 当调用api失败的时候会返回false
                                if response_data is False:
                                    error_data = '更新应用删除应用旧检出目录失败，SaltAPI调用file_directory_exists_api请求出错'
                                    logger.error(error_data)
                                    result['result'] = error_data
                                    return JsonResponse(result)
                                else:
                                    response_data = response_data['return'][0][settings.SITE_SALT_MASTER]
                                    # 判断一下svn检出的目录是否存在，因为如果没发布过，目录还没生成，存在的话删，
                                    # 不然新svn检出和旧svn检出内容会重叠导致后面同步文件不对
                                    if response_data is True:
                                        response_data = saltapi.file_remove_api(tgt=settings.SITE_SALT_MASTER,
                                                                                arg=[app_svn_co_path])
                                        # 当调用api失败的时候会返回false
                                        if response_data is False:
                                            error_data = '删除应用旧检出目录失败，SaltAPI调用file_remove_api请求出错'
                                            logger.error(error_data)
                                            result['result'] = error_data
                                            return JsonResponse(result)
                                        else:
                                            response_data = response_data['return'][0][settings.SITE_SALT_MASTER]
                                            if response_data is True:
                                                # 删除成功后提交更新，记得把应用svn版本和成功svn版本还原None,可以不更新检出目录
                                                # 检出目录还是不变，在做发布的时候会重新自动创建出来不用担心
                                                AppRelease.objects.filter(app_name=obj.cleaned_data["app_name"]).update(
                                                    sys_type=obj.cleaned_data["sys_type"],
                                                    minion_id=obj.cleaned_data["minion_id"],
                                                    app_path=obj.cleaned_data["app_path"],
                                                    app_path_owner=obj.cleaned_data["app_path_owner"],
                                                    app_svn_url=obj.cleaned_data["app_svn_url"],
                                                    app_svn_user=obj.cleaned_data["app_svn_user"],
                                                    app_svn_password=obj.cleaned_data["app_svn_password"],
                                                    execution_style=obj.cleaned_data["execution_style"],
                                                    operation_content=obj.cleaned_data["operation_content"],
                                                    operation_arguments=obj.cleaned_data["operation_arguments"],
                                                    app_backup_path=obj.cleaned_data["app_backup_path"],
                                                    description=obj.cleaned_data["description"],
                                                    app_svn_version=None, app_svn_version_success=None)
                                                result['result'] = '成功'
                                                result['status'] = True
                                            else:
                                                logger.error('更新应用删除应用检出目录结果错误：' + str(response_data))
                                                result['result'] = '更新应用删除应用检出目录结果错误：' + str(response_data)
                                                return JsonResponse(result)
                                    else:
                                        AppRelease.objects.filter(app_name=obj.cleaned_data["app_name"]).update(
                                            sys_type=obj.cleaned_data["sys_type"],
                                            minion_id=obj.cleaned_data["minion_id"],
                                            app_path=obj.cleaned_data["app_path"],
                                            app_path_owner=obj.cleaned_data["app_path_owner"],
                                            app_svn_url=obj.cleaned_data["app_svn_url"],
                                            app_svn_user=obj.cleaned_data["app_svn_user"],
                                            app_svn_password=obj.cleaned_data["app_svn_password"],
                                            execution_style=obj.cleaned_data["execution_style"],
                                            operation_content=obj.cleaned_data["operation_content"],
                                            operation_arguments=obj.cleaned_data["operation_arguments"],
                                            app_backup_path=obj.cleaned_data["app_backup_path"],
                                            description=obj.cleaned_data["description"])
                                        result['result'] = '成功'
                                        result['status'] = True

                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('app_tag_key') == 'app_delete' and request.user.is_superuser:
                app_name = request.POST.get('app_name')
                delete_app_file_select = request.POST.get('delete_app_file_select')
                try:
                    app_svn_co_path = AppRelease.objects.get(app_name=app_name).app_svn_co_path
                    app_path = AppRelease.objects.get(app_name=app_name).app_path
                    app_backup_path = AppRelease.objects.get(app_name=app_name).app_backup_path
                    minion_id = AppRelease.objects.get(app_name=app_name).minion_id
                    minion_id_list = minion_id.split(',')

                    app_group_exist = AppGroup.objects.filter(
                        app_group_members__regex=r'^%s$|^%s,|,%s$|,%s,' % (app_name, app_name, app_name, app_name)).exists()
                    if app_group_exist:
                        result['result'] = '该应用属于应用发布组的成员，请先从应用发布组中踢除该应用，再执行删除操作'
                        return JsonResponse(result)

                    with requests.Session() as s:
                        saltapi = SaltAPI(session=s)
                        if saltapi.get_token() is False:
                            error_data = '删除应用获取SaltAPI调用get_token请求出错'
                            logger.error(error_data)
                            result['result'] = error_data
                            return JsonResponse(result)
                        else:
                            # 判断一下svn检出的目录是否存在，因为如果没发布过，目录还没生成，存在的话删除项目的时候要顺带删除
                            response_data = saltapi.file_directory_exists_api(tgt=settings.SITE_SALT_MASTER,
                                                                              arg=[app_svn_co_path])
                            # 当调用api失败的时候会返回false
                            if response_data is False:
                                error_data = '删除应用失败，SaltAPI调用file_directory_exists_api请求出错'
                                logger.error(error_data)
                                result['result'] = error_data
                                return JsonResponse(result)
                            else:
                                response_data = response_data['return'][0][settings.SITE_SALT_MASTER]
                                if response_data is True:
                                    # 删除master端项目的检出目录
                                    response_data = saltapi.file_remove_api(tgt=settings.SITE_SALT_MASTER,
                                                                            arg=[app_svn_co_path])
                                    # 当调用api失败的时候会返回false
                                    if response_data is False:
                                        error_data = '删除应用失败，SaltAPI调用file_remove_api请求出错'
                                        logger.error(error_data)
                                        result['result'] = error_data
                                        return JsonResponse(result)
                                    else:
                                        response_data = response_data['return'][0][settings.SITE_SALT_MASTER]
                                        if response_data is True:
                                            pass
                                        else:
                                            logger.error('删除应用结果错误：' + str(response_data))
                                            result['result'] = '删除应用结果错误：' + str(response_data)
                                            return JsonResponse(result)
                            if delete_app_file_select == 'delete_app_file':
                                for minion in minion_id_list:
                                    # 删除应用目录
                                    response_data = saltapi.file_directory_exists_api(tgt=minion, arg=[app_path])
                                    # 当调用api失败的时候会返回false
                                    if response_data is False:
                                        error_data = '删除应用时删除应用目录失败，SaltAPI调用file_directory_exists_api请求出错'
                                        logger.error(error_data)
                                        result['result'] = error_data
                                        return JsonResponse(result)
                                    else:
                                        response_data = response_data['return'][0][minion]
                                        # 判断一下svn检出的目录是否存在，因为如果没发布过，目录还没生成，存在的话删除项目的时候要顺带删除
                                        if response_data is True:
                                            response_data = saltapi.file_remove_api(tgt=minion, arg=[app_path])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                error_data = '删除应用时删除应用目录失败，SaltAPI调用file_remove_api请求出错'
                                                logger.error(error_data)
                                                result['result'] = error_data
                                                return JsonResponse(result)
                                            else:
                                                response_data = response_data['return'][0][minion]
                                                if response_data is True:
                                                    pass
                                                else:
                                                    logger.error('删除应用时删除应用目录结果错误：' + str(response_data))
                                                    result['result'] = '删除应用时删除应用目录结果错误：' + str(response_data)
                                                    return JsonResponse(result)
                                    # 删除备份目录
                                    response_data = saltapi.file_directory_exists_api(tgt=minion, arg=[app_backup_path])
                                    # 当调用api失败的时候会返回false
                                    if response_data is False:
                                        error_data = '删除应用时删除应用备份目录失败，SaltAPI调用file_directory_exists_api请求出错'
                                        logger.error(error_data)
                                        result['result'] = error_data
                                        return JsonResponse(result)
                                    else:
                                        response_data = response_data['return'][0][minion]
                                        # 判断一下svn检出的目录是否存在，因为如果没发布过，目录还没生成，存在的话删除项目的时候要顺带删除
                                        if response_data is True:
                                            response_data = saltapi.file_remove_api(tgt=minion, arg=[app_backup_path])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                error_data = '删除应用时删除应用备份目录失败，SaltAPI调用file_remove_api请求出错'
                                                logger.error(error_data)
                                                result['result'] = error_data
                                                return JsonResponse(result)
                                            else:
                                                response_data = response_data['return'][0][minion]
                                                if response_data is True:
                                                    pass
                                                else:
                                                    logger.error('删除应用时删除应用备份目录结果错误：' + str(response_data))
                                                    result['result'] = '删除应用时删除应用备份目录结果错误：' + str(response_data)
                                                    return JsonResponse(result)
                            AppRelease.objects.get(app_name=app_name).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.GET.get('app_tag_key') == 'check_svn':
                app_name = request.GET.get("app_name")
                try:
                    app_data = AppRelease.objects.get(app_name=app_name)
                    with requests.Session() as s:
                        saltapi = SaltAPI(session=s)
                        if saltapi.get_token() is False:
                            error_data = '应用发布查询svn版本获取SaltAPI调用get_token请求出错'
                            result['result'] = error_data
                            return JsonResponse(result)
                        else:
                            response_data = saltapi.cmd_run_api(tgt=settings.SITE_SALT_MASTER, arg=[
                                'svn info %s  --username=%s --password=%s --no-auth-cache' % (
                                    app_data.app_svn_url, app_data.app_svn_user,
                                    app_data.app_svn_password), 'reset_system_locale=false'])
                            # 当调用api失败的时候会返回false
                            if response_data is False:
                                error_data = '应用发布查询svn版本失败，SaltAPI调用async_cmd_run_api请求出错'
                                result['result'] = error_data
                                return JsonResponse(result)
                            else:
                                response_data = response_data['return'][0][settings.SITE_SALT_MASTER]
                                try:
                                    svn_version = re.search(r'Revision: (\d+)', response_data).group(1)
                                    result['result'] = svn_version
                                    result['status'] = True
                                    return JsonResponse(result)
                                except Exception as e:
                                    logger.error('检查svn版本结果错误：' + str(e) + str(response_data))
                                    result['result'] = '检查svn版本结果错误：' + str(response_data)
                                    return JsonResponse(result)
                except Exception as e:
                    logger.error('检查svn版本出错：'+str(e))
                    result['result'] = '检查出错'
                    return JsonResponse(result)
            elif request.GET.get('app_tag_key') == 'search_app_log':
                try:
                    app_name = request.GET.get("app_name")
                    log_data = AppReleaseLog.objects.filter(app_name=app_name).order_by(
                        '-create_time')
                    data_list = getPage(request, log_data, 1)
                    log_content = ''
                    create_time = ''
                    release_result = ''
                    log_app_username = ''
                    for data in data_list:
                        log_content = eval(data.log_content)
                        create_time = data.create_time.strftime('%Y-%m-%d %X')
                        release_result = data.release_result
                        log_app_username = data.username
                    if data_list.has_previous():
                        previous_page_number = data_list.previous_page_number()
                    else:
                        previous_page_number = 0
                    if data_list.has_next():
                        next_page_number = data_list.next_page_number()
                    else:
                        next_page_number = 0
                    result = {'status': True, 'log_app_username': log_app_username, 'create_time': create_time, 'release_result': release_result, 'has_previous': data_list.has_previous(), 'previous_page_number': previous_page_number, 'number': data_list.number, 'num_pages': data_list.paginator.num_pages, 'has_next': data_list.has_next(), 'next_page_number': next_page_number, 'log_content': log_content}
                    return JsonResponse(result)
                except Exception as e:
                    logger.error('查询日志错误了'+str(e))
                    return JsonResponse(result)
            elif request.POST.get('app_tag_key') == 'release_app':
                app_name = request.POST.get('app_name')
                release_svn_version = request.POST.get('release_svn_version')
                app_data = AppRelease.objects.get(app_name=app_name)
                # 判断执行的是否为单项操作执行判断，如果不是就是执行操作步骤顺序的操作
                single_cmd = request.POST.get('single_cmd')
                if single_cmd:
                    operation_content = [single_cmd]
                else:
                    operation_content = app_data.operation_content.split(',')
                operation_arguments = app_data.operation_arguments
                operation_arguments = eval(operation_arguments)
                app_svn_co_path = app_data.app_svn_co_path
                app_svn_url = app_data.app_svn_url
                app_svn_user = app_data.app_svn_user
                app_svn_password = app_data.app_svn_password
                app_svn_version_success = ''
                app_path = app_data.app_path
                sys_type = app_data.sys_type
                app_path_owner = app_data.app_path_owner
                try:
                    # 由于用的salt来做发布所以如果minion离线或不存在删除了就无法执行，所以要判断，另外还有一个原因是minion管理表如果
                    # 删除了某个minion会触发try的except
                    try:
                        minion_id_list = app_data.minion_id.split(',')
                        for minion_id in minion_id_list:
                            minion_status = MinionList.objects.get(minion_id=minion_id).minion_status
                            if minion_status == '离线':
                                app_log.append('\n应用minion_id:%s离线了，请确认全部在线或移除离线minino_id才可执行应用发布' % minion_id)
                                result['result'] = app_log
                                return JsonResponse(result)
                    except Exception as e:
                        logger.error('\n检查应用的Minion_ID出错，可能有Minion已经不存在了，报错信息:'+str(e))
                        app_log.append('\n检查应用的Minion_ID出错，可能有Minion已经不存在了，报错信息:'+str(e))
                        result['result'] = app_log
                        return JsonResponse(result)
                    for minion_id in minion_id_list:
                        # svn版本需要在这里获取，因为下面需要用到版本判断是检出还是更新操作，
                        # 如果在上面就定义好，那么多个minion_id的新项目第一个id是检出后也全是检出，因为判断版本的时候都是空
                        app_svn_version = AppRelease.objects.get(app_name=app_name).app_svn_version
                        app_log.append(('-'*10+('Minion_ID:%s开始发布 时间戳%s' % (minion_id, time.strftime('%X')))+'-'*10).center(88)+'\n')
                        for operation in operation_content:
                            if operation == 'SVN更新':
                                app_log.append('\n\n开始执行SVN更新-> 时间戳%s\n' % time.strftime('%X'))
                                with requests.Session() as s:
                                    saltapi = SaltAPI(session=s)
                                    if saltapi.get_token() is False:
                                        app_log.append('\n更新svn后台出错_error(0)，请联系管理员. 时间戳%s\n' % time.strftime('%X'))
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    else:
                                        # 判断是否有应用svn版本号，如果有说明已经检出过，那就使用更新up，如果没有就用检出co
                                        if app_svn_version:
                                            cmd_data = 'svn up -r %s %s --no-auth-cache --non-interactive  --username=%s --password=%s' % (
                                                release_svn_version, app_svn_co_path, app_svn_user, app_svn_password)
                                            # 用来做执行结果判断的，因为结果有很多意外情况，下面是对的情况下会出现的关键字
                                            check_data = "Updating '%s'" % app_svn_co_path
                                        else:
                                            cmd_data = 'svn co -r %s %s  %s --username=%s --password=%s --non-interactive --no-auth-cache' % (
                                                release_svn_version, app_svn_url, app_svn_co_path, app_svn_user, app_svn_password)
                                            check_data = 'Checked out revision'
                                        response_data = saltapi.cmd_run_api(tgt=settings.SITE_SALT_MASTER, arg=[
                                            cmd_data, 'reset_system_locale=false'])
                                        # 当调用api失败的时候会返回false
                                        if response_data is False:
                                            app_log.append('\n更新svn后台出错_error(1)，请联系管理员. 时间戳%s\n' % time.strftime('%X'))
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            response_data = response_data['return'][0][settings.SITE_SALT_MASTER]

                                            if check_data in response_data:
                                                # 用正则获取版本号，并更新一下数据表,这里发现有出错的可能就是正则没匹配到，所以再加一层try
                                                try:
                                                    app_svn_version = re.search(r'revision (\d+)\.', response_data).group(1)
                                                    AppRelease.objects.filter(app_name=app_name).update(
                                                        app_svn_version=app_svn_version)
                                                    app_svn_version_success = app_svn_version
                                                    app_log.append('\n'+str(response_data)+'\n\nSVN更新完成<- 时间戳%s\n' % time.strftime('%X'))
                                                except Exception as e:
                                                    app_log.append('\nSVN更新失败:\n'+str(response_data)+'\n时间戳%s' % time.strftime('%X'))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                            else:
                                                app_log.append('\nSVN更新失败:'+str(response_data)+'\n时间戳%s' % time.strftime('%X'))
                                                result['result'] = app_log
                                                return JsonResponse(result)
                            elif operation == '同步文件':
                                sync_file_method = operation_arguments.get('文件同步方法', 'salt')
                                if sync_file_method == 'salt':
                                    source_path = app_svn_co_path.rstrip('/').rsplit('/', 1)[1]
                                    sync_file_style = operation_arguments['文件同步方式']
                                    svn_symlink_path = settings.SITE_BASE_SVN_SYMLINK_PATH + source_path
                                    app_log.append('\n\n开始执行同步文件-> 时间戳%s\n' % time.strftime('%X'))
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n同步文件后台出错_error(0)，请联系管理员.  时间戳%s\n' % time.strftime('%X'))
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            # 先创建软连接
                                            response_data = saltapi.file_symlink_api(tgt=settings.SITE_SALT_MASTER,
                                                                                     arg=[app_svn_co_path, svn_symlink_path])
                                            if response_data is False:
                                                app_log.append('\n同步文件后台出错_error(1)，请联系管理员. 时间戳%s\n' % time.strftime('%X'))
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                if response_data['return'][0][settings.SITE_SALT_MASTER] is not True:
                                                    # 如果软连接创建失败会返回：{'return': [{'192.168.100.170': False}]}
                                                    app_log.append('同步文件过程中，创建软连接失败\n' + str(response_data))
                                                    app_log.append('\n' + '文件同步失败！！ 时间戳%s\n' % time.strftime('%X'))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                            # 执行文件同步
                                            jid = saltapi.async_state_api(tgt=minion_id, arg=["rsync_dir", "pillar={'sync_file_method':'%s','source_path':'%s','name_path':'%s','user':'%s','sync_file_style':'%s'}" % (sync_file_method, source_path, app_path, app_path_owner, sync_file_style), "queue=True"])
                                            if jid is False:
                                                app_log.append('\n同步文件后台出错,SaltAPI调用async_state_api请求出错，请联系管理员. 时间戳%s\n' % time.strftime('%X'))
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    jid = jid['return'][0]['jid']
                                                    check_count = 400
                                                    re_count = 0
                                                    time.sleep(10)
                                                    while check_count:
                                                        job_status = saltapi.job_active_api(tgt=minion_id, arg=jid)
                                                        if job_status is False:
                                                            app_log.append(
                                                                '\n同步文件后台出错,SaltAPI调用job_active_api请求出错，请联系管理员. 时间戳%s\n' % time.strftime('%X'))
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                        else:
                                                            value = job_status['return'][0][minion_id]
                                                            if value:
                                                                # 为真说明job还在执行，刚好用来恢复断线false的计数器
                                                                if re_count > 0:
                                                                    re_count = 0
                                                            # 这个留在这里做个说明，我发现在调用job_active_api接口的时候经常失败返回false了，感觉是接口有问题
                                                            # 而如果出现都是false用jid_api接口取到的结果就会是[{}]所以下面对这个要做一层判断，以免因为接口不稳导致没取到结果
                                                            # 另外注意这里value is False看上去好像和上面是if value是相反的可以直接用else代替，但是不行！因为当执行完毕返回是{}而{}和False是不同的！
                                                            elif value is False:
                                                                # 连续监测2次都是那就不用跑了直接返回离线结束呵呵
                                                                if re_count == 2:
                                                                    app_log.append('\n同步文件后台出错,您要发布的主机%s离线了，请联系管理员. 时间戳%s\n' % (minion_id, time.strftime('%X')))
                                                                    result['result'] = app_log
                                                                    return JsonResponse(result)
                                                                # re计数器不到3次则+1，继续下一轮循环
                                                                else:
                                                                    re_count += 1
                                                            # 当value等于[{}]时候说明job执行完毕了，则执行下面
                                                            else:
                                                                jid_data = saltapi.jid_api(jid=jid)
                                                                # 注意[{}] ！= False所以不能用if jid_data['return']判断是否有数据，这个坑埋了好久奶奶的！！！
                                                                if jid_data is False:
                                                                    app_log.append('\n同步文件后台出错,SaltAPI调用jid_api请求出错,jid:%s，请联系管理员. 时间戳%s\n' % (jid, time.strftime('%X')))
                                                                    result['result'] = app_log
                                                                    return JsonResponse(result)
                                                                elif jid_data['return'] == [{}]:
                                                                    # 这个判断没必要，只是留这里做个说明，我之前上面没有做if value is False判断的时候，如果job_active_api
                                                                    # 的结果全部false了也会正常跳出for循环，然后在这里会出现jid_data['return'] == [{}]的情况，因为false
                                                                    # 说明minion断线了，结果肯定取到空了；还有另一种情况就是还没有返回值的时候也会等于[{}],
                                                                    # 不过后面我在上面加了对false做判断这里就没必要了呵呵
                                                                    pass
                                                                else:
                                                                    format_result = format_state(jid_data)
                                                                    if type(format_result) == str:
                                                                        # 如果minion客户端停了会返回：{'return': [{'192.168.100.170': False}]}
                                                                        app_log.append(format_result)
                                                                        app_log.append(
                                                                            '\n' + '文件同步失败！！ 时间戳%s\n' % time.strftime('%X'))
                                                                        result['result'] = app_log
                                                                        return JsonResponse(result)
                                                                    else:
                                                                        try:
                                                                            failed_result = re.search(r'Failed:     (\d+)',
                                                                                                      format_result[0]).group(1)
                                                                            if int(failed_result) != 0:
                                                                                app_log.extend(format_result)
                                                                                app_log.append(
                                                                                    '\n' + '文件同步失败！！ 时间戳%s\n' % time.strftime(
                                                                                        '%X'))
                                                                                result['result'] = app_log
                                                                                return JsonResponse(result)
                                                                            else:
                                                                                app_log.extend(format_result)
                                                                                app_log.append(
                                                                                    '\n\n文件同步完成<- 时间戳%s\n' % time.strftime(
                                                                                        '%X'))
                                                                                break
                                                                        except Exception as e:
                                                                            app_log.append('\n' + '文件同步代码出错：' + str(
                                                                                e) + '\n时间戳%s' % time.strftime('%X'))
                                                                            result['result'] = app_log
                                                                            return JsonResponse(result)
                                                            check_count -= 1
                                                            time.sleep(15)
                                                    else:
                                                        app_log.append('\n' + '文件同步超过100分钟还没有结束，系统默认同步失败，如需获取同步结果请联系管理员通过jid：%s查看！！ 时间戳%s\n' % (jid, time.strftime('%X')))
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append(str(e))
                                                    app_log.append('\n' + '文件同步失败！！ 时间戳%s\n' % time.strftime('%X'))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                                finally:
                                                    # 释放掉软连接
                                                    response_data = saltapi.file_remove_api(tgt=settings.SITE_SALT_MASTER,
                                                                                            arg=[svn_symlink_path])
                                                    # 当调用api失败的时候会返回false
                                                    if response_data is False:
                                                        app_log.append('删除软连接失败(0)，未避免目录不断膨胀请联系管理员删除软连接\n')
                                                    else:
                                                        response_data = response_data['return'][0][
                                                            settings.SITE_SALT_MASTER]
                                                        if response_data is True:
                                                            app_log.append('\n释放软连接成功\n')
                                                        else:
                                                            app_log.append('\n释放软连接失败(1)，未避免目录不断膨胀请联系管理员删除软连接\n')
                                elif sync_file_method == 'rsync':
                                    source_path = app_svn_co_path.rstrip('/').rsplit('/', 1)[1]
                                    sync_file_style = operation_arguments.get('文件同步方式', 'check_file')
                                    rsync_ip = operation_arguments.get('rsync源IP', '192.168.18.18')
                                    # salt-2018.3.0以前rsync的参数中没有additional_opts，无法指定很多东西，2018版本就有了，留这里为了新版使用
                                    rsync_port = operation_arguments.get('rsync源端口', '873')
                                    app_log.append('\n\n开始执行同步文件-> 时间戳%s\n' % time.strftime('%X'))
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n同步文件后台出错_error(0)，请联系管理员.  时间戳%s\n' % time.strftime('%X'))
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            if sys_type == 'windows':
                                                # windows下的rsync语法中路径写法比较特殊，所以要做下修改来适应，name_path存传递给后端SLS的name字段
                                                name_path = '/cygdrive/' + app_path.replace(':\\', '/').replace('\\', '/')
                                            else:
                                                name_path = app_path
                                            jid = saltapi.async_state_api(tgt=minion_id, arg=["rsync_dir",
                                                                                              "pillar={'sync_file_method':'%s','mkdir_path':'%s','rsync_ip':'%s','rsync_port':'%s','source_path':'%s','name_path':'%s','user':%s,'sync_file_style':'%s'}" % (
                                                                                              sync_file_method,app_path,
                                                                                              rsync_ip, rsync_port, source_path,
                                                                                              name_path, app_path_owner,
                                                                                              sync_file_style),
                                                                                              "queue=True"])

                                            if jid is False:
                                                app_log.append(
                                                    '\n同步文件后台出错,SaltAPI调用async_rsync_rsync_api请求出错，请联系管理员. 时间戳%s\n' % time.strftime(
                                                        '%X'))
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    jid = jid['return'][0]['jid']
                                                    check_count = 400
                                                    re_count = 0
                                                    time.sleep(10)
                                                    while check_count:
                                                        job_status = saltapi.job_active_api(tgt=minion_id, arg=jid)
                                                        if job_status is False:
                                                            app_log.append(
                                                                '\n同步文件后台出错,SaltAPI调用job_active_api请求出错，请联系管理员. 时间戳%s\n' % time.strftime(
                                                                    '%X'))
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                        else:
                                                            value = job_status['return'][0][minion_id]
                                                            if value:
                                                                # 为真说明job还在执行，刚好用来恢复断线false的计数器
                                                                if re_count > 0:
                                                                    re_count = 0
                                                            # 这个留在这里做个说明，我发现在调用job_active_api接口的时候经常失败返回false了，感觉是接口有问题
                                                            # 而如果出现都是false用jid_api接口取到的结果就会是[{}]所以下面对这个要做一层判断，以免因为接口不稳导致没取到结果
                                                            # 另外注意这里value is False看上去好像和上面是if value是相反的可以直接用else代替，但是不行！因为当执行完毕返回是{}而{}和False是不同的！
                                                            elif value is False:
                                                                # 连续监测2次都是那就不用跑了直接返回离线结束呵呵
                                                                if re_count == 2:
                                                                    app_log.append(
                                                                        '\n同步文件后台出错,您要发布的主机%s离线了，请联系管理员. 时间戳%s\n' % (
                                                                        minion_id, time.strftime('%X')))
                                                                    result['result'] = app_log
                                                                    return JsonResponse(result)
                                                                # re计数器不到3次则+1，继续下一轮循环
                                                                else:
                                                                    re_count += 1
                                                            # 当value等于[{}]时候说明job执行完毕了，则执行下面
                                                            else:
                                                                jid_data = saltapi.jid_api(jid=jid)
                                                                # 注意[{}] ！= False所以不能用if jid_data['return']判断是否有数据，这个坑埋了好久奶奶的！！！
                                                                if jid_data is False:
                                                                    app_log.append(
                                                                        '\n同步文件后台出错,SaltAPI调用jid_api请求出错，请联系管理员. 时间戳%s\n' % time.strftime(
                                                                            '%X'))
                                                                    result['result'] = app_log
                                                                    return JsonResponse(result)
                                                                elif jid_data['return'] == [{}]:
                                                                    # 这个判断没必要，只是留这里做个说明，我之前上面没有做if value is False判断的时候，如果job_active_api
                                                                    # 的结果全部false了也会正常跳出for循环，然后在这里会出现jid_data['return'] == [{}]的情况，因为false
                                                                    # 说明minion断线了，结果肯定取到空了；还有另一种情况就是还没有返回值的时候也会等于[{}],
                                                                    # 不过后面我在上面加了对false做判断这里就没必要了呵呵
                                                                    pass
                                                                else:
                                                                    format_result = format_state(jid_data)
                                                                    if type(format_result) == str:
                                                                        # 如果minion客户端停了会返回：{'return': [{'192.168.100.170': False}]}
                                                                        app_log.append(format_result)
                                                                        app_log.append(
                                                                            '\n' + '文件同步失败！！ 时间戳%s\n' % time.strftime(
                                                                                '%X'))
                                                                        result['result'] = app_log
                                                                        return JsonResponse(result)
                                                                    else:
                                                                        try:
                                                                            failed_result = re.search(
                                                                                r'Failed:     (\d+)',
                                                                                format_result[0]).group(1)
                                                                            if int(failed_result) != 0:
                                                                                app_log.extend(format_result)
                                                                                app_log.append(
                                                                                    '\n' + '文件同步失败！！ 时间戳%s\n' % time.strftime(
                                                                                        '%X'))
                                                                                result['result'] = app_log
                                                                                return JsonResponse(result)
                                                                            else:
                                                                                app_log.extend(format_result)
                                                                                app_log.append(
                                                                                    '\n\n文件同步完成<- 时间戳%s\n' % time.strftime(
                                                                                        '%X'))
                                                                                break
                                                                        except Exception as e:
                                                                            app_log.append('\n' + '文件同步代码出错：' + str(
                                                                                e) + '\n时间戳%s' % time.strftime('%X'))
                                                                            result['result'] = app_log
                                                                            return JsonResponse(result)
                                                            check_count -= 1
                                                            time.sleep(15)
                                                    else:
                                                        app_log.append(
                                                            '\n' + '文件同步超过100分钟还没有结束，系统默认同步失败，如需获取同步结果请联系管理员通过jid：%s查看！！ 时间戳%s\n' % (
                                                            jid, time.strftime('%X')))
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append(str(e))
                                                    app_log.append('\n' + '文件同步失败！！ 时间戳%s\n' % time.strftime('%X'))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                            elif operation == '应用停止':
                                app_log.append('\n\n开始执行应用服务停止操作->')
                                if '停止服务名' in operation_arguments:
                                    stop_server_name = operation_arguments['停止服务名']
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用停止后台出错_error(0)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            response_data = saltapi.service_available_api(tgt=minion_id, arg=[stop_server_name])
                                            if response_data is False:
                                                app_log.append('\n应用停止后台出错_error(1)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                if response_data['return'][0][minion_id] is False:
                                                    app_log.append('\n' + '应用停止失败,请确定是否存在该服务！！')
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                                elif response_data['return'][0][minion_id] is True:
                                                    response_data = saltapi.service_stop_api(tgt=minion_id, arg=[stop_server_name])
                                                    # 当调用api失败的时候会返回false
                                                    if response_data is False:
                                                        app_log.append('\n应用停止后台出错_error(2)，请联系管理员')
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                    else:
                                                        stop_data = response_data['return'][0][minion_id]
                                                        response_data = saltapi.service_status_api(tgt=minion_id,
                                                                                                   arg=[stop_server_name])
                                                        # 当调用api失败的时候会返回false
                                                        if response_data is False:
                                                            app_log.append('\n应用停止后台出错_error(3)，请联系管理员')
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                        elif response_data['return'][0][minion_id] is False:
                                                            app_log.append('\n'+'应用停止成功<-\n')
                                                        elif response_data['return'][0][minion_id] is True:
                                                            app_log.append('\n'+'应用停止失败，程序还在运行中。')
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                        else:
                                                            app_log.append('\n'+'应用停止失败,执行结果：'+str(stop_data)+str(response_data['return'][0][minion_id]))
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                else:
                                                    app_log.append('\n' + '应用停止失败查询服务时没有返回正确结果,执行结果：' + str(
                                                        response_data['return'][0][minion_id]))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                elif '停止命令' in operation_arguments:
                                    stop_cmd = operation_arguments['停止命令']
                                    if sys_type == 'windows':
                                        stop_cmd = stop_cmd+'&& echo %errorlevel%'
                                        split_cmd = '\r\n'
                                    else:
                                        stop_cmd = stop_cmd + '; echo $?'
                                        split_cmd = '\n'
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用停止命令后台出错_error(4)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            response_data = saltapi.cmd_run_api(tgt=minion_id, arg=[stop_cmd])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                app_log.append('\n应用停止命令后台出错_error(5)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    response_data = response_data['return'][0][minion_id].rsplit(split_cmd, 1)
                                                    # 发现有的命令没有输出那么最终只会有成功失败的0、1返回这时候列表长度就=1
                                                    if len(response_data) == 1:
                                                        if response_data[0] == '0':
                                                            app_log.append('\n' + '应用停止成功<-\n')
                                                        else:
                                                            app_log.append('\n' + '应用停止失败:' + response_data[0])
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                    else:
                                                        if response_data[1] == '0':
                                                            app_log.append('\n' + '应用停止成功<-\n')
                                                        else:
                                                            app_log.append('\n' + '应用停止失败:' + response_data[0])
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append('\n' + '应用停止失败_error(6):' + str(response_data))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                elif '任务计划停止' in operation_arguments:
                                    start_cmd = operation_arguments['任务计划停止']
                                    if sys_type == 'linux':
                                        logger.error('应用停止失败，应用停止中《任务计划启动》启动方式只适用于windows')
                                        app_log.append('\n\n应用停止失败，应用停止中《任务计划停止》停止方式只适用于windows')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用停止命令后台出错_error(1)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            response_data = saltapi.task_stop_api(tgt=minion_id, arg=[start_cmd])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                app_log.append('\n应用停止命令后台出错_error(2)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    response_data = response_data['return'][0][minion_id]
                                                    if response_data is True:
                                                        app_log.append('\n' + '应用停止成功<-\n')
                                                    else:
                                                        app_log.append('\n'+'应用停止失败:'+response_data)
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append('\n' + '应用停止后台出错_error(3):' + str(e))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                elif '映像名称和命令行' in operation_arguments:
                                    stop_cmd = operation_arguments['映像名称和命令行']
                                    data = stop_cmd.split('|')
                                    if len(data) != 2:
                                        logger.error('应用停止失败,填写的命令不符合规范')
                                        app_log.append('\n\n应用停止失败,填写的命令不符合规范')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    exe_name = data[0].strip()
                                    cmdline = data[1].strip()
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用停止命令后台出错_error(7)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            # 查看是否有映像名称的id存在，支持模糊搜索
                                            response_data = saltapi.ps_pgrep_api(tgt=minion_id, arg=[exe_name])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                app_log.append('\n应用停止命令后台出错_error(8)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    response_data = response_data['return'][0][minion_id]
                                                    if isinstance(response_data, list):
                                                        for pid in response_data:
                                                            response_data = saltapi.ps_proc_info_api(tgt=minion_id, arg=['pid=%s' % pid, 'attrs=["cmdline","status"]'])
                                                            # 当调用api失败的时候会返回false
                                                            if response_data is False:
                                                                app_log.append('\n应用停止命令后台出错_error(9)，请联系管理员\n')
                                                                result['result'] = app_log
                                                                return JsonResponse(result)
                                                            else:
                                                                # 返回的cmdline会根据命令中空格（文件名字里有空格不算）进行分割成列表，所以下面用空格合并列表
                                                                cmdline_result = ' '.join(response_data['return'][0][minion_id]['cmdline'])
                                                                if cmdline == cmdline_result:
                                                                    response_data = saltapi.ps_kill_pid_api(
                                                                        tgt=minion_id, arg=['pid=%s' % pid])
                                                                    if response_data is False:
                                                                        app_log.append(
                                                                            '\n应用停止命令后台出错_error(10)，请联系管理员\n')
                                                                        result['result'] = app_log
                                                                        return JsonResponse(result)
                                                                    else:
                                                                        if response_data['return'][0][minion_id]:
                                                                            app_log.append('\n' + '应用服务停止成功<-\n')
                                                                        else:
                                                                            app_log.append(
                                                                                '\n' + '应用停止在结束进程pid时返回结果为失败，系统默认为停止失败\n')
                                                                            result['result'] = app_log
                                                                            return JsonResponse(result)
                                                                else:
                                                                    app_log.append(
                                                                        '\n' + '应用停止在匹配命令行时没有发现可以匹配的命令行，系统默认为已经停止成功\n')

                                                    else:
                                                        app_log.append('\n'+'应用停止在查看进程时没有发现指定的进程，系统默认为已经停止成功\n')
                                                except Exception as e:
                                                    logger.error('应用服务停止代码出错：'+str(e))
                                                    app_log.append('\n' + '应用服务停止后台出错_error(11):' + str(e))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                elif 'supervisor_stop' in operation_arguments:
                                    stop_cmd = operation_arguments['supervisor_stop']
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用停止命令后台出错_error(12)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            # 直接执行supervisor停止命令，只要不出现False，就执行查询状态命令，就看状态来决定成功与否
                                            response_data = saltapi.supervisord_stop_api(tgt=minion_id, arg=[stop_cmd])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                app_log.append('\n应用停止命令后台出错_error(13)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                # 查看是否有supervisor名称存在，不支持模糊搜索
                                                response_data = saltapi.supervisord_status_api(tgt=minion_id, arg=[stop_cmd])
                                                # 当调用api失败的时候会返回false
                                                if response_data is False:
                                                    app_log.append('\n应用停止命令后台出错_error(14)，请联系管理员')
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                                else:
                                                    try:
                                                        status_result = response_data['return'][0][minion_id][stop_cmd]['state']
                                                        # 这里有发现一个问题，返回的state值可能不是STOPPED可能是FATAL或者BACKOFF,所以判断只要不是RUNNING都算停止
                                                        if status_result == 'STOPPED':
                                                            app_log.append('\n' + '应用服务停止成功<-\n')
                                                        else:
                                                            if status_result != 'RUNNING':
                                                                app_log.append(
                                                                    '\n' + '返回的状态码为%s,只要不是RUNNING应用服务都默认为停止成功<-\n' % status_result)
                                                            else:
                                                                app_log.append('\n' + '应用停止查询状态结果为RUNNING，停止失败\n')
                                                                result['result'] = app_log
                                                                return JsonResponse(result)
                                                    except Exception as e:
                                                        logger.error('应用服务停止代码出错：' + str(e))
                                                        app_log.append('\n' + '应用停止结果有错，返回结果：' + str(response_data))
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                            elif operation == '应用启动':
                                app_log.append('\n\n开始执行应用启动操作->\n')
                                if '启动服务名' in operation_arguments:
                                    start_server_name = operation_arguments['启动服务名']
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用启动后台出错_error(0)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            response_data = saltapi.service_available_api(tgt=minion_id, arg=[start_server_name])
                                            if response_data is False:
                                                app_log.append('\n应用启动后台出错_error(1)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                if response_data['return'][0][minion_id] is False:
                                                    app_log.append('\n' + '应用启动失败,请确定是否存在该服务！！')
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                                elif response_data['return'][0][minion_id] is True:
                                                    response_data = saltapi.service_start_api(tgt=minion_id, arg=[start_server_name])
                                                    # 当调用api失败的时候会返回false
                                                    if response_data is False:
                                                        app_log.append('\n应用启动后台出错_error(2)，请联系管理员')
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                    else:
                                                        start_data = response_data['return'][0][minion_id]
                                                        response_data = saltapi.service_status_api(tgt=minion_id,
                                                                                                   arg=[start_server_name])
                                                        # 当调用api失败的时候会返回false
                                                        if response_data is False:
                                                            app_log.append('\n应用启动后台出错_error(3)，请联系管理员')
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                        elif response_data['return'][0][minion_id] is False:
                                                            app_log.append('\n'+'应用启动失败。')
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                        elif response_data['return'][0][minion_id] is True:
                                                            app_log.append('\n'+'应用启动成功<-\n')
                                                        else:
                                                            app_log.append('\n'+'应用启动失败,执行结果：'+str(start_data)+str(response_data['return'][0][minion_id]))
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                else:
                                                    app_log.append('\n' + '应用启动失败查询服务时没有返回正确结果,执行结果：' + str(
                                                        response_data['return'][0][minion_id]))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                elif '启动命令' in operation_arguments:
                                    start_cmd = operation_arguments['启动命令']
                                    if sys_type == 'windows':
                                        start_cmd = start_cmd+'&& echo %errorlevel%'
                                        split_cmd = '\r\n'
                                    else:
                                        start_cmd = start_cmd + '; echo $?'
                                        split_cmd = '\n'
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用启动命令后台出错_error(3)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            response_data = saltapi.cmd_run_api(tgt=minion_id, arg=[start_cmd])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                app_log.append('\n应用启动命令后台出错_error(4)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    response_data = response_data['return'][0][minion_id].rsplit(split_cmd, 1)
                                                    # 发现有的命令没有输出那么最终只会有成功失败的0、1返回这时候列表长度就=1
                                                    if len(response_data) == 1:
                                                        if response_data[0] == '0':
                                                            app_log.append('\n' + '应用服务启动成功<-\n')
                                                        else:
                                                            app_log.append('\n' + '应用启动失败:' + response_data[0])
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                    else:
                                                        if response_data[1] == '0':
                                                            app_log.append('\n' + '应用服务启动成功<-\n')
                                                        else:
                                                            app_log.append('\n' + '应用启动失败:' + response_data[0])
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append('\n' + '应用启动失败_error(5):' + str(response_data))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                elif '任务计划启动' in operation_arguments:
                                    start_cmd = operation_arguments['任务计划启动']
                                    if sys_type == 'linux':
                                        logger.error('应用启动失败，应用启动中《任务计划启动》启动方式只适用于windows')
                                        app_log.append('\n\n应用启动失败，应用启动中《任务计划启动》启动方式只适用于windows')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用启动命令后台出错_error(6)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            response_data = saltapi.task_run_api(tgt=minion_id, arg=[start_cmd])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                app_log.append('\n应用启动命令后台出错_error(7)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    response_data = response_data['return'][0][minion_id]
                                                    if response_data is True:
                                                        app_log.append('\n' + '应用启动成功<-\n')
                                                    else:
                                                        app_log.append('\n'+'应用启动失败:'+response_data)
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append('\n' + '应用启动后台出错_error(8):' + str(e))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                elif 'supervisor_start' in operation_arguments:
                                    stop_cmd = operation_arguments['supervisor_start']
                                    with requests.Session() as s:
                                        saltapi = SaltAPI(session=s)
                                        if saltapi.get_token() is False:
                                            app_log.append('\n应用启动命令后台出错_error(9)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            # 直接执行supervisor启动命令，只要不出现False，就执行查询状态命令，就看状态来决定成功与否
                                            response_data = saltapi.supervisord_start_api(tgt=minion_id, arg=[stop_cmd])
                                            # 当调用api失败的时候会返回false
                                            if response_data is False:
                                                app_log.append('\n应用启动命令后台出错_error(10)，请联系管理员')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                # 查看是否有supervisor名称存在，不支持模糊搜索
                                                response_data = saltapi.supervisord_status_api(tgt=minion_id, arg=[stop_cmd])
                                                # 当调用api失败的时候会返回false
                                                if response_data is False:
                                                    app_log.append('\n应用启动命令后台出错_error(11)，请联系管理员')
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                                else:
                                                    try:
                                                        status_result = response_data['return'][0][minion_id][stop_cmd]['state']
                                                        if status_result == 'RUNNING':
                                                            app_log.append('\n' + '应用启动成功<-\n')
                                                        else:
                                                            app_log.append('\n' + '应用启动查询状态结果有错，返回结果：'+str(response_data))
                                                            result['result'] = app_log
                                                            return JsonResponse(result)
                                                    except Exception as e:
                                                        logger.error('应用服务启动代码出错：' + str(e))
                                                        app_log.append('\n' + '应用启动结果有错，返回结果：' + str(response_data))
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                            elif operation == '执行命令1':
                                execute_cmd = operation_arguments['执行命令1']
                                if sys_type == 'windows':
                                    execute_cmd = execute_cmd + '&& echo %errorlevel%'
                                    split_cmd = '\r\n'
                                else:
                                    execute_cmd = execute_cmd + '; echo $?'
                                    split_cmd = '\n'
                                with requests.Session() as s:
                                    saltapi = SaltAPI(session=s)
                                    if saltapi.get_token() is False:
                                        app_log.append('\n执行命令1后台出错_error(0)，请联系管理员')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    else:
                                        response_data = saltapi.cmd_run_api(tgt=minion_id, arg=[execute_cmd])
                                        # 当调用api失败的时候会返回false
                                        if response_data is False:
                                            app_log.append('\n执行命令1后台出错_error(1)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            try:
                                                response_data = response_data['return'][0][minion_id].rsplit(split_cmd, 1)
                                                # 发现有的命令没有输出那么最终只会有成功失败的0、1返回这时候列表长度就=1
                                                if len(response_data) == 1:
                                                    if response_data[0] == '0':
                                                        app_log.append('\n' + '执行命令1成功<-\n')
                                                    else:
                                                        app_log.append('\n' + '执行命令1失败:' + response_data[0])
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                else:
                                                    if response_data[1] == '0':
                                                        app_log.append('\n' + '执行命令1成功<-\n')
                                                    else:
                                                        app_log.append('\n' + '执行命令1失败:' + response_data[0])
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                            except Exception as e:
                                                app_log.append('\n' + '执行命令1失败_error(2):' + str(response_data))
                                                result['result'] = app_log
                                                return JsonResponse(result)
                            elif operation == '执行命令2':
                                execute_cmd = operation_arguments['执行命令2']
                                if sys_type == 'windows':
                                    execute_cmd = execute_cmd + '&& echo %errorlevel%'
                                    split_cmd = '\r\n'
                                else:
                                    execute_cmd = execute_cmd + '; echo $?'
                                    split_cmd = '\n'
                                with requests.Session() as s:
                                    saltapi = SaltAPI(session=s)
                                    if saltapi.get_token() is False:
                                        app_log.append('\n执行命令2后台出错_error(0)，请联系管理员')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                                    else:
                                        response_data = saltapi.cmd_run_api(tgt=minion_id, arg=[execute_cmd])
                                        # 当调用api失败的时候会返回false
                                        if response_data is False:
                                            app_log.append('\n执行命令2后台出错_error(1)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            try:
                                                response_data = response_data['return'][0][minion_id].rsplit(split_cmd, 1)
                                                # 发现有的命令没有输出那么最终只会有成功失败的0、1返回这时候列表长度就=1
                                                if len(response_data) == 1:
                                                    if response_data[0] == '0':
                                                        app_log.append('\n' + '执行命令2成功<-\n')
                                                    else:
                                                        app_log.append('\n' + '执行命令2失败:' + response_data[0])
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                else:
                                                    if response_data[1] == '0':
                                                        app_log.append('\n' + '执行命令2成功<-\n')
                                                    else:
                                                        app_log.append('\n' + '执行命令2失败:' + response_data[0])
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                            except Exception as e:
                                                app_log.append('\n' + '执行命令2失败_error(2):' + str(response_data))
                                                result['result'] = app_log
                                                return JsonResponse(result)

                        app_log.append(('-' * 10 + ('Minion_ID:%s发布完成 时间戳%s' % (minion_id, time.strftime('%X')))+'-'*10).center(88) + '\n\n\n\n\n\n')
                    result['status'] = True
                    result['result'] = app_log
                    return JsonResponse(result)
                except Exception as e:
                    logger.error(str(e))
                    result['result'] = app_log
                    result['result'].append('\n出错了：'+str(e))
                    return JsonResponse(result)
                finally:
                    if result['status']:
                        release_result = '发布成功'
                        if 'SVN更新' in operation_content:
                            AppRelease.objects.filter(app_name=app_name).update(app_svn_version_success=app_svn_version_success)
                    else:
                        release_result = '发布失败'
                    username = request.user.username
                    AppReleaseLog.objects.create(app_name=app_name, log_content=app_log, release_result=release_result, username=username)
                    AppRelease.objects.filter(app_name=app_name).update(update_time=time.strftime('%Y年%m月%d日 %X'))
            elif request.POST.get('app_tag_key') == 'app_backup':
                app_name = request.POST.get('app_name')
                try:
                    app_backup_path = AppRelease.objects.get(app_name=app_name).app_backup_path
                    app_path = AppRelease.objects.get(app_name=app_name).app_path
                    minion_id = AppRelease.objects.get(app_name=app_name).minion_id
                    minion_id_list = minion_id.split(',')
                    app_log.append(('-' * 20 + ('Minion_ID:%s 备份任务启动' % minion_id) + '-' * 20).center(88) + '\n')
                    app_log.append('\n\n开始备份->\n')
                    with requests.Session() as s:
                        saltapi = SaltAPI(session=s)
                        if saltapi.get_token() is False:
                            app_log.append('\n备份应用后台出错_error(0)，请联系管理员')
                            result['result'] = app_log
                            return JsonResponse(result)
                        else:
                            # 判断客户端应用目录是否存在，存在也要删除
                            for minion in minion_id_list:
                                response_data = saltapi.file_directory_exists_api(tgt=minion, arg=[app_path])
                                # 当调用api失败的时候会返回false
                                if response_data is False:
                                    app_log.append('\n备份应用后台出错_error(1)，请联系管理员')
                                    result['result'] = app_log
                                    return JsonResponse(result)
                                else:
                                    response_data = response_data['return'][0][minion]
                                    if response_data is True:
                                        response_data = saltapi.state_api(tgt=minion_id, arg=["copy_dir", "pillar={'source_path':'%s','name_path':'%s'}" % (app_path, app_backup_path), "queue=True"])
                                        if response_data is False:
                                            app_log.append('\n备份应用后台出错_error(2)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            format_result = format_state(response_data)
                                            # 这个是对格式化输出的一个判断，类型str说明格式化出错了呵呵，一般在minion一个sls未执行完成又执行会出现
                                            if type(format_result) == str:
                                                # 如果minion客户端停了会返回：{'return': [{'192.168.100.170': False}]}
                                                app_log.append(format_result)
                                                app_log.append('\n' + '备份应用失败！！')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    failed_result = re.search(r'Failed:     (\d+)', format_result[0]).group(1)
                                                    if int(failed_result) != 0:
                                                        app_log.extend(format_result)
                                                        app_log.append('\n' + '备份应用失败！！')
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append('\n' + '备份应用代码出错：'+str(e))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                            app_log.extend(format_result)
                                            app_log.append('\n备份应用完成<-\n\n')
                                    else:
                                        app_log.append('\n备份应用失败,应用目录不存在，无法备份，请确认是否发布过')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                            app_log.append(
                                ('-' * 20 + ('Minion_ID:%s备份任务结束' % minion_id) + '-' * 20).center(88) + '\n\n\n\n\n\n')
                    result['status'] = True
                    result['result'] = app_log
                    return JsonResponse(result)
                except Exception as e:
                    logger.error(str(e))
                    result['result'] = app_log
                    result['result'].append('\n出错了：' + str(e))
                    return JsonResponse(result)
                finally:
                    if result['status']:
                        release_result = '备份成功'
                    else:
                        release_result = '备份失败'
                    username = request.user.username
                    AppReleaseLog.objects.create(app_name=app_name, log_content=app_log, release_result=release_result, username=username)
                    AppRelease.objects.filter(app_name=app_name).update(update_time=time.strftime('%Y年%m月%d日 %X'))
            elif request.POST.get('app_tag_key') == 'app_restore':
                app_name = request.POST.get('app_name')
                try:
                    app_backup_path = AppRelease.objects.get(app_name=app_name).app_backup_path
                    app_path = AppRelease.objects.get(app_name=app_name).app_path
                    minion_id = AppRelease.objects.get(app_name=app_name).minion_id
                    minion_id_list = minion_id.split(',')
                    app_log.append(('-' * 20 + ('Minion_ID:%s 还原任务启动' % minion_id) + '-' * 20).center(88) + '\n')
                    app_log.append('\n\n开始还原->\n')
                    with requests.Session() as s:
                        saltapi = SaltAPI(session=s)
                        if saltapi.get_token() is False:
                            app_log.append('\n还原应用后台出错_error(0)，请联系管理员')
                            result['result'] = app_log
                            return JsonResponse(result)
                        else:
                            # 判断客户端应用目录是否存在，存在也要删除
                            for minion in minion_id_list:
                                response_data = saltapi.file_directory_exists_api(tgt=minion, arg=[app_backup_path])
                                # 当调用api失败的时候会返回false
                                if response_data is False:
                                    app_log.append('\n还原应用后台出错_error(1)，请联系管理员')
                                    result['result'] = app_log
                                    return JsonResponse(result)
                                else:
                                    response_data = response_data['return'][0][minion]
                                    if response_data is True:
                                        response_data = saltapi.state_api(tgt=minion_id, arg=["copy_dir", "pillar={'source_path':'%s','name_path':'%s'}" % (app_backup_path, app_path), "queue=True"])
                                        if response_data is False:
                                            app_log.append('\n还原应用后台出错_error(2)，请联系管理员')
                                            result['result'] = app_log
                                            return JsonResponse(result)
                                        else:
                                            format_result = format_state(response_data)
                                            # 这个是对格式化输出的一个判断，类型str说明格式化出错了呵呵，一般在minion一个sls未执行完成又执行会出现
                                            if type(format_result) == str:
                                                # 如果minion客户端停了会返回：{'return': [{'192.168.100.170': False}]}
                                                app_log.append(format_result)
                                                app_log.append('\n' + '还原应用失败！！')
                                                result['result'] = app_log
                                                return JsonResponse(result)
                                            else:
                                                try:
                                                    failed_result = re.search(r'Failed:     (\d+)', format_result[0]).group(1)
                                                    if int(failed_result) != 0:
                                                        app_log.extend(format_result)
                                                        app_log.append('\n' + '还原应用失败！！')
                                                        result['result'] = app_log
                                                        return JsonResponse(result)
                                                except Exception as e:
                                                    app_log.append('\n' + '还原应用代码出错：'+str(e))
                                                    result['result'] = app_log
                                                    return JsonResponse(result)
                                            app_log.extend(format_result)
                                            app_log.append('\n还原应用完成<-\n\n')
                                    else:
                                        app_log.append('\n还原应用失败,应用备份目录不存在，无法还原，请确认是否备份过')
                                        result['result'] = app_log
                                        return JsonResponse(result)
                            app_log.append(
                                ('-' * 20 + ('Minion_ID:%s还原任务结束' % minion_id) + '-' * 20).center(88) + '\n\n\n\n\n\n')
                    result['status'] = True
                    result['result'] = app_log
                    return JsonResponse(result)
                except Exception as e:
                    logger.error(str(e))
                    result['result'] = app_log
                    result['result'].append('\n出错了：' + str(e))
                    return JsonResponse(result)
                finally:
                    if result['status']:
                        release_result = '还原成功'
                    else:
                        release_result = '还原失败'
                    username = request.user.username
                    AppReleaseLog.objects.create(app_name=app_name, log_content=app_log, release_result=release_result, username=username)
                    AppRelease.objects.filter(app_name=app_name).update(update_time=time.strftime('%Y年%m月%d日 %X'))
            else:
                result['result'] = '应用发布页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('应用发布页ajax提交处理有问题', e)
        result['result'] = '应用发布页ajax提交处理有问题'
        return JsonResponse(result)


# 发布系统 应用发布组 主
def app_group(request):
    try:
        if request.method == 'GET':
            # 默认如果没有get到的话值为None，这里我需要为空''，所以下面修改默认值为''
            search_field = request.GET.get('search_field', '')
            search_content = request.GET.get('search_content', '')
            if request.user.is_superuser:
                if search_content is '':
                    app_group_data = AppGroup.objects.all().order_by('id')
                    data_list = getPage(request, app_group_data, 12)
                else:
                    if search_field == 'search_app_group_name':
                        app_data = AppGroup.objects.filter(
                            app_group_name__icontains=search_content).order_by(
                            'id')
                        data_list = getPage(request, app_data, 12)
                    elif search_field == 'search_app_group_members':
                        app_data = AppGroup.objects.filter(
                            app_group_members__icontains=search_content).order_by(
                            'id')
                        data_list = getPage(request, app_data, 12)
                    else:
                        data_list = ""
                return render(request, 'app_group.html',
                              {'data_list': data_list, 'search_field': search_field, 'search_content': search_content})
            else:
                username = request.user.username
                try:
                    app_auth_app_group_data = AppAuth.objects.get(username=username).app_group_perms.split(',')
                except Exception as e:
                    app_auth_app_group_data = ''
                if search_content is '':
                    app_group_data = AppGroup.objects.filter(app_group_name__in=app_auth_app_group_data).order_by('id')
                    data_list = getPage(request, app_group_data, 12)
                else:
                    if search_field == 'search_app_group_name':
                        app_data = AppGroup.objects.filter(app_group_name__in=app_auth_app_group_data).filter(
                            app_group_name__icontains=search_content).order_by(
                            'id')
                        data_list = getPage(request, app_data, 12)
                    elif search_field == 'search_app_group_members':
                        app_data = AppGroup.objects.filter(app_group_name__in=app_auth_app_group_data).filter(
                            app_group_members__icontains=search_content).order_by(
                            'id')
                        data_list = getPage(request, app_data, 12)
                    else:
                        data_list = ""
                return render(request, 'app_group.html',
                              {'data_list': data_list, 'search_field': search_field, 'search_content': search_content})
    except Exception as e:
        logger.error('应用发布组页面有问题', e)
        return render(request, 'app_group.html')


# 发布系统 应用发布组 成员管理页 主
def app_group_members_manage(request):
    try:
        if request.method == 'GET':
            app_group_name = request.GET.get('app_group_name')
            app_group_members_data = AppGroup.objects.get(app_group_name=app_group_name).app_group_members
            app_data = []
            if app_group_members_data:
                for app_name in app_group_members_data.split(','):
                    app_data.extend(AppRelease.objects.filter(app_name=app_name))
            # 下面这个app_data_list的作用是复制一份，因为刚开始下面做判断的时候直接用for  app in app_data然后内
            # 部又用了app_data.remove这样会导致app_data变了连锁导致开头for出现奇怪的现象逻辑上面自己搞错了
            app_data_list = app_data[:]
            search_field = request.GET.get('search_field', '')
            search_content = request.GET.get('search_content', '')
            if search_content is '':
                data_list = getPage(request, app_data_list, 12)
            else:
                if search_field == 'search_app_name':
                    for app in app_data:
                        if search_content in app.app_name:
                            pass
                        else:
                            app_data_list.remove(app)
                    data_list = getPage(request, app_data_list, 12)
                elif search_field == 'search_minion_id':
                    for app in app_data:
                        if search_content in app.minion_id:
                            pass
                        else:
                            app_data_list.remove(app)
                    data_list = getPage(request, app_data_list, 12)
                else:
                    data_list = ""
            return render(request, 'app_group_members_manage.html',
                          {'data_list': data_list, 'search_field': search_field,
                           'search_content': search_content,
                           'app_group_name': app_group_name,
                           'app_group_members_data': app_group_members_data})
    except Exception as e:
        logger.error('成员管理页面有问题', e)
        return render(request, 'app_group_members_manage.html')


# 发布系统 应用发布组 ajax提交处理
def app_group_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.POST.get('app_group_tag_key') == 'app_group_add' and request.user.is_superuser:
                obj = AppGroupAddForm(request.POST)
                if obj.is_valid():
                    AppGroup.objects.create(app_group_name=obj.cleaned_data["app_group_name"], description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.GET.get('app_group_tag_key') == 'modal_search_app_name':

                app_name = request.GET.get('app_name')
                app_name_list = AppRelease.objects.filter(app_name__icontains=app_name).order_by(
                    'create_time').values_list('app_name', flat=True)
                result['result'] = list(app_name_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.POST.get('app_group_tag_key') == 'app_group_update' and request.user.is_superuser:
                obj = AppGroupUpdateForm(request.POST)
                if obj.is_valid():
                    AppGroup.objects.filter(app_group_name=obj.cleaned_data["app_group_name"]).update(
                        description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('app_group_tag_key') == 'app_group_delete' and request.user.is_superuser:
                app_group_name = request.POST.get('app_group_name')
                try:
                    AppGroup.objects.filter(app_group_name=app_group_name).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('app_group_tag_key') == 'app_group_member_add' and request.user.is_superuser:
                obj = AppGroupUpdateForm(request.POST)
                if obj.is_valid():
                    AppGroup.objects.filter(app_group_name=obj.cleaned_data["app_group_name"]).update(
                        app_group_members=obj.cleaned_data["app_group_members"])
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                    return JsonResponse(result)
                result['result'] = '成功'
                result['status'] = True
                return JsonResponse(result)
            elif request.POST.get('app_group_tag_key') == 'app_group_member_delete' and request.user.is_superuser:
                app_name = request.POST.get('app_name')
                app_group_name = request.POST.get('app_group_name')
                try:
                    app_group_members = AppGroup.objects.get(app_group_name=app_group_name).app_group_members
                    app_group_members_list = app_group_members.split(',')
                    app_group_members_list.remove(app_name)
                    app_group_members = ','.join(app_group_members_list)
                    AppGroup.objects.filter(app_group_name=app_group_name).update(app_group_members=app_group_members)
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            else:
                result['result'] = '应用发布组ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('应用发布组ajax提交处理有问题', e)
        result['result'] = '应用发布组ajax提交处理有问题'
        return JsonResponse(result)


# 发布系统 应用授权 主
def app_auth(request):
    try:
        if request.method == 'GET':
            # 默认如果没有get到的话值为None，这里我需要为空''，所以下面修改默认值为''
            search_field = request.GET.get('search_field', '')
            search_content = request.GET.get('search_content', '')
            username_list = list(AppAuth.objects.values('my_user_id', 'username'))
            if search_content is '':
                app_auth_data = AppAuth.objects.all().order_by('my_user_id')
                data_list = getPage(request, app_auth_data, 12)
            else:
                if search_field == 'search_myuser_username':
                    app_auth_data = AppAuth.objects.filter(username__icontains=search_content).order_by('my_user_id')
                    data_list = getPage(request, app_auth_data, 12)
                elif search_field == 'search_app_name':
                    app_auth_data = AppAuth.objects.filter(app_perms__icontains=search_content).order_by('my_user_id')
                    data_list = getPage(request, app_auth_data, 12)
                elif search_field == 'search_app_group_name':
                    app_auth_data = AppAuth.objects.filter(app_group_perms__icontains=search_content).order_by('my_user_id')
                    data_list = getPage(request, app_auth_data, 12)
                else:
                    data_list = ""
            return render(request, 'app_auth.html',
                          {'data_list': data_list, 'search_field': search_field, 'search_content': search_content,
                           'username_list': username_list})
    except Exception as e:
        logger.error('应用授权页面有问题', e)
        return render(request, 'app_auth.html')


# 发布系统 应用授权 应用权限管理页
def app_auth_app_manage(request):
    try:
        if request.method == 'GET':
            my_user_id = request.GET.get('my_user_id')
            username = request.GET.get('username')
            app_perms_data = AppAuth.objects.get(username=username).app_perms
            if app_perms_data:
                app_data_list = AppRelease.objects.filter(app_name__in=app_perms_data.split(',')).order_by('app_name')
            else:
                app_data_list = []
            # 默认如果没有get到的话值为None，这里我需要为空''，所以下面修改默认值为''
            search_field = request.GET.get('search_field', '')
            search_content = request.GET.get('search_content', '')
            if search_content is '':
                data_list = getPage(request, app_data_list, 12)
            else:
                if search_field == 'search_app_name':
                    app_data_list = app_data_list.filter(app_name__icontains=search_content)
                    data_list = getPage(request, app_data_list, 12)
                elif search_field == 'search_minion_id':
                    app_data_list = app_data_list.filter(minion_id__icontains=search_content)
                    data_list = getPage(request, app_data_list, 12)
                else:
                    data_list = ""
            return render(request, 'app_auth_app_manage.html', {'data_list': data_list, 'search_field': search_field,
                                                                'search_content': search_content, 'username': username,
                                                                'app_perms_data': app_perms_data, 'my_user_id': my_user_id})
    except Exception as e:
        logger.error('应用授权应用权限管理页面有问题', e)
        return render(request, 'app_auth_app_manage.html')


# 发布系统 应用授权 应用组权限管理页
def app_auth_app_group_manage(request):
    try:
        if request.method == 'GET':
            my_user_id = request.GET.get('my_user_id')
            username = request.GET.get('username')
            app_group_perms_data = AppAuth.objects.get(username=username).app_group_perms
            if app_group_perms_data:
                app_data_list = AppGroup.objects.filter(app_group_name__in=app_group_perms_data.split(',')).order_by('id')
            else:
                app_data_list = []
            # 默认如果没有get到的话值为None，这里我需要为空''，所以下面修改默认值为''
            search_field = request.GET.get('search_field', '')
            search_content = request.GET.get('search_content', '')
            if search_content is '':
                data_list = getPage(request, app_data_list, 12)
            else:
                if search_field == 'search_app_group_name':
                    app_data_list = app_data_list.filter(app_group_name__icontains=search_content).order_by('id')
                    data_list = getPage(request, app_data_list, 12)
                elif search_field == 'search_app_group_members':
                    app_data_list = app_data_list.filter(app_group_members__icontains=search_content).order_by('id')
                    data_list = getPage(request, app_data_list, 12)
                else:
                    data_list = ""
            return render(request, 'app_auth_app_group_manage.html', {'data_list': data_list,
                                                                      'search_field': search_field,
                                                                      'search_content': search_content,
                                                                      'username': username,
                                                                      'app_group_perms_data': app_group_perms_data,
                                                                      'my_user_id': my_user_id})
    except Exception as e:
        logger.error('应用授权应用组权限管理页面有问题', e)
        return render(request, 'app_auth_app_group_manage.html')


# 发布系统 应用授权 ajax提交处理
def app_auth_ajax(request):
    result = {'result': None, 'status': False}
    try:
        if request.is_ajax():
            if request.POST.get('app_auth_tag_key') == 'app_auth_add':
                username_list = request.POST.get('username_list').split(',')
                for id_and_username in username_list:
                    id_and_username = id_and_username.split(' ')
                    data = {'my_user_id': id_and_username[0], 'username': id_and_username[1]}
                    obj = AppAuthCreateForm(data)
                    if obj.is_valid():
                        AppAuth.objects.create(my_user_id=obj.cleaned_data["my_user_id"], username=obj.cleaned_data["username"])
                        result['result'] = '成功'
                        result['status'] = True
                    else:
                        error_str = obj.errors.as_json()
                        result['result'] = json.loads(error_str)
                        return JsonResponse(result)
                return JsonResponse(result)
            elif request.GET.get('app_auth_tag_key') == 'modal_search_username':
                username = request.GET.get('username')
                username_list = MyUser.objects.filter(username__icontains=username).order_by('id').values('id', 'username')
                result['result'] = list(username_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.GET.get('app_auth_tag_key') == 'modal_search_app_name':
                app_name = request.GET.get('app_name')
                app_name_list = AppRelease.objects.filter(app_name__icontains=app_name).order_by(
                    'create_time').values_list('app_name', flat=True)
                result['result'] = list(app_name_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.GET.get('app_auth_tag_key') == 'modal_search_app_group':
                app_group_name = request.GET.get('app_group_name')
                app_group_list = AppGroup.objects.filter(app_group_name__icontains=app_group_name).order_by(
                    'id').values_list('app_group_name', flat=True)
                result['result'] = list(app_group_list)
                result['status'] = True
                return JsonResponse(result)
            elif request.POST.get('app_auth_tag_key') == 'app_auth_update':
                obj = AppAuthUpdateForm(request.POST)
                if obj.is_valid():
                    AppAuth.objects.filter(my_user_id=obj.cleaned_data['my_user_id'],
                                           username=obj.cleaned_data["username"]).update(
                        app_perms=obj.cleaned_data["app_perms"], app_group_perms=obj.cleaned_data["app_group_perms"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('app_auth_tag_key') == 'app_auth_description_update':
                obj = AppAuthUpdateForm(request.POST)
                if obj.is_valid():
                    AppAuth.objects.filter(my_user_id=obj.cleaned_data['my_user_id'],
                                           username=obj.cleaned_data["username"]).update(description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('app_auth_tag_key') == 'app_auth_app_update':
                obj = AppAuthUpdateForm(request.POST)
                if obj.is_valid():
                    AppAuth.objects.filter(my_user_id=obj.cleaned_data['my_user_id'],
                                           username=obj.cleaned_data["username"]).update(
                        app_perms=obj.cleaned_data["app_perms"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('app_auth_tag_key') == 'app_auth_app_delete':
                username = request.POST.get('username')
                my_user_id = request.POST.get('my_user_id')
                app_name = request.POST.get('app_name')
                try:
                    app_perms = AppAuth.objects.get(my_user_id=my_user_id, username=username).app_perms
                    app_perms_list = app_perms.split(',')
                    # 为了结合单个移除和批量移除，对传过来的app_name做列表化因为批量删除就是逗号隔开的字符串，然后移除操作
                    for data in app_name.split(','):
                        app_perms_list.remove(data) if data in app_perms_list else app_perms_list
                    logger.error(app_perms_list)
                    app_perms = ','.join(app_perms_list)
                    AppAuth.objects.filter(my_user_id=my_user_id, username=username).update(app_perms=app_perms)
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('app_auth_tag_key') == 'app_auth_app_group_update':
                obj = AppAuthUpdateForm(request.POST)
                if obj.is_valid():
                    AppAuth.objects.filter(my_user_id=obj.cleaned_data['my_user_id'],
                                           username=obj.cleaned_data["username"]).update(
                        app_group_perms=obj.cleaned_data["app_group_perms"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('app_auth_tag_key') == 'app_auth_app_group_delete':
                username = request.POST.get('username')
                my_user_id = request.POST.get('my_user_id')
                app_group_name = request.POST.get('app_group_name')
                try:
                    app_group_perms = AppAuth.objects.get(my_user_id=my_user_id, username=username).app_group_perms
                    app_group_perms_list = app_group_perms.split(',')
                    # 为了结合单个移除和批量移除，对传过来的app_name做列表化因为批量删除就是逗号隔开的字符串，然后移除操作
                    for data in app_group_name.split(','):
                        app_group_perms_list.remove(data) if data in app_group_perms_list else app_group_perms_list
                    app_group_perms = ','.join(app_group_perms_list)
                    AppAuth.objects.filter(my_user_id=my_user_id, username=username).update(app_group_perms=app_group_perms)
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('app_auth_tag_key') == 'app_auth_delete':
                my_user_id = request.POST.get('my_user_id')
                try:
                    AppAuth.objects.filter(my_user_id=my_user_id).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            else:
                result['result'] = '应用发布组ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('应用发布组ajax提交处理有问题', e)
        result['result'] = '应用发布组ajax提交处理有问题'
        return JsonResponse(result)