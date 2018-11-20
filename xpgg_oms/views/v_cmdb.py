#!/usr/bin/env python3
# -.- coding:utf-8 -.-

from django.shortcuts import render, HttpResponse
import json
from django.http import JsonResponse, FileResponse  # 1.7以后版本json数据返回方法
import os
from xpgg_oms.forms import *
from xpgg_oms.models import *
from django.conf import settings
import openpyxl  # 操作excel读写
from io import BytesIO
from django.utils.encoding import escape_uri_path  # 下载文件中文名时使用
from django.utils import timezone  # 调用django的时间参数timezone.now()
from .v_general import getPage


import logging
# Create your views here.
logger = logging.getLogger('xpgg_oms.views')

# 主机资源列表
def server_list(request):
    try:
        # 这里主要是担心通过ajax提交了get请求，其实多此一举因为这个页面没有ajax的get请求哈哈
        if request.is_ajax() is not True:
            if request.method == 'GET':
                search_field = request.GET.get('search_field', '')
                search_content = request.GET.get('search_content', '')
                if search_content is '':
                    server_data = ServerList.objects.all().order_by('create_date')
                    data_list = getPage(request, server_data, 9)
                else:
                    if search_field == 'search_server_name':
                        server_data = ServerList.objects.filter(server_name__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                    elif search_field == 'search_ip':
                        server_data = ServerList.objects.filter(ip__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                    elif search_field == 'search_sys':
                        server_data = ServerList.objects.filter(sys__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                    else:
                        server_data = ServerList.objects.filter(server_type__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, server_data, 9)
                return render(request, 'cmdb/server_list.html',
                              {'data_list': data_list, 'search_field': search_field, 'search_content': search_content})

    except Exception as e:
        logger.error('主机列表管理页面有问题', e)
        return render(request, 'cmdb/server_list.html')


# 主机资源列表ajax
def server_list_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.GET.get('server_list_tag_key') == 'modal_search_minion_id':
                minion_id = request.GET.get('minion_id')
                sys = request.GET.get('sys_type')
                minion_id_list = MinionList.objects.filter(minion_id__icontains=minion_id, sys=sys).order_by(
                    'create_date').values_list('minion_id', flat=True)
                result['result'] = list(minion_id_list)
                result['status'] = True
                # 返回字典之外的需要把参数safe改成false如：JsonResponse([1, 2, 3], safe=False)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'add_server_list':
                obj = ServerListAddForm(request.POST)
                if obj.is_valid():
                    ServerList.objects.create(server_name=obj.cleaned_data["server_name"],
                                              server_type=obj.cleaned_data["server_type"],
                                              localhost=obj.cleaned_data["localhost"],
                                              ip=obj.cleaned_data["ip"],
                                              system_issue=obj.cleaned_data["system_issue"],
                                              sn=obj.cleaned_data["sn"],
                                              cpu_num=obj.cleaned_data["cpu_num"],
                                              cpu_model=obj.cleaned_data["cpu_model"],
                                              sys_type=obj.cleaned_data["sys_type"],
                                              kernel=obj.cleaned_data["kernel"],
                                              product_name=obj.cleaned_data["product_name"],
                                              ipv4_address=obj.cleaned_data["ipv4_address"],
                                              mac_address=obj.cleaned_data["mac_address"],
                                              mem_total=obj.cleaned_data["mem_total"],
                                              mem_explain=obj.cleaned_data["mem_explain"],
                                              disk_total=obj.cleaned_data["disk_total"],
                                              disk_explain=obj.cleaned_data["disk_explain"],
                                              minion_id=obj.cleaned_data["minion_id"],
                                              idc_name=obj.cleaned_data["idc_name"],
                                              idc_num=obj.cleaned_data["idc_num"],
                                              login_ip=obj.cleaned_data["login_ip"],
                                              login_port=obj.cleaned_data["login_port"],
                                              login_user=obj.cleaned_data["login_user"],
                                              login_password=obj.cleaned_data["login_password"],
                                              description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'update_server_list':
                obj = ServerListUpdateForm(request.POST)
                if obj.is_valid():
                    ServerList.objects.filter(server_name=obj.cleaned_data["server_name"]).update(
                        server_type=obj.cleaned_data["server_type"],
                        localhost=obj.cleaned_data["localhost"],
                        ip=obj.cleaned_data["ip"],
                        system_issue=obj.cleaned_data["system_issue"],
                        sn=obj.cleaned_data["sn"],
                        cpu_num=obj.cleaned_data["cpu_num"],
                        cpu_model=obj.cleaned_data["cpu_model"],
                        sys_type=obj.cleaned_data["sys_type"],
                        kernel=obj.cleaned_data["kernel"],
                        product_name=obj.cleaned_data["product_name"],
                        ipv4_address=obj.cleaned_data["ipv4_address"],
                        mac_address=obj.cleaned_data["mac_address"],
                        mem_total=obj.cleaned_data["mem_total"],
                        mem_explain=obj.cleaned_data["mem_explain"],
                        disk_total=obj.cleaned_data["disk_total"],
                        disk_explain=obj.cleaned_data["disk_explain"],
                        minion_id=obj.cleaned_data["minion_id"],
                        idc_name=obj.cleaned_data["idc_name"],
                        idc_num=obj.cleaned_data["idc_num"],
                        login_ip=obj.cleaned_data["login_ip"],
                        login_port=obj.cleaned_data["login_port"],
                        login_user=obj.cleaned_data["login_user"],
                        login_password=obj.cleaned_data["login_password"],
                        update_time=timezone.now(),
                        description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'delete_server_list':
                server_name = request.POST.get('server_name')
                try:
                    ServerList.objects.get(server_name=server_name).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('server_list_tag_key') == 'import_server_list':
                # 获取的数据类型如果有文件，则要像下面这么写form才能正确获取到文件
                obj = ServerListImportForm(request.POST, request.FILES)
                if obj.is_valid():
                    # request.FILES['file'].read()是从内存中直接读取文件内容
                    # openpyxl的filename除了能打开某个路径的文件，还可以接受BytesIO读取的二进制数据所以如下即可获取xlsx内容
                    wb = openpyxl.load_workbook(filename=BytesIO(request.FILES['file'].read()))
                    table = wb[wb.sheetnames[0]]
                    server_list_cover_select = request.POST.get('server_list_cover_select')
                    # 结果需要存一些内容所以修改成列表
                    result['result'] = []
                    for row in table.iter_rows(min_row=2, max_col=25):
                        data = {'server_name': row[0].value, 'server_type': row[1].value, 'localhost': row[2].value,
                                'ip': row[3].value, 'system_issue': row[4].value, 'sn': row[5].value,
                                'cpu_num': row[6].value, 'cpu_model': row[7].value, 'sys_type': row[8].value,
                                'kernel': row[9].value, 'product_name': row[10].value, 'ipv4_address': row[11].value,
                                'mac_address': row[12].value, 'mem_total': row[13].value, 'mem_explain': row[14].value,
                                'disk_total': row[15].value, 'disk_explain': row[16].value, 'minion_id': row[17].value,
                                'idc_name': row[18].value, 'idc_num': row[17].value, 'login_ip': row[20].value,
                                'login_port': row[21].value, 'login_user': row[22].value,
                                'login_password': row[23].value, 'description': row[24].value}
                        obj = ServerListAddForm(data)
                        if obj.is_valid():
                            ServerList.objects.create(server_name=obj.cleaned_data["server_name"],
                                                      server_type=obj.cleaned_data["server_type"],
                                                      localhost=obj.cleaned_data["localhost"],
                                                      ip=obj.cleaned_data["ip"],
                                                      system_issue=obj.cleaned_data["system_issue"],
                                                      sn=obj.cleaned_data["sn"],
                                                      cpu_num=obj.cleaned_data["cpu_num"],
                                                      cpu_model=obj.cleaned_data["cpu_model"],
                                                      sys_type=obj.cleaned_data["sys_type"],
                                                      kernel=obj.cleaned_data["kernel"],
                                                      product_name=obj.cleaned_data["product_name"],
                                                      ipv4_address=obj.cleaned_data["ipv4_address"],
                                                      mac_address=obj.cleaned_data["mac_address"],
                                                      mem_total=obj.cleaned_data["mem_total"],
                                                      mem_explain=obj.cleaned_data["mem_explain"],
                                                      disk_total=obj.cleaned_data["disk_total"],
                                                      disk_explain=obj.cleaned_data["disk_explain"],
                                                      minion_id=obj.cleaned_data["minion_id"],
                                                      idc_name=obj.cleaned_data["idc_name"],
                                                      idc_num=obj.cleaned_data["idc_num"],
                                                      login_ip=obj.cleaned_data["login_ip"],
                                                      login_port=obj.cleaned_data["login_port"],
                                                      login_user=obj.cleaned_data["login_user"],
                                                      login_password=obj.cleaned_data["login_password"],
                                                      description=obj.cleaned_data["description"])
                        else:
                            error_str = json.loads(obj.errors.as_json())
                            if error_str.get('server_name') and server_list_cover_select == 'cover':
                                if error_str['server_name'][0]['message'] == '服务器名称已存在，请检查':
                                    obj = ServerListUpdateForm(data)
                                    if obj.is_valid():
                                        ServerList.objects.filter(server_name=obj.cleaned_data["server_name"]).update(
                                            server_type=obj.cleaned_data["server_type"],
                                            localhost=obj.cleaned_data["localhost"],
                                            ip=obj.cleaned_data["ip"],
                                            system_issue=obj.cleaned_data["system_issue"],
                                            sn=obj.cleaned_data["sn"],
                                            cpu_num=obj.cleaned_data["cpu_num"],
                                            cpu_model=obj.cleaned_data["cpu_model"],
                                            sys_type=obj.cleaned_data["sys_type"],
                                            kernel=obj.cleaned_data["kernel"],
                                            product_name=obj.cleaned_data["product_name"],
                                            ipv4_address=obj.cleaned_data["ipv4_address"],
                                            mac_address=obj.cleaned_data["mac_address"],
                                            mem_total=obj.cleaned_data["mem_total"],
                                            mem_explain=obj.cleaned_data["mem_explain"],
                                            disk_total=obj.cleaned_data["disk_total"],
                                            disk_explain=obj.cleaned_data["disk_explain"],
                                            minion_id=obj.cleaned_data["minion_id"],
                                            idc_name=obj.cleaned_data["idc_name"],
                                            idc_num=obj.cleaned_data["idc_num"],
                                            login_ip=obj.cleaned_data["login_ip"],
                                            login_port=obj.cleaned_data["login_port"],
                                            login_user=obj.cleaned_data["login_user"],
                                            login_password=obj.cleaned_data["login_password"],
                                            update_time=timezone.now(),
                                            description=obj.cleaned_data["description"])
                                    else:
                                        error_str = json.loads(obj.errors.as_json())
                                        result['result'].append(error_str)
                            else:
                                if error_str['server_name'][0]['message'] == '服务器名称已存在，请检查':
                                    result['result'].append('服务器名称:{} 已存在'.format(data['server_name']))
                                else:
                                    result['result'].append(error_str)
                    if result['result']:
                        pass
                    else:
                        result['result'] = '成功'
                        result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            else:
                result['result'] = '主机管理页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('主机管理页ajax提交处理有问题', e)
        result['result'] = '主机管理页ajax提交处理有问题'
        return JsonResponse(result)


# 主机资源管理模板下载
def server_list_template_down(request):
    try:
        # 基本上django下载文件就按这个模板来就可以，更复杂的参考下面的主机管理列表导出方法
        # FileResponse方法继承了StreamingHttpResponse并且封装了迭代方法，是django最好的大文件流传送方式了，用法直接按下面，很简单
        file = settings.STATICFILES_DIRS[0] + "/download_files/主机模板.xlsx"
        response = FileResponse(open(file, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        # 带中文的文件名需要如下用escape_uri_path和utf8才能识别
        response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path(os.path.basename(file)))
        return response
    except Exception as e:
        logger.error(str(e))
        return str(e)


# 主机资源列表理导出下载
def server_list_down(request):
    from openpyxl.utils import get_column_letter
    from openpyxl.writer.excel import save_virtual_workbook
    # 只写模式会加快速度，不过写法和普通读写有区别，按下面一行一行插入即可
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = "Sheet1"
    columns = ("服务器名称", "服务器类型(0是物理机，1是虚拟机)", "主机名", "IP地址", "系统版本", "SN", "CPU核数", "CPU型号",
               "系统类型", "内核", "品牌名称", "ipv4列表", "mac地址列表", "内存大小(M)", "内存说明", "磁盘大小(G)", "磁盘说明",
               "minion_id", "机房名称", "机柜号", "远程访问IP", "远程访问端口", "远程访问用户", "远程访问密码", "描述备注")

    ws.append(columns)
    queryset = ServerList.objects.all()
    for obj in queryset:
        row = (obj.server_name, int(obj.server_type), obj.localhost, obj.ip, obj.system_issue, obj.sn, obj.cpu_num,
               obj.cpu_model, obj.sys_type, obj.kernel, obj.product_name, obj.ipv4_address, obj.mac_address,
               obj.mem_total, obj.mem_explain, obj.disk_total, obj.disk_explain, obj.minion_id, obj.idc_name,
               obj.idc_num, obj.login_ip, obj.login_port, obj.login_user, obj.login_password, obj.update_time, obj.description)
        ws.append(row)
    # 用(save_virtual_workbook(wb)来保存到内存中供django调用,无法使用StreamingHttpResponse或者FileResponse在openpyxl官方例子就是这样的
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path('主机列表.xlsx'))
    # wb.save(response)
    return response


# 网络设备列表
def network_list(request):
    try:
        # 这里主要是担心通过ajax提交了get请求，其实多此一举因为这个页面没有ajax的get请求
        if request.is_ajax() is not True:
            if request.method == 'GET':
                search_field = request.GET.get('search_field', '')
                search_content = request.GET.get('search_content', '')
                if search_content is '':
                    device_data = NetworkList.objects.all().order_by('create_date')
                    data_list = getPage(request, device_data, 15)
                else:
                    if search_field == 'search_device_name':
                        device_data = NetworkList.objects.filter(device_name__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 15)
                    elif search_field == 'search_manage_ip':
                        device_data = NetworkList.objects.filter(manage_ip__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 15)
                    elif search_field == 'search_device_type':
                        device_data = NetworkList.objects.filter(device_type__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 15)
                    else:
                        device_data = NetworkList.objects.filter(server_type__icontains=search_content).order_by(
                            'create_date')
                        data_list = getPage(request, device_data, 15)
                data_form = NetworkListAddForm()
                return render(request, 'cmdb/network_list.html',
                              {'data_list': data_list, 'search_field': search_field,
                               'search_content': search_content, 'data_form': data_form})

    except Exception as e:
        logger.error('网络设备列表管理页面有问题', e)
        return render(request, 'cmdb/network_list.html')


# 网络设备列表ajax
def network_list_ajax(request):
    result = {'result': None, 'status': False}
    app_log = []
    try:
        if request.is_ajax():
            # 在ajax提交时候多一个字段作为标识，来区分多个ajax提交哈，厉害！
            if request.POST.get('network_list_tag_key') == 'add_network_list':
                obj = NetworkListAddForm(request.POST)
                if obj.is_valid():
                    NetworkList.objects.create(device_name=obj.cleaned_data["device_name"],
                                               device_type=obj.cleaned_data["device_type"],
                                               manage_ip=obj.cleaned_data["manage_ip"],
                                               product_name=obj.cleaned_data["product_name"],
                                               product_type=obj.cleaned_data["product_type"],
                                               sn=obj.cleaned_data["sn"],
                                               idc_name=obj.cleaned_data["idc_name"],
                                               idc_num=obj.cleaned_data["idc_num"],
                                               login_ip=obj.cleaned_data["login_ip"],
                                               login_port=obj.cleaned_data["login_port"],
                                               login_user=obj.cleaned_data["login_user"],
                                               login_password=obj.cleaned_data["login_password"],
                                               description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('network_list_tag_key') == 'update_network_list':
                obj = NetworkListUpdateForm(request.POST)
                if obj.is_valid():
                    NetworkList.objects.filter(device_name=obj.cleaned_data["device_name"]).update(
                        device_type=obj.cleaned_data["device_type"],
                        manage_ip=obj.cleaned_data["manage_ip"],
                        product_name=obj.cleaned_data["product_name"],
                        product_type=obj.cleaned_data["product_type"],
                        sn=obj.cleaned_data["sn"],
                        idc_name=obj.cleaned_data["idc_name"],
                        idc_num=obj.cleaned_data["idc_num"],
                        login_ip=obj.cleaned_data["login_ip"],
                        login_port=obj.cleaned_data["login_port"],
                        login_user=obj.cleaned_data["login_user"],
                        login_password=obj.cleaned_data["login_password"],
                        update_time=timezone.now(),
                        description=obj.cleaned_data["description"])
                    result['result'] = '成功'
                    result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            elif request.POST.get('network_list_tag_key') == 'delete_network_list':
                device_name = request.POST.get('device_name')
                try:
                    NetworkList.objects.get(device_name=device_name).delete()
                    result['result'] = '成功'
                    result['status'] = True
                except Exception as e:
                    result['result'] = str(e)
                return JsonResponse(result)
            elif request.POST.get('network_list_tag_key') == 'import_network_list':
                # 获取的数据类型如果有文件，则要像下面这么写form才能正确获取到文件
                obj = NetworkListImportForm(request.POST, request.FILES)
                if obj.is_valid():
                    # request.FILES['file'].read()是从内存中直接读取文件内容
                    # openpyxl的filename除了能打开某个路径的文件，还可以接受BytesIO读取的二进制数据所以如下即可获取xlsx内容
                    wb = openpyxl.load_workbook(filename=BytesIO(request.FILES['file'].read()))
                    table = wb[wb.sheetnames[0]]
                    network_list_cover_select = request.POST.get('network_list_cover_select')
                    # 结果需要存一些内容所以修改成列表
                    result['result'] = []
                    for row in table.iter_rows(min_row=2, max_col=25):
                        data = {'device_name': row[0].value, 'device_type': row[1].value, 'manage_ip': row[2].value,
                                'product_name': row[3].value, 'product_type': row[4].value, 'sn': row[5].value,
                                'idc_name': row[18].value, 'idc_num': row[17].value, 'login_ip': row[20].value,
                                'login_port': row[21].value, 'login_user': row[22].value,
                                'login_password': row[23].value, 'description': row[24].value}
                        obj = NetworkListAddForm(data)
                        if obj.is_valid():
                            NetworkList.objects.create(device_name=obj.cleaned_data["device_name"],
                                                       device_type=obj.cleaned_data["device_type"],
                                                       manage_ip=obj.cleaned_data["manage_ip"],
                                                       product_name=obj.cleaned_data["product_name"],
                                                       product_type=obj.cleaned_data["product_type"],
                                                       sn=obj.cleaned_data["sn"],
                                                       idc_name=obj.cleaned_data["idc_name"],
                                                       idc_num=obj.cleaned_data["idc_num"],
                                                       login_ip=obj.cleaned_data["login_ip"],
                                                       login_port=obj.cleaned_data["login_port"],
                                                       login_user=obj.cleaned_data["login_user"],
                                                       login_password=obj.cleaned_data["login_password"],
                                                       description=obj.cleaned_data["description"])
                        else:
                            error_str = json.loads(obj.errors.as_json())
                            if error_str.get('device_name') and network_list_cover_select == 'cover':
                                if error_str['device_name'][0]['message'] == '设备名称已存在，请检查':
                                    obj = NetworkListUpdateForm(data)
                                    if obj.is_valid():
                                        NetworkList.objects.filter(device_name=obj.cleaned_data["device_name"]).update(
                                            device_type=obj.cleaned_data["device_type"],
                                            manage_ip=obj.cleaned_data["manage_ip"],
                                            product_name=obj.cleaned_data["product_name"],
                                            product_type=obj.cleaned_data["product_type"],
                                            sn=obj.cleaned_data["sn"],
                                            idc_name=obj.cleaned_data["idc_name"],
                                            idc_num=obj.cleaned_data["idc_num"],
                                            login_ip=obj.cleaned_data["login_ip"],
                                            login_port=obj.cleaned_data["login_port"],
                                            login_user=obj.cleaned_data["login_user"],
                                            login_password=obj.cleaned_data["login_password"],
                                            update_time=timezone.now(),
                                            description=obj.cleaned_data["description"])
                                    else:
                                        error_str = json.loads(obj.errors.as_json())
                                        result['result'].append(error_str)
                            else:
                                if error_str['device_name'][0]['message'] == '设备名称已存在，请检查':
                                    result['result'].append('设备名称:{} 已存在'.format(data['device_name']))
                                else:
                                    result['result'].append(error_str)
                    if result['result']:
                        pass
                    else:
                        result['result'] = '成功'
                        result['status'] = True
                else:
                    error_str = obj.errors.as_json()
                    result['result'] = json.loads(error_str)
                return JsonResponse(result)
            else:
                result['result'] = '网络设备管理页ajax提交了错误的tag'
                return JsonResponse(result)
    except Exception as e:
        logger.error('网络设备管理页ajax提交处理有问题', e)
        result['result'] = '网络设备管理页ajax提交处理有问题'
        return JsonResponse(result)


# 网络设备列表模板下载
def network_list_template_down(request):
    try:
        # 基本上django下载文件就按这个模板来就可以，更复杂的参考下面的主机管理列表导出方法
        # FileResponse方法继承了StreamingHttpResponse并且封装了迭代方法，是django最好的大文件流传送方式了，用法直接按下面，很简单
        file = settings.STATICFILES_DIRS[0] + "/download_files/网络设备模板.xlsx"
        response = FileResponse(open(file, 'rb'))
        response['Content-Type'] = 'application/octet-stream'
        # 带中文的文件名需要如下用escape_uri_path和utf8才能识别
        response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path(os.path.basename(file)))
        return response
    except Exception as e:
        logger.error(str(e))
        return str(e)


# 网络设备列表理导出下载
def network_list_down(request):
    from openpyxl.utils import get_column_letter
    from openpyxl.writer.excel import save_virtual_workbook
    # 只写模式会加快速度，不过写法和普通读写有区别，按下面一行一行插入即可
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = "Sheet1"
    columns = ("设备名称", "设备类型", "管理IP", "设备厂家", "产品型号", "序列号",
               "机房名称", "机柜号", "远程管理IP", "远程管理端口", "远程管理用户", "远程管理密码", "描述备注")
    ws.append(columns)
    queryset = NetworkList.objects.all()
    for obj in queryset:
        row = (obj.device_name, obj.device_type, obj.manage_ip, obj.product_name, obj.product_type, obj.sn, obj.idc_name,
               obj.idc_num, obj.login_ip, obj.login_port, obj.login_user, obj.login_password, obj.description)
        ws.append(row)
    # 用(save_virtual_workbook(wb)来保存到内存中供django调用,无法使用StreamingHttpResponse或者FileResponse在openpyxl官方例子就是这样的
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment; filename*=utf-8''{0}".format(escape_uri_path('网络设备列表.xlsx'))
    # wb.save(response)
    return response